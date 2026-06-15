#!/usr/bin/env python3
"""Dashboard SolarZ — Análise Planos de Negócio 2026 (versão detalhada)."""

import openpyxl, os, json, unicodedata, base64, csv, io, time, re, requests
from datetime import datetime
from io import BytesIO

ANO = 2026

# ── Configuração via variáveis de ambiente (GitHub Actions) ou defaults locais ─
MASTER_SHEET_ID = os.environ.get('MASTER_SHEET_ID', '')
OUTPUT_DIR  = os.environ.get('OUTPUT_DIR',
              r"c:\Users\usuario\Desktop\Claude Code\Analise Planos De Negocio")
OUTPUT      = os.path.join(OUTPUT_DIR, 'index.html')
OUTPUT_TV   = os.path.join(OUTPUT_DIR, 'dashboard_tv.html')
DOWNLOAD_DELAY = float(os.environ.get('DOWNLOAD_DELAY', '1.5'))
MAX_RETRIES    = int(os.environ.get('MAX_RETRIES', '2'))

# ── Compatibilidade local: pasta de planilhas para modo offline ─────────────────
PASTA = os.environ.get('PASTA_LOCAL',
        r"c:\Users\usuario\Desktop\Claude Code\Analise Planos De Negocio\planilhas")

MESES_PT = {1:'Jan',2:'Fev',3:'Mar',4:'Abr',5:'Mai',6:'Jun',
            7:'Jul',8:'Ago',9:'Set',10:'Out',11:'Nov',12:'Dez'}

# ── HELPERS ──────────────────────────────────────────────────────────────────

def norm(t):
    if not t: return ''
    return unicodedata.normalize('NFKD', str(t).strip()).encode('ASCII','ignore').decode('ASCII').lower()

def parse_num(v):
    if v is None: return None
    if isinstance(v, bool): return None
    if isinstance(v, (int, float)):
        return None if v != v else float(v)
    s = str(v).strip().replace('\xa0','').replace(' ','').replace('%','')
    if ',' in s and '.' in s:
        s = s.replace('.','').replace(',','.')
    else:
        s = s.replace(',','.')
    try:
        f = float(s)
        return None if f != f else f
    except: return None

def parse_eff(v, meta=None, exec_val=None):
    n = parse_num(v)
    if n is not None:
        if isinstance(v, str) and '%' in v: return n / 100
        if n > 1.5: return n / 100
        return n
    if meta is not None and meta != 0 and exec_val is not None:
        return exec_val / meta
    if meta == 0 and exec_val is not None:
        return 0.0
    return None

def status_eff(eff):
    if eff is None: return 'nd'
    if eff >= 0.80: return 'verde'
    if eff >= 0.60: return 'amarelo'
    return 'vermelho'

def fmt_money(v):
    if v is None: return '—'
    if abs(v) >= 1_000_000: return f'R$ {v/1_000_000:.1f}M'
    if abs(v) >= 1_000: return f'R$ {v/1_000:.1f}K'
    return f'R$ {v:,.0f}'

def fmt_num(v):
    if v is None: return '—'
    return f'{int(v):,}'.replace(',','.')

def fmt_pct(v):
    if v is None: return '—'
    return f'{v*100:.1f}%'

# ── SHEET FINDER ─────────────────────────────────────────────────────────────

def find_sheet(wb):
    best_score, best_name = 0, wb.sheetnames[0]
    for name in wb.sheetnames:
        n = norm(name)
        score = 0
        if 'plano' in n: score += 1
        if any(x in n for x in ['negocio','negocios']): score += 1
        if '2026' in name or ' 26' in name or name.endswith('26'): score += 2
        if score > best_score:
            best_score, best_name = score, name
    return wb[best_name]

def get_period(ws):
    b2, c2 = ws['B2'].value, ws['C2'].value
    if isinstance(b2,(int,float)) and 1<=b2<=12:
        s=int(b2); e=int(c2) if isinstance(c2,(int,float)) and 1<=c2<=12 else s
        return s,e
    b1,c1 = ws['B1'].value, ws['C1'].value
    if isinstance(b1,(int,float)) and 1<=b1<=12:
        s=int(b1); e=int(c1) if isinstance(c1,(int,float)) and 1<=c1<=12 else s
        return s,e
    return 1,12

def map_monthly_cols(ws):
    row1=[cell.value for cell in ws[1]]
    row2=[cell.value for cell in ws[2]]
    def scan(check_row,date_row):
        result={}
        i=6
        while i<len(check_row):
            val=check_row[i]
            if isinstance(val,(int,float)) and 1<=int(val)<=12:
                mn=int(val); pc,ec=i-1,i
                if mn==12 and not result:
                    dv=date_row[pc] if pc<len(date_row) else None
                    if dv and hasattr(dv,'year') and dv.year<ANO:
                        i+=2; continue
                if mn not in result: result[mn]=(pc,ec)
            i+=2
        return result
    monthly=scan(row2,row1)
    if not monthly: monthly=scan(row1,row2)
    return monthly

# ── BLOCK PARSER ─────────────────────────────────────────────────────────────

ANALYZED={'acompanhamento','seguro','servicos'}

def canonical_block(name):
    n=norm(name)
    if 'acompanhamento' in n: return 'acompanhamento'
    if 'seguro' in n: return 'seguro'
    if 'servico' in n: return 'servicos'
    return None

def parse_file(source, empresa_id='', empresa_nome=''):
    """Aceita path local (str) ou conteúdo xlsx (bytes/BytesIO)."""
    if isinstance(source, (bytes, BytesIO)):
        source = BytesIO(source) if isinstance(source, bytes) else source
    else:
        partes = os.path.basename(source).replace('.xlsx','').split(' - ',1)
        empresa_id   = partes[0].strip()
        empresa_nome = partes[1].strip() if len(partes) > 1 else partes[0]

    empresa = f'{empresa_id} - {empresa_nome}' if empresa_nome else empresa_id
    result={'empresa':empresa,'id':empresa_id,'nome':empresa_nome or empresa_id,
            'periodo':None,'sheet':None,'blocos':{},'erros':[]}
    try:
        wb=openpyxl.load_workbook(source,data_only=True)
    except Exception as e:
        result['erros'].append(str(e)); return result

    ws=find_sheet(wb)
    result['sheet']=ws.title
    sm,em=get_period(ws)
    result['periodo']={'inicio':sm,'fim':em}
    monthly_map=map_monthly_cols(ws)
    max_col=ws.max_column
    blocos={}; cur_block=None; cur_parent=None; consec_empty=0

    for row in ws.iter_rows(min_row=2,values_only=True):
        row=list(row)
        while len(row)<max(max_col,6): row.append(None)
        a,b,c,d,e=row[0],row[1],row[2],row[3],row[4]

        is_empty=all(v is None or str(v).strip()=='' for v in (a,b,c,d,e))
        if is_empty:
            consec_empty+=1
            if consec_empty>=2: break
            continue
        else: consec_empty=0

        b_str=norm(str(b)) if b else ''
        if b_str in ('plano','plano anual') and a:
            canonical=canonical_block(str(a))
            cur_block=canonical; cur_parent=None
            if canonical and canonical not in blocos:
                blocos[canonical]={'nome':str(a).strip(),'indicadores':[]}
            continue

        if cur_block not in ANALYZED or cur_block not in blocos: continue
        if not a or str(a).strip()=='': continue

        a_str=str(a)
        is_sub=len(a_str)-len(a_str.lstrip())>=3
        ind_name=a_str.strip()
        meta=parse_num(c); exec_val=parse_num(d)
        eff=parse_eff(e,meta,exec_val)
        mensal={}
        for m,(pi,ei) in monthly_map.items():
            pv=parse_num(row[pi] if pi<len(row) else None)
            ev=parse_num(row[ei] if ei<len(row) else None)
            if pv is not None or ev is not None:
                mensal[m]={'p':pv,'e':ev}
        ind={'n':ind_name,'m':meta,'e':exec_val,'eff':eff,
             's':status_eff(eff),'sub':is_sub,
             'parent':cur_parent if is_sub else None,'mensal':mensal}
        if not is_sub: cur_parent=ind_name
        blocos[cur_block]['indicadores'].append(ind)

    # Pós-processamento: MRR Acumulado usa último mês com exec preenchido.
    # Coluna D em algumas planilhas contém soma de todos os meses (ERRADO).
    # O valor correto é sempre o executado do último mês registrado.
    for bloco in blocos.values():
        for i, ind in enumerate(bloco['indicadores']):
            if 'mrr acumulado' in norm(ind['n']) and not ind['sub']:
                mensal = ind.get('mensal', {})
                if mensal:
                    ultimo_exec = None
                    for m in sorted(mensal.keys(), reverse=True):
                        ev = mensal[m].get('e')
                        if ev is not None:
                            ultimo_exec = ev
                            break
                    if ultimo_exec is not None and ultimo_exec != ind['e']:
                        new_ind = dict(ind)
                        new_ind['e'] = ultimo_exec
                        meta = ind['m']
                        if meta and meta != 0:
                            new_ind['eff'] = ultimo_exec / meta
                        elif meta == 0:
                            new_ind['eff'] = 0.0
                        else:
                            new_ind['eff'] = None
                        new_ind['s'] = status_eff(new_ind['eff'])
                        bloco['indicadores'][i] = new_ind

    result['blocos']=blocos
    return result

# ── SCORING (ponderado: verde=1, amarelo=0.5, vermelho=0) ────────────────────

OBRIG={
    'acompanhamento':['mrr vendido','volume abordado','volume vendido',
                      'taxa de conversao','safra de negocio','clientes em carteira','total em carteira',
                      'volume indicacoes','volume de indicacoes'],
    'seguro':['vendido','volume abordado','volume vendido','volume vendas','taxa de conversao'],
    'servicos':['valor vendido','vendido','volume abordado','volume vendido','volume vendas','taxa de conversao']
}

def score_empresa(data):
    pts=0.0; max_pts=0.0
    verde=amarelo=vermelho=nd=0
    matched={bk:set() for bk in OBRIG}
    for bk,req_list in OBRIG.items():
        bloco=data['blocos'].get(bk)
        if not bloco: continue
        for ind in bloco['indicadores']:
            if ind['sub']: continue
            n=norm(ind['n'])
            for req in req_list:
                if req in n and req not in matched[bk]:
                    matched[bk].add(req)
                    s=ind['s']
                    if s=='verde':   pts+=1.0; max_pts+=1.0; verde+=1
                    elif s=='amarelo': pts+=0.5; max_pts+=1.0; amarelo+=1
                    elif s=='vermelho': max_pts+=1.0; vermelho+=1
                    else: nd+=1
                    break
    score=round(pts/max_pts*100,1) if max_pts>0 else None
    return {'score':score,'verde':verde,'amarelo':amarelo,
            'vermelho':vermelho,'nd':nd,'total':verde+amarelo+vermelho+nd}

def get_ind(blocos,bk,*keywords):
    bloco=blocos.get(bk)
    if not bloco: return None
    for ind in bloco['indicadores']:
        if ind['sub']: continue
        n=norm(ind['n'])
        if any(kw in n for kw in keywords): return ind
    return None

# ── DIAGNÓSTICO AUTOMÁTICO ───────────────────────────────────────────────────

def generate_diagnosis(data):
    b=data['blocos']
    obs=[]

    if not b:
        obs.append({'tipo':'critico','texto':'Nenhum bloco de dados encontrado na planilha'})
        return obs

    mrr_acc  = get_ind(b,'acompanhamento','mrr acumulado')
    mrr_vend = get_ind(b,'acompanhamento','mrr vendido')
    vol_abord= get_ind(b,'acompanhamento','volume abordado')
    vol_vend = get_ind(b,'acompanhamento','volume vendido')
    safra    = get_ind(b,'acompanhamento','safra de negocio')
    carteira = get_ind(b,'acompanhamento','clientes em carteira','total em carteira')
    agentes  = get_ind(b,'acompanhamento','quantidade de agente','quantidade de operador','quantida de operador')
    indicacoes=get_ind(b,'acompanhamento','volume indicacoes','indicacoes')
    conversao= get_ind(b,'acompanhamento','taxa de conversao')
    seguro   = get_ind(b,'seguro','vendido','valor vendido')
    servicos = get_ind(b,'servicos','valor vendido','vendido')
    churn    = get_ind(b,'acompanhamento','mrr churn')

    # Verifica se há algum executado preenchido
    has_exec=any(
        ind['e'] is not None and ind['e']!=0
        for blk in b.values()
        for ind in blk['indicadores']
        if not ind['sub']
    )
    if not has_exec:
        obs.append({'tipo':'critico','texto':'Planilha sem dados executados preenchidos — consultor não atualizou o plano'})
        return obs

    # ── POSITIVOS ──
    if mrr_acc and mrr_acc['eff'] and mrr_acc['eff']>=1.0:
        obs.append({'tipo':'positivo','texto':f'MRR Acumulado acima da meta ({fmt_pct(mrr_acc["eff"])}) — base sólida construída'})
    elif mrr_acc and mrr_acc['eff'] and mrr_acc['eff']>=0.8:
        obs.append({'tipo':'positivo','texto':f'MRR Acumulado saudável ({fmt_pct(mrr_acc["eff"])} da meta)'})

    if mrr_vend and mrr_vend['eff'] and mrr_vend['eff']>=1.0:
        obs.append({'tipo':'positivo','texto':f'MRR Vendido acima da meta ({fmt_pct(mrr_vend["eff"])}) — bom ritmo de novas vendas'})

    if seguro and seguro['eff'] and seguro['eff']>=0.8:
        obs.append({'tipo':'positivo','texto':f'Seguro com boa performance ({fmt_pct(seguro["eff"])} da meta)'})
    if seguro and seguro['eff'] and seguro['eff']>=1.0:
        obs[-1]['texto']=f'Seguro acima da meta ({fmt_pct(seguro["eff"])}) — destaque em produto financeiro'

    if servicos and servicos['eff'] and servicos['eff']>=0.8:
        obs.append({'tipo':'positivo','texto':f'Serviços com boa performance ({fmt_pct(servicos["eff"])} da meta)'})
    if servicos and servicos['eff'] and servicos['eff']>=1.0:
        obs[-1]['texto']=f'Serviços acima da meta ({fmt_pct(servicos["eff"])}) — receita avulsa acima do esperado'

    if conversao and conversao['eff'] and conversao['eff']>=1.0:
        obs.append({'tipo':'positivo','texto':f'Taxa de conversão acima da meta ({fmt_pct(conversao["eff"])}) — equipe eficiente'})

    if indicacoes and indicacoes['e'] and indicacoes['e']>0:
        obs.append({'tipo':'positivo','texto':f'{fmt_num(indicacoes["e"])} indicações geradas — base gerando novos leads organicamente'})

    if carteira and carteira['eff'] and carteira['eff']>=1.0:
        obs.append({'tipo':'positivo','texto':f'Carteira acima da meta ({fmt_pct(carteira["eff"])}) — expansão da base'})

    # ── CRÍTICOS ──
    # Vol. abordado composto = soma das 4 abordagens (igual ao painel)
    _va_comp_parts=[(get_ind(b,bk,kw) or {}).get('e') or 0 for bk,kw in
                    [('acompanhamento','volume abordado'),('servicos','volume abordado'),
                     ('seguro','volume abordado'),('acompanhamento','safra de negocio')]]
    _va_total=sum(_va_comp_parts)
    _va_any=any((get_ind(b,bk,kw) is not None) for bk,kw in
                [('acompanhamento','volume abordado'),('servicos','volume abordado'),
                 ('seguro','volume abordado'),('acompanhamento','safra de negocio')])
    if _va_any and _va_total==0:
        obs.append({'tipo':'critico','texto':'Volume total de abordagem = 0 (monitoramento + serviços + seguro + safra) — sem operação ativa'})
    elif vol_abord and vol_abord['eff'] and vol_abord['eff']<0.3:
        obs.append({'tipo':'critico','texto':f'Volume Abordado (monitoramento) muito baixo ({fmt_pct(vol_abord["eff"])} da meta) — abordagem insuficiente'})

    if mrr_vend and mrr_vend['e']==0:
        obs.append({'tipo':'critico','texto':'MRR Vendido = 0 — nenhuma nova venda de plano de monitoramento no período'})

    if agentes and agentes['e']==0:
        obs.append({'tipo':'critico','texto':'Nenhum agente/operador registrado — equipe de pós-vendas indefinida'})

    if safra and safra['e']==0 and safra['m'] and safra['m']>0:
        obs.append({'tipo':'critico','texto':'Safra de Negócio vazia com meta definida — pipeline esgotado'})

    # ── ALERTAS ──
    if safra and safra['e']==0 and (safra['m'] is None or safra['m']==0):
        obs.append({'tipo':'alerta','texto':'Safra de Negócio = 0 — sem pipeline ativo no período'})

    if indicacoes and indicacoes['e']==0:
        obs.append({'tipo':'alerta','texto':'Zero indicações geradas — clientes não estão sendo engajados para indicar'})

    if mrr_acc and mrr_acc['eff'] and mrr_acc['eff']<0.6:
        obs.append({'tipo':'alerta','texto':f'MRR Acumulado abaixo de 60% da meta ({fmt_pct(mrr_acc["eff"])}) — construção de base está lenta'})

    if carteira and carteira['eff'] and carteira['eff']<0.5:
        obs.append({'tipo':'alerta','texto':f'Clientes em Carteira abaixo de 50% da meta ({fmt_pct(carteira["eff"])})'})

    if churn and churn['e'] and churn['e']>0:
        obs.append({'tipo':'alerta','texto':f'Churn de MRR registrado: {fmt_money(churn["e"])} — monitorar cancelamentos'})

    if vol_vend and vol_vend['e']==0 and vol_abord and vol_abord['e'] and vol_abord['e']>0:
        obs.append({'tipo':'alerta','texto':f'Abordagem realizada mas zero vendas — problema de conversão'})

    if not obs:
        obs.append({'tipo':'neutro','texto':'Dados parcialmente preenchidos — análise incompleta'})

    return obs

# ── CONSOLIDADO ──────────────────────────────────────────────────────────────

def consolidate(all_data):
    total_mrr=total_mrr_vend=total_carteira=total_indicacoes=0
    total_seguro=total_servicos=total_vol_abord=total_vol_vend=0
    qtd_seguro=qtd_servicos=0
    counts={'verde':0,'amarelo':0,'vermelho':0,'nd':0}
    mrr_mensal={m:0 for m in range(1,13)}
    mrr_vend_mensal={m:0 for m in range(1,13)}

    def _sum_mensal(ind):
        if not ind or not ind.get('mensal'): return None
        vals=[v['e'] for v in ind['mensal'].values() if v.get('e') is not None]
        return sum(vals) if vals else None

    def _last_mensal(ind):
        if not ind: return None
        if not ind.get('mensal'): return ind.get('e')
        filled=[(int(m),v['e']) for m,v in ind['mensal'].items() if v.get('e') is not None]
        if not filled: return ind.get('e')
        return max(filled,key=lambda x:x[0])[1]

    for d in all_data:
        s=d.get('_status_geral','nd')
        counts[s]=counts.get(s,0)+1
        b=d['blocos']

        # MRR Acumulado: usa ind['e'] pós-processado (último mês)
        ind_mrr=get_ind(b,'acompanhamento','mrr acumulado')
        if ind_mrr and ind_mrr['e'] is not None:
            total_mrr+=ind_mrr['e']
            for m,v in ind_mrr.get('mensal',{}).items():
                if v.get('e') is not None:
                    mrr_mensal[m]=mrr_mensal.get(m,0)+v['e']

        # MRR Vendido: soma mensal executada (fallback col.D)
        ind_vend=get_ind(b,'acompanhamento','mrr vendido')
        if ind_vend:
            sm=_sum_mensal(ind_vend)
            val=sm if sm is not None else ind_vend.get('e')
            if val is not None: total_mrr_vend+=val
            for m,v in ind_vend.get('mensal',{}).items():
                if v.get('e') is not None:
                    mrr_vend_mensal[m]=mrr_vend_mensal.get(m,0)+v['e']

        # Carteira: último mês preenchido
        lm=_last_mensal(get_ind(b,'acompanhamento','clientes em carteira','total em carteira'))
        if lm is not None: total_carteira+=lm

        # Indicações: soma mensal (fallback col.D)
        v=get_ind(b,'acompanhamento','volume indicacoes','indicacoes')
        sm=_sum_mensal(v); val=sm if sm is not None else (v.get('e') if v else None)
        if val is not None: total_indicacoes+=val

        # Vol. Abordado e Vol. Vendido: col.D (base para Conversão Geral)
        v=get_ind(b,'acompanhamento','volume abordado')
        if v and v['e']: total_vol_abord+=v['e']
        v=get_ind(b,'acompanhamento','volume vendido')
        if v and v['e']: total_vol_vend+=v['e']

        # Seguro: valor soma mensal (fallback col.D)
        v=get_ind(b,'seguro','vendido','valor vendido')
        sm=_sum_mensal(v); val=sm if sm is not None else (v.get('e') if v else None)
        if val is not None: total_seguro+=val

        # Seguro: quantidade soma mensal (fallback col.D)
        v=get_ind(b,'seguro','volume vendido','volume vendas')
        sm=_sum_mensal(v); val=sm if sm is not None else (v.get('e') if v else None)
        if val is not None: qtd_seguro+=val

        # Serviços: valor soma mensal (fallback col.D)
        v=get_ind(b,'servicos','valor vendido','vendido')
        sm=_sum_mensal(v); val=sm if sm is not None else (v.get('e') if v else None)
        if val is not None: total_servicos+=val

        # Serviços: quantidade soma mensal (fallback col.D)
        v=get_ind(b,'servicos','volume vendido','volume vendas')
        sm=_sum_mensal(v); val=sm if sm is not None else (v.get('e') if v else None)
        if val is not None: qtd_servicos+=val

    conv=(total_vol_vend/total_vol_abord) if total_vol_abord>0 else None
    return {
        'total_mrr':total_mrr,'total_mrr_vend':total_mrr_vend,
        'total_carteira':int(total_carteira),'total_indicacoes':int(total_indicacoes),
        'total_seguro':total_seguro,'total_servicos':total_servicos,
        'qtd_seguro':int(qtd_seguro),'qtd_servicos':int(qtd_servicos),
        'ticket_seguro':round(total_seguro/qtd_seguro,2) if qtd_seguro>0 else None,
        'ticket_servicos':round(total_servicos/qtd_servicos,2) if qtd_servicos>0 else None,
        'conversao_geral':conv,'counts':counts,
        'mrr_mensal':{str(k):round(v,2) for k,v in mrr_mensal.items()},
        'mrr_vend_mensal':{str(k):round(v,2) for k,v in mrr_vend_mensal.items()},
    }

# ── DIAGNÓSTICO DE INDICADORES FALTANTES ─────────────────────────────────────

DIAG_CHECKS = {
    'acompanhamento': [
        ('MRR Acumulado',        ['mrr acumulado']),
        ('MRR Vendido',          ['mrr vendido']),
        ('Vol. Abordado',        ['volume abordado']),
        ('Vol. Vendido',         ['volume vendido']),
        ('Taxa de Conversão',    ['taxa de conversao']),
        ('Safra de Negócio',     ['safra de negocio']),
        ('Clientes em Carteira', ['clientes em carteira','total em carteira']),
        ('Indicações',           ['volume indicacoes','indicacoes']),
        ('Agentes/Operadores',   ['quantidade de agente','quantidade de operador','quantida de operador']),
    ],
    'seguro': [
        ('Valor Vendido',     ['vendido','valor vendido']),
        ('Vol. Abordado',     ['volume abordado']),
        ('Vol. Vendido',      ['volume vendido','volume vendas']),
        ('Taxa de Conversão', ['taxa de conversao']),
    ],
    'servicos': [
        ('Valor Vendido',     ['valor vendido','vendido']),
        ('Vol. Abordado',     ['volume abordado']),
        ('Vol. Vendido',      ['volume vendido','volume vendas']),
        ('Taxa de Conversão', ['taxa de conversao']),
    ],
}

def gerar_diagnostico(all_data, output_dir):
    relatorio = []
    for d in all_data:
        b = d.get('blocos', {})
        if not b:
            continue
        problemas = {}
        for bloco, checks in DIAG_CHECKS.items():
            bloco_data = b.get(bloco)
            if bloco_data is None:
                continue
            nomes_no_bloco = [ind['n'] for ind in bloco_data['indicadores'] if not ind['sub']]
            for label, keywords in checks:
                encontrado = any(
                    any(kw in norm(ind['n']) for kw in keywords)
                    for ind in bloco_data['indicadores'] if not ind['sub']
                )
                if not encontrado:
                    problemas.setdefault(bloco, {})[label] = {
                        'buscava': keywords,
                        'encontrados_no_bloco': nomes_no_bloco,
                    }
        if problemas:
            relatorio.append({'id': d.get('id',''), 'nome': d.get('nome',''), 'problemas': problemas})

    out_path = os.path.join(output_dir, 'diagnostico.json')
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(relatorio, f, ensure_ascii=False, indent=2)
    return relatorio

# ── INSIGHTS GLOBAIS ─────────────────────────────────────────────────────────

def global_insights(all_data):
    insights=[]
    n=len(all_data)

    # MRR acima da meta
    mrr_ok=[d for d in all_data if get_ind(d['blocos'],'acompanhamento','mrr acumulado') and
            get_ind(d['blocos'],'acompanhamento','mrr acumulado')['eff'] and
            get_ind(d['blocos'],'acompanhamento','mrr acumulado')['eff']>=1.0]
    if mrr_ok:
        insights.append({'tipo':'positivo','texto':f'{len(mrr_ok)} empresa(s) com MRR Acumulado acima da meta',
                         'detalhe':', '.join(d['nome'][:20] for d in mrr_ok[:5])})

    # Volume Abordado composto zerado (soma das 4 abordagens)
    def _va_total(d):
        return sum((get_ind(d['blocos'],bk,kw) or {}).get('e') or 0
                   for bk,kw in [('acompanhamento','volume abordado'),('servicos','volume abordado'),
                                  ('seguro','volume abordado'),('acompanhamento','safra de negocio')])
    def _va_exists(d):
        return any(get_ind(d['blocos'],bk,kw) is not None
                   for bk,kw in [('acompanhamento','volume abordado'),('servicos','volume abordado'),
                                  ('seguro','volume abordado'),('acompanhamento','safra de negocio')])
    vol_zero=[d for d in all_data if _va_exists(d) and _va_total(d)==0]
    if vol_zero:
        insights.append({'tipo':'critico','texto':f'{len(vol_zero)} empresa(s) com abordagem total = 0 — sem operação ativa',
                         'detalhe':', '.join(d['nome'][:20] for d in vol_zero[:5])})

    # MRR Vendido zerado
    mrr_vz=[d for d in all_data if get_ind(d['blocos'],'acompanhamento','mrr vendido') and
            get_ind(d['blocos'],'acompanhamento','mrr vendido')['e']==0]
    if mrr_vz:
        insights.append({'tipo':'alerta','texto':f'{len(mrr_vz)} empresa(s) sem novas vendas de MRR no período',
                         'detalhe':', '.join(d['nome'][:20] for d in mrr_vz[:5])})

    # Safra vazia
    safra_z=[d for d in all_data if get_ind(d['blocos'],'acompanhamento','safra de negocio') and
             get_ind(d['blocos'],'acompanhamento','safra de negocio')['e']==0]
    if safra_z:
        insights.append({'tipo':'alerta','texto':f'{len(safra_z)} empresa(s) com Safra de Negócio vazia — pipeline em risco',
                         'detalhe':', '.join(d['nome'][:20] for d in safra_z[:5])})

    # Zero indicações
    ind_z=[d for d in all_data if get_ind(d['blocos'],'acompanhamento','volume indicacoes','indicacoes') and
           get_ind(d['blocos'],'acompanhamento','volume indicacoes','indicacoes')['e']==0]
    if ind_z:
        insights.append({'tipo':'alerta','texto':f'{len(ind_z)} empresa(s) sem indicações geradas no período',
                         'detalhe':', '.join(d['nome'][:20] for d in ind_z[:5])})

    # Seguro destaque
    seg_ok=[d for d in all_data if get_ind(d['blocos'],'seguro','vendido','valor vendido') and
            get_ind(d['blocos'],'seguro','vendido','valor vendido')['eff'] and
            get_ind(d['blocos'],'seguro','vendido','valor vendido')['eff']>=1.0]
    if seg_ok:
        insights.append({'tipo':'positivo','texto':f'{len(seg_ok)} empresa(s) com Seguro acima da meta — produto financeiro bem posicionado',
                         'detalhe':', '.join(d['nome'][:20] for d in seg_ok[:5])})

    # Plano não preenchido
    sem_dados=[d for d in all_data if all(
        not any(ind['e'] is not None and ind['e']!=0
                for ind in blk['indicadores'] if not ind['sub'])
        for blk in d['blocos'].values()
    ) and d['blocos']]
    if sem_dados:
        insights.append({'tipo':'critico','texto':f'{len(sem_dados)} empresa(s) com plano sem dados executados — planilha desatualizada',
                         'detalhe':', '.join(d['nome'][:20] for d in sem_dados[:5])})

    return insights

# ── GOOGLE SHEETS HELPERS ────────────────────────────────────────────────────

def extract_sheet_id(url):
    """Extrai o ID do Google Sheets de qualquer URL de compartilhamento."""
    m = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', url)
    return m.group(1) if m else None

def fetch_master(master_id):
    """Baixa a planilha mestre como CSV e retorna lista de empresas ativas.

    Estrutura real da planilha:
      Empresa          → "ID - Nome" (ex: "14393 - IRRADIAR")
      Consultor
      Data de Inicio
      Link do Plano de Negocio  → URL do Google Sheets
      Status           → vazio (ativo) | CRÍTICO | ALERTA | inativo | excluido
      Observação
    """
    url = f"https://docs.google.com/spreadsheets/d/{master_id}/export?format=csv"
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    # Google Sheets exporta em UTF-8 (requests erra o default para ISO-8859-1)
    content = resp.content.decode('utf-8')
    reader = csv.DictReader(io.StringIO(content))
    # Detecta coluna de observação dinamicamente (ignora espaços e acentuação)
    obs_field = next((f for f in (reader.fieldnames or []) if 'obs' in norm(f)), None)
    result = []
    for row in reader:
        status_raw = row.get('Status', row.get('status', '')).strip()
        status = status_raw.lower()
        # Excluir apenas inativos explícitos — CRÍTICO e ALERTA ainda processam
        if status in ('inativo', 'excluido', 'excluído'):
            continue
        link = row.get('Link do Plano de Negocio',
                       row.get('link do plano de negocio', '')).strip()
        if not link or '/spreadsheets/d/' not in link:
            continue
        # Parsear "ID - Nome" da coluna Empresa
        empresa_raw = row.get('Empresa', row.get('empresa', '')).strip()
        partes = empresa_raw.split(' - ', 1)
        emp_id   = partes[0].strip()
        emp_nome = partes[1].strip() if len(partes) > 1 else empresa_raw
        st_upper = status_raw.upper()
        obs_raw  = row.get(obs_field, '').strip() if obs_field else ''
        result.append({
            'id':          emp_id,
            'nome':        emp_nome,
            'link':        link,
            'status':      status_raw,
            'consultor':   row.get('Consultor', '').strip(),
            'data_inicio': row.get('Data de Inicio', '').strip(),
            'status_tag':  status_raw if st_upper in ('CRÍTICO', 'CRITICO', 'ALERTA') else '',
            'observacao':  obs_raw if st_upper in ('CRÍTICO', 'CRITICO', 'ALERTA') else '',
        })
    return result

def download_plan(sheet_id, session):
    """Baixa o xlsx do plano de negócio com retry. Retorna bytes ou None."""
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    for attempt in range(MAX_RETRIES + 1):
        try:
            resp = session.get(url, timeout=45)
            if resp.status_code == 200 and resp.content[:4] == b'PK\x03\x04':
                return resp.content
            if resp.status_code in (401, 403):
                return None
        except requests.RequestException:
            pass
        if attempt < MAX_RETRIES:
            time.sleep(DOWNLOAD_DELAY * (attempt + 1))
    return None

# ── MAIN ─────────────────────────────────────────────────────────────────────

def _processar(data):
    score = score_empresa(data)
    data['_score'] = score
    s = score['score']
    if s is None:  data['_status_geral'] = 'nd'
    elif s >= 60:  data['_status_geral'] = 'verde'
    elif s >= 35:  data['_status_geral'] = 'amarelo'
    else:          data['_status_geral'] = 'vermelho'
    data['_diagnosis'] = generate_diagnosis(data)
    return data

def main():
    all_data = []

    if MASTER_SHEET_ID:
        # ── Modo online: Google Sheets ──────────────────────────────────────
        print(f'Buscando planilha mestre {MASTER_SHEET_ID}...')
        empresas = fetch_master(MASTER_SHEET_ID)
        print(f'Encontradas {len(empresas)} empresas ativas.\n')

        session = requests.Session()
        for emp in empresas:
            sheet_id = extract_sheet_id(emp['link'])
            nome_log = emp['nome'] or emp['id'] or emp['link'][:40]
            if not sheet_id:
                print(f'  SKIP: {nome_log} — link inválido')
                all_data.append({'empresa': nome_log, 'id': emp['id'],
                                 'nome': emp['nome'], 'blocos': {}, 'erros': ['link inválido'],
                                 '_score': None, '_status_geral': 'nd', '_diagnosis': [],
                                 'consultor': emp.get('consultor',''), 'data_inicio': emp.get('data_inicio',''), 'status_tag': emp.get('status_tag',''), 'observacao': emp.get('observacao','')})
                continue
            content = download_plan(sheet_id, session)
            if content is None:
                print(f'  ERRO download: {nome_log}')
                all_data.append({'empresa': nome_log, 'id': emp['id'],
                                 'nome': emp['nome'], 'blocos': {}, 'erros': ['falha no download'],
                                 '_score': None, '_status_geral': 'nd', '_diagnosis': [],
                                 'consultor': emp.get('consultor',''), 'data_inicio': emp.get('data_inicio',''), 'status_tag': emp.get('status_tag',''), 'observacao': emp.get('observacao','')})
                time.sleep(DOWNLOAD_DELAY)
                continue
            try:
                data = parse_file(content, empresa_id=emp['id'], empresa_nome=emp['nome'])
                data = _processar(data)
                s = (data['_score'] or {}).get('score')
                print(f'  OK: {nome_log[:45]:<45} | Score: {s}%')
            except Exception as ex:
                data = {'empresa': nome_log, 'id': emp['id'], 'nome': emp['nome'],
                        'blocos': {}, 'erros': [str(ex)],
                        '_score': None, '_status_geral': 'nd', '_diagnosis': []}
                print(f'  ERRO parse: {nome_log} -> {ex}')
            data['consultor']   = emp.get('consultor', '')
            data['data_inicio'] = emp.get('data_inicio', '')
            data['status_tag']  = emp.get('status_tag', '')
            data['observacao']  = emp.get('observacao', '')
            all_data.append(data)
            time.sleep(DOWNLOAD_DELAY)

    else:
        # ── Modo offline: pasta local (fallback) ────────────────────────────
        print(f'MASTER_SHEET_ID não definido — usando pasta local: {PASTA}\n')
        files = sorted([os.path.join(PASTA, f) for f in os.listdir(PASTA)
                        if f.endswith('.xlsx') and not f.startswith('~')])
        print(f'Processando {len(files)} arquivos...\n')
        for fpath in files:
            nome = os.path.basename(fpath)
            try:
                data = parse_file(fpath)
                data = _processar(data)
                s = (data['_score'] or {}).get('score')
                blocos_ok = list(data['blocos'].keys())
                print(f'  OK: {nome[:45]:<45} | Blocos: {blocos_ok} | Score: {s}%')
            except Exception as ex:
                data = {'empresa': nome, 'id': '', 'nome': nome, 'blocos': {}, 'erros': [str(ex)],
                        '_score': None, '_status_geral': 'nd', '_diagnosis': []}
                print(f'  ERRO: {nome} -> {ex}')
            all_data.append(data)

    all_data.sort(key=lambda x: (x['_score'] or {}).get('score') or -1, reverse=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    gerar_diagnostico(all_data, OUTPUT_DIR)
    consolidado = consolidate(all_data)
    insights = global_insights(all_data)
    print(f'\nConsolidado: MRR={fmt_money(consolidado["total_mrr"])} | Carteira={fmt_num(consolidado["total_carteira"])} | {consolidado["counts"]}')
    html = generate_html(all_data, consolidado, insights)
    with open(OUTPUT, 'w', encoding='utf-8') as f: f.write(html)
    print(f'Dashboard gerado: {OUTPUT}')
    tv_html = generate_tv_html(all_data, consolidado, insights)
    with open(OUTPUT_TV, 'w', encoding='utf-8') as f: f.write(tv_html)
    print(f'Dashboard TV gerado: {OUTPUT_TV}')

# ── HTML GENERATOR ────────────────────────────────────────────────────────────

def generate_html(all_data, consolidado, insights):
    # ── JS data ──
    js_data=[]
    for d in all_data:
        blocos_js={}
        for bk,bloco in d['blocos'].items():
            inds=[{'n':i['n'],'m':i['m'],'e':i['e'],
                   'eff':round(i['eff']*100,1) if i['eff'] is not None else None,
                   's':i['s'],'sub':i['sub'],
                   'mensal':{str(k):v for k,v in i.get('mensal',{}).items()}}
                  for i in bloco['indicadores']]
            blocos_js[bk]={'nome':bloco['nome'],'inds':inds}
        b=d['blocos']
        def iv(k,*kw): i=get_ind(b,k,*kw); return i['e'] if i else None
        def ie(k,*kw):
            i=get_ind(b,k,*kw)
            return round(i['eff']*100,1) if i and i['eff'] is not None else None
        def iss(k,*kw): i=get_ind(b,k,*kw); return i['s'] if i else 'nd'
        sc=d.get('_score') or {}
        # Vol. Abordado = soma de 4 abordagens
        _va_ks=[('acompanhamento','volume abordado'),('servicos','volume abordado'),
                ('seguro','volume abordado'),('acompanhamento','safra de negocio')]
        _va_execs=[x for x in [iv(bk,kw) for bk,kw in _va_ks] if x is not None]
        _va_metas=[gi['m'] for bk,kw in _va_ks for gi in [get_ind(b,bk,kw)] if gi and gi['m'] is not None]
        _va_e=sum(_va_execs) if _va_execs else None
        _va_m=sum(_va_metas) if _va_metas else None
        _va_eff=round(_va_e/_va_m*100,1) if _va_e is not None and _va_m and _va_m>0 else None
        _va_s=status_eff(_va_eff/100) if _va_eff is not None else 'nd'
        # Agentes PV
        _agentes=get_ind(b,'acompanhamento','quantidade de agente','quantidade de operador','quantida de operador')
        _agentes_e=_agentes['e'] if _agentes else None
        js_data.append({
            'id':d['id'],'nome':d['nome'],'periodo':d.get('periodo'),
            'status':d.get('_status_geral','nd'),'score':sc.get('score'),
            'verde':sc.get('verde',0),'amarelo':sc.get('amarelo',0),
            'vermelho':sc.get('vermelho',0),'nd':sc.get('nd',0),'total_inds':sc.get('total',0),
            'erros':d.get('erros',[]),'diagnosis':d.get('_diagnosis',[]),
            'mrr_acc_e':iv('acompanhamento','mrr acumulado'),
            'mrr_acc_m':get_ind(b,'acompanhamento','mrr acumulado') and get_ind(b,'acompanhamento','mrr acumulado')['m'],
            'mrr_acc_eff':ie('acompanhamento','mrr acumulado'),
            'mrr_acc_s':iss('acompanhamento','mrr acumulado'),
            'mrr_vend_e':iv('acompanhamento','mrr vendido'),
            'mrr_vend_eff':ie('acompanhamento','mrr vendido'),
            'mrr_vend_s':iss('acompanhamento','mrr vendido'),
            'carteira_e':iv('acompanhamento','clientes em carteira','total em carteira'),
            'carteira_s':iss('acompanhamento','clientes em carteira','total em carteira'),
            'conversao_e':(lambda i: round(i['e']*100,1) if i and i['e'] is not None else None)(get_ind(b,'acompanhamento','taxa de conversao')),
            'conversao_eff':ie('acompanhamento','taxa de conversao'),
            'conversao_s':iss('acompanhamento','taxa de conversao'),
            'conv_acomp_e':(lambda i: round(i['e']*100,1) if i and i['e'] is not None else None)(get_ind(b,'acompanhamento','taxa de conversao')),
            'conv_acomp_s':iss('acompanhamento','taxa de conversao'),
            'conv_srv_e':(lambda i: round(i['e']*100,1) if i and i['e'] is not None else None)(get_ind(b,'servicos','taxa de conversao')),
            'conv_srv_s':iss('servicos','taxa de conversao'),
            'conv_seg_e':(lambda i: round(i['e']*100,1) if i and i['e'] is not None else None)(get_ind(b,'seguro','taxa de conversao')),
            'conv_seg_s':iss('seguro','taxa de conversao'),
            'vol_abord_e':_va_e,
            'vol_abord_eff':_va_eff,
            'vol_abord_s':_va_s,
            'seguro_e':iv('seguro','vendido','valor vendido'),
            'seguro_eff':ie('seguro','vendido','valor vendido'),
            'seguro_s':iss('seguro','vendido','valor vendido'),
            'servicos_e':iv('servicos','valor vendido','vendido'),
            'servicos_eff':ie('servicos','valor vendido','vendido'),
            'servicos_s':iss('servicos','valor vendido','vendido'),
            'indicacoes_e':iv('acompanhamento','volume indicacoes','indicacoes'),
            'agentes_e':_agentes_e,
            'safra_e':iv('acompanhamento','safra de negocio'),
            'safra_eff':ie('acompanhamento','safra de negocio'),
            'consultor':d.get('consultor',''),
            'data_inicio':d.get('data_inicio',''),
            'status_tag':d.get('status_tag',''),
            'observacao':d.get('observacao',''),
            'blocos':blocos_js
        })

    data_json=json.dumps(js_data,ensure_ascii=False,default=str)
    consol_json=json.dumps(consolidado,ensure_ascii=False,default=str)
    insights_json=json.dumps(insights,ensure_ascii=False)
    now=datetime.now().strftime('%d/%m/%Y às %H:%M')
    n=len(all_data)

    _logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'assets', 'solarz-logo-branca.png')
    try:
        with open(_logo_path,'rb') as _lf:
            _logo_b64 = base64.b64encode(_lf.read()).decode('ascii')
        logo_html = '<img src="data:image/png;base64,' + _logo_b64 + '" height="26" alt="SolarZ">'
    except Exception:
        logo_html = '<span style="color:#FF5500;font-weight:800;font-size:16px;letter-spacing:-0.02em">SolarZ</span>'

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>SolarZ · Planos de Negócio 2026</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Inter',system-ui,-apple-system,sans-serif;background:#0a0f1e;background-image:linear-gradient(160deg,#0a0f1e 0%,#0d1835 55%,#0a0f1e 100%);background-attachment:fixed;color:#F0F0F0;font-size:14px;line-height:1.5}}
:root{{
  --verde:#22C55E;--verde-bg:rgba(34,197,94,0.12);
  --amarelo:#F59E0B;--amarelo-bg:rgba(245,158,11,0.12);
  --vermelho:#EF4444;--vermelho-bg:rgba(239,68,68,0.12);
  --azul:#3B82F6;--laranja:#FF5500;
  --bg:#181D1E;--card:rgba(255,255,255,0.04);--card2:rgba(255,255,255,0.07);
  --card-border:rgba(255,255,255,0.08);--text:#F0F0F0;--muted:rgba(255,255,255,0.55);--faint:rgba(255,255,255,0.3)
}}
/* ── SCROLLBAR ── */
::-webkit-scrollbar{{width:6px;height:6px}}
::-webkit-scrollbar-track{{background:#181D1E}}
::-webkit-scrollbar-thumb{{background:rgba(255,255,255,0.15);border-radius:3px}}
::-webkit-scrollbar-thumb:hover{{background:rgba(255,255,255,0.25)}}
/* ── HEADER ── */
header{{background:rgba(0,0,0,0.75);backdrop-filter:blur(10px);-webkit-backdrop-filter:blur(10px);border-bottom:1px solid rgba(255,255,255,0.08);padding:0 24px;height:54px;display:flex;align-items:center;gap:12px;position:sticky;top:0;z-index:200;box-shadow:0 2px 16px rgba(0,0,0,0.4)}}
.logo{{display:flex;align-items:center;gap:10px}}
.logo img{{height:26px;filter:drop-shadow(0 0 6px rgba(255,85,0,0.2))}}
.logo-sep{{color:rgba(255,255,255,0.2);font-size:15px;font-weight:300;margin:0 2px}}
.logo-title{{font-weight:600;font-size:14px;color:rgba(255,255,255,0.65);letter-spacing:-0.01em}}
.header-meta{{margin-left:auto;color:rgba(255,255,255,0.32);font-size:12px;font-weight:500}}
.sync-btn{{padding:5px 13px;border-radius:20px;border:1px solid rgba(59,130,246,0.35);background:rgba(59,130,246,0.1);color:#93c5fd;cursor:pointer;font-size:12px;font-weight:600;transition:.15s;font-family:inherit;margin-left:10px;white-space:nowrap}}
.sync-btn:hover:not(:disabled){{background:rgba(59,130,246,0.2);border-color:rgba(59,130,246,0.6)}}
.sync-btn:disabled{{opacity:0.6;cursor:not-allowed}}
.sync-cfg{{background:none;border:none;color:rgba(255,255,255,0.22);cursor:pointer;font-size:15px;padding:2px 6px;transition:.15s;line-height:1;margin-left:2px}}
.sync-cfg:hover{{color:rgba(255,255,255,0.65)}}
#syncModal{{display:none;position:fixed;inset:0;background:rgba(0,0,0,0.72);z-index:1000;align-items:center;justify-content:center}}
.smbox{{background:#1a1f2e;border:1px solid rgba(255,255,255,0.12);border-radius:12px;padding:24px;width:390px;max-width:92vw}}
.smbox h3{{color:#F0F0F0;font-size:15px;font-weight:700;margin:0 0 7px}}
.smbox p{{color:rgba(255,255,255,0.42);font-size:12px;margin:0 0 18px;line-height:1.6}}
.smbox p strong{{color:rgba(255,255,255,0.65)}}
.smlabel{{color:rgba(255,255,255,0.38);font-size:10px;font-weight:700;letter-spacing:.08em;display:block;margin-bottom:4px;text-transform:uppercase}}
.sminput{{width:100%;padding:8px 12px;border-radius:6px;border:1px solid rgba(255,255,255,0.1);background:rgba(255,255,255,0.05);color:#F0F0F0;font-size:13px;margin-bottom:12px;box-sizing:border-box;font-family:inherit;outline:none;transition:.15s}}
.sminput:focus{{border-color:rgba(59,130,246,0.5);background:rgba(255,255,255,0.07)}}
.smactions{{display:flex;gap:8px;margin-top:4px}}
.smsave{{flex:1;padding:9px;border-radius:6px;border:none;background:#3b82f6;color:#fff;font-weight:700;font-size:13px;cursor:pointer;font-family:inherit;transition:.15s}}
.smsave:hover{{background:#2563eb}}
.smcancel{{padding:9px 16px;border-radius:6px;border:1px solid rgba(255,255,255,0.1);background:transparent;color:rgba(255,255,255,0.42);font-size:13px;cursor:pointer;font-family:inherit;transition:.15s}}
.smcancel:hover{{color:#F0F0F0}}
/* ── INTELIGÊNCIA OPERACIONAL ── */
.io-topbar{{display:flex;align-items:center;gap:16px;margin-bottom:18px;flex-wrap:wrap}}
.io-filter-wrap{{display:flex;align-items:center;gap:8px;margin-left:auto}}
.io-filter-label{{color:rgba(255,255,255,0.4);font-size:11px;font-weight:600;letter-spacing:.06em;text-transform:uppercase;white-space:nowrap}}
.io-filter-input{{padding:6px 12px;border-radius:20px;border:1px solid rgba(255,255,255,0.12);background:rgba(255,255,255,0.05);color:#F0F0F0;font-size:12px;width:210px;outline:none;font-family:inherit;transition:.15s}}
.io-filter-input:focus{{border-color:rgba(59,130,246,0.5);background:rgba(255,255,255,0.08)}}
.io-clear-btn{{background:none;border:none;color:rgba(255,255,255,0.3);cursor:pointer;font-size:14px;padding:4px 6px;transition:.15s;line-height:1}}
.io-clear-btn:hover{{color:#F0F0F0}}
.io-grid-3{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:14px}}
.obs-counts{{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:14px}}
.obs-count-pill{{display:flex;align-items:center;gap:7px;padding:5px 12px;border-radius:20px;border:1px solid;font-size:11px;font-weight:700}}
.obs-count-num{{font-size:18px;font-weight:800}}
.obs-kw-row{{display:flex;flex-wrap:wrap;gap:6px;margin-bottom:16px;padding:10px 12px;background:rgba(255,255,255,0.03);border-radius:8px;border:1px solid rgba(255,255,255,0.06)}}
.obs-kw-label{{font-size:10px;color:rgba(255,255,255,0.4);font-weight:700;text-transform:uppercase;letter-spacing:.07em;align-self:center;margin-right:4px;white-space:nowrap}}
.obs-kw{{display:inline-flex;align-items:center;gap:4px;padding:3px 9px;border-radius:12px;font-size:11px;font-weight:600;background:rgba(255,255,255,0.07);color:rgba(255,255,255,0.6)}}
.obs-kw-n{{font-size:10px;font-weight:800;color:rgba(255,255,255,0.4)}}
.obs-group-title{{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.08em;padding:6px 0;margin:4px 0 8px;border-bottom:1px solid rgba(255,255,255,0.07)}}
.obs-item{{display:flex;gap:10px;padding:9px 0;border-bottom:1px solid rgba(255,255,255,0.05)}}
.obs-item:last-child{{border-bottom:none}}
.obs-item-left{{flex-shrink:0;min-width:170px;max-width:200px}}
.obs-item-nome{{font-size:12px;font-weight:600;color:rgba(255,255,255,0.8);line-height:1.3}}
.obs-item-id{{font-size:10px;color:rgba(255,255,255,0.35);margin-top:2px}}
.obs-item-text{{font-size:12px;color:rgba(255,255,255,0.62);line-height:1.55;flex:1}}
.obs-empty{{font-size:12px;color:rgba(255,255,255,0.3);padding:12px 0;font-style:italic}}
@media(max-width:1100px){{.io-grid-3{{grid-template-columns:1fr 1fr}}}}
@media(max-width:700px){{.io-grid-3{{grid-template-columns:1fr}}}}
.io-card{{background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.08);border-radius:10px;padding:16px 18px}}
.io-card h3{{color:rgba(255,255,255,0.55);font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.08em;margin:0 0 14px}}
.io-card-wide{{background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.08);border-radius:10px;padding:16px 18px;margin-bottom:14px}}
.io-card-wide h3{{color:rgba(255,255,255,0.55);font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.08em;margin:0 0 14px}}
/* donut */
.io-donut-wrap{{display:flex;align-items:center;gap:18px}}
.io-donut{{width:130px;height:130px;border-radius:50%;flex-shrink:0;position:relative;display:flex;align-items:center;justify-content:center;background:rgba(255,255,255,0.08)}}
.io-donut-hole{{width:80px;height:80px;background:#0f1117;border-radius:50%;display:flex;flex-direction:column;align-items:center;justify-content:center;position:absolute}}
.io-donut-num{{font-size:22px;font-weight:700;color:#F0F0F0;line-height:1}}
.io-donut-lbl{{font-size:9px;color:rgba(255,255,255,0.38);letter-spacing:.06em;margin-top:2px}}
.io-legend{{display:flex;flex-direction:column;gap:8px;flex:1}}
.io-leg-item{{display:flex;align-items:center;gap:7px;font-size:12px;color:rgba(255,255,255,0.6)}}
.io-leg-item b{{color:#F0F0F0}}
.io-leg-dot{{width:8px;height:8px;border-radius:50%;flex-shrink:0}}
/* barras horizontais */
.io-hbar-list{{display:flex;flex-direction:column;gap:7px}}
.io-hbar-scroll{{max-height:260px;overflow-y:auto}}
.io-hbar{{display:flex;align-items:center;gap:8px}}
.io-hbar-label{{width:90px;font-size:11px;color:rgba(255,255,255,0.55);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;flex-shrink:0}}
.io-hbar-track{{flex:1;height:8px;background:rgba(255,255,255,0.07);border-radius:4px;position:relative;overflow:visible}}
.io-hbar-fill{{height:100%;border-radius:4px;transition:width .3s}}
.io-hbar-val{{width:60px;font-size:11px;color:rgba(255,255,255,0.55);text-align:right;flex-shrink:0;white-space:nowrap}}
.io-threshold{{position:absolute;top:-3px;bottom:-3px;width:2px;background:#F59E0B;border-radius:1px;opacity:.9}}
/* barras verticais */
.io-vbar-list{{display:flex;align-items:flex-end;gap:10px;padding:4px 0 0}}
.io-vbar{{display:flex;flex-direction:column;align-items:center;gap:4px;flex:1}}
.io-vbar-val{{font-size:11px;font-weight:700;color:rgba(255,255,255,0.7);white-space:nowrap;min-height:16px;display:flex;align-items:flex-end}}
.io-vbar-track{{width:100%;height:100px;background:rgba(255,255,255,0.07);border-radius:4px 4px 0 0;flex-shrink:0;position:relative;overflow:hidden}}
.io-vbar-fill{{position:absolute;bottom:0;left:0;right:0;border-radius:4px 4px 0 0;min-height:3px;transition:height .3s}}
.io-vbar-label{{font-size:9px;color:rgba(255,255,255,0.4);text-align:center;line-height:1.3;width:100%;margin-top:4px}}
/* ── TABS ── */
.tabs{{display:flex;background:rgba(0,0,0,0.45);backdrop-filter:blur(6px);border-bottom:1px solid rgba(255,255,255,0.07);padding:0 24px;overflow-x:auto}}
.tab{{padding:12px 18px;cursor:pointer;border-bottom:2px solid transparent;color:rgba(255,255,255,0.42);font-weight:600;font-size:13px;white-space:nowrap;transition:.15s;letter-spacing:0.01em}}
.tab.active{{color:#F59E0B;border-bottom-color:#F59E0B}}
.tab:hover:not(.active){{color:rgba(255,255,255,0.78)}}
/* ── SUMMARY STRIP ── */
.summary{{display:grid;grid-template-columns:repeat(auto-fill,minmax(170px,1fr));gap:12px;padding:18px 24px;background:#181D1E;border-bottom:1px solid rgba(255,255,255,0.06)}}
.scard{{background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:8px;padding:18px 20px;display:flex;flex-direction:column;gap:4px;box-shadow:0 2px 12px rgba(0,0,0,0.25);transition:.18s;cursor:default}}
.scard:hover{{background:rgba(255,255,255,0.06);border-color:rgba(255,255,255,0.12);transform:translateY(-1px)}}
.scard .lbl{{color:rgba(255,255,255,0.42);font-size:11px;text-transform:uppercase;letter-spacing:.07em;font-weight:700}}
.scard .val{{font-size:26px;font-weight:800;color:#F0F0F0;line-height:1.1;letter-spacing:-0.02em;font-variant-numeric:tabular-nums}}
.scard .sub{{font-size:11px;color:rgba(255,255,255,0.28)}}
.scard.cl-verde{{border-left:3px solid var(--verde)}}
.scard.cl-amarelo{{border-left:3px solid var(--amarelo)}}
.scard.cl-vermelho{{border-left:3px solid var(--vermelho)}}
.scard.cl-laranja{{border-left:3px solid var(--laranja)}}
/* ── KPI TV-STYLE (aba Visão Geral) ── */
.kpi-tv{{display:grid;grid-template-columns:repeat(6,1fr);gap:14px;padding:18px 24px;background:rgba(0,0,0,0.15);border-bottom:1px solid rgba(255,255,255,0.06)}}
.kpi-tv-card{{background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:14px;padding:18px 20px;backdrop-filter:blur(20px);-webkit-backdrop-filter:blur(20px);display:flex;flex-direction:column;gap:8px;position:relative;overflow:hidden}}
.kpi-tv-card::before{{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:linear-gradient(90deg,transparent,var(--kac),transparent);opacity:.55}}
.kpi-tv-head{{display:flex;align-items:center;gap:8px}}
.kpi-tv-icon{{width:30px;height:30px;border-radius:8px;display:grid;place-items:center;background:rgba(var(--kac-rgb),.12);flex-shrink:0}}
.kpi-tv-icon svg{{width:15px;height:15px;stroke:var(--kac);stroke-width:2;fill:none;stroke-linecap:round;stroke-linejoin:round}}
.kpi-tv-lbl{{font-size:9px;font-weight:700;letter-spacing:.08em;text-transform:uppercase;color:rgba(255,255,255,.42);line-height:1.3}}
.kpi-tv-val{{font-size:26px;font-weight:800;letter-spacing:-.03em;color:#fff;font-variant-numeric:tabular-nums;line-height:1}}
.kpi-tv-sub{{font-size:10px;color:rgba(255,255,255,.3)}}
.kpi-tv-ticket{{font-size:10px;color:rgba(255,255,255,.5);background:rgba(255,255,255,.05);border-radius:5px;padding:3px 7px;margin-top:2px;display:inline-block}}
/* ── RANK COLS (aba Visão Geral) ── */
.rank1st-divider{{height:1px;background:linear-gradient(90deg,transparent,rgba(255,255,255,.07) 20%,rgba(255,255,255,.07) 80%,transparent);margin:0 24px}}
.rank1st-cols{{display:grid;grid-template-columns:repeat(3,1fr);gap:16px;padding:16px 24px 20px}}
.rank1st-col{{background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:14px;overflow:hidden}}
.rank1st-hdr{{padding:14px 18px 12px;border-bottom:1px solid rgba(255,255,255,.06);display:flex;align-items:center;gap:12px}}
.rank1st-hdr-icon{{width:36px;height:36px;border-radius:10px;display:grid;place-items:center;flex-shrink:0}}
.rank1st-hdr-title{{font-size:12px;font-weight:700;letter-spacing:.04em;text-transform:uppercase;color:rgba(255,255,255,.78)}}
.rank1st-hdr-sub{{font-size:10px;color:rgba(255,255,255,.32);margin-top:2px}}
.rank1st-items{{display:flex;flex-direction:column}}
.rank1st-item{{padding:10px 18px;border-bottom:1px solid rgba(255,255,255,.05);display:flex;flex-direction:column;gap:4px;transition:background .15s}}
.rank1st-item:last-child{{border-bottom:none}}
.rank1st-item:hover{{background:rgba(255,255,255,.02)}}
.rank1st-top{{display:flex;align-items:center;gap:10px}}
.rank1st-badge{{width:24px;height:24px;border-radius:7px;display:grid;place-items:center;font-size:11px;font-weight:800;flex-shrink:0;border:1px solid var(--rbd,rgba(255,255,255,.15));background:var(--rbg,rgba(255,255,255,.06));color:var(--rcl,rgba(255,255,255,.45))}}
.rank1st-nome{{font-size:13px;font-weight:700;color:#fff;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
.rank1st-val{{font-size:20px;font-weight:800;letter-spacing:-.03em;font-variant-numeric:tabular-nums;line-height:1;padding-left:34px}}
.rank1st-bar{{display:flex;align-items:center;gap:8px;padding-left:34px}}
.rank1st-bar-bg{{flex:1;height:4px;background:rgba(255,255,255,.07);border-radius:2px;overflow:hidden}}
.rank1st-bar-fill{{height:100%;border-radius:2px}}
.rank1st-bar-pct{{font-size:10px;color:rgba(255,255,255,.35);font-weight:600;min-width:28px}}
/* ── PORTFOLIO SECTION (Velocidade + Cobertura) ── */
.portfolio-section{{display:grid;grid-template-columns:3fr 2fr;gap:16px;padding:0 24px 20px}}
.portfolio-card{{background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:14px;overflow:hidden}}
.portfolio-card-hdr{{padding:14px 18px 12px;border-bottom:1px solid rgba(255,255,255,.06);display:flex;align-items:center;gap:12px}}
.portfolio-card-hdr-icon{{width:36px;height:36px;border-radius:10px;display:grid;place-items:center;flex-shrink:0}}
.portfolio-card-title{{font-size:12px;font-weight:700;letter-spacing:.04em;text-transform:uppercase;color:rgba(255,255,255,.78)}}
.portfolio-card-sub{{font-size:10px;color:rgba(255,255,255,.32);margin-top:2px}}
.growth-list{{display:flex;flex-direction:column}}
.growth-item{{padding:10px 18px;border-bottom:1px solid rgba(255,255,255,.05);display:flex;align-items:center;gap:12px}}
.growth-item:last-child{{border-bottom:none}}
.growth-badge{{width:24px;height:24px;border-radius:7px;display:grid;place-items:center;font-size:11px;font-weight:800;flex-shrink:0;background:rgba(16,185,129,.12);border:1px solid rgba(16,185,129,.35);color:#10b981}}
.growth-info{{flex:1;min-width:0}}
.growth-name{{font-size:12px;font-weight:700;color:#fff;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
.growth-detail{{font-size:10px;color:rgba(255,255,255,.38);margin-top:2px}}
.growth-pct{{font-size:18px;font-weight:800;color:#10b981;letter-spacing:-.02em;font-variant-numeric:tabular-nums;white-space:nowrap}}
.coverage-list{{display:flex;flex-direction:column}}
.coverage-item{{padding:13px 18px;border-bottom:1px solid rgba(255,255,255,.05)}}
.coverage-item:last-child{{border-bottom:none}}
.coverage-head{{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px}}
.coverage-label{{font-size:11px;font-weight:700;color:rgba(255,255,255,.65)}}
.coverage-count{{font-size:10px;color:rgba(255,255,255,.32)}}
.coverage-bar-bg{{height:6px;background:rgba(255,255,255,.07);border-radius:3px;overflow:hidden}}
.coverage-bar-fill{{height:100%;border-radius:3px}}
.coverage-pct{{font-size:22px;font-weight:800;letter-spacing:-.03em;margin-top:6px;font-variant-numeric:tabular-nums}}
/* ── CONTENT ── */
.content{{padding:18px 24px}}
/* ── INSIGHTS ── */
.insights-box{{margin-bottom:20px}}
.insights-title{{font-size:11px;text-transform:uppercase;letter-spacing:.08em;color:rgba(255,255,255,0.42);font-weight:700;margin-bottom:10px;display:flex;align-items:center;gap:8px}}
.insight-list{{display:flex;flex-direction:column;gap:6px}}
.insight{{display:flex;align-items:flex-start;gap:10px;background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.07);border-radius:8px;padding:10px 14px;font-size:13px}}
.insight.positivo{{border-left:3px solid var(--verde)}}
.insight.alerta{{border-left:3px solid var(--amarelo)}}
.insight.critico{{border-left:3px solid var(--vermelho)}}
.insight-icon{{font-size:15px;flex-shrink:0;margin-top:1px}}
.insight-body{{flex:1}}
.insight-txt{{font-weight:500;color:#F0F0F0}}
.insight-det{{font-size:11px;color:rgba(255,255,255,0.36);margin-top:2px}}
/* ── FILTERS ── */
.filters{{display:flex;gap:8px;margin-bottom:14px;align-items:center;flex-wrap:wrap}}
.fbtn{{padding:5px 14px;border-radius:20px;border:1px solid rgba(255,255,255,0.1);background:rgba(255,255,255,0.05);color:rgba(255,255,255,0.52);cursor:pointer;font-size:12px;font-weight:600;transition:.15s;font-family:inherit}}
.fbtn:hover,.fbtn.active{{background:rgba(255,255,255,0.12);color:#F0F0F0;border-color:rgba(255,255,255,0.2)}}
.fbtn.fv.active{{background:rgba(34,197,94,0.14);border-color:var(--verde);color:var(--verde)}}
.fbtn.fa.active{{background:rgba(245,158,11,0.14);border-color:var(--amarelo);color:var(--amarelo)}}
.fbtn.fr.active{{background:rgba(239,68,68,0.14);border-color:var(--vermelho);color:var(--vermelho)}}
.mfbtn{{padding:4px 10px;border-radius:20px;border:1px solid rgba(255,255,255,0.08);background:transparent;color:rgba(255,255,255,0.33);cursor:pointer;font-size:11px;font-weight:600;transition:.15s;font-family:inherit}}
.mfbtn:hover{{background:rgba(255,255,255,0.08);color:rgba(255,255,255,0.72)}}
.mfbtn.active{{background:rgba(59,130,246,0.15);border-color:var(--azul);color:#93c5fd;font-weight:700}}
.search-box{{padding:6px 14px;border-radius:20px;border:1px solid rgba(255,255,255,0.1);background:rgba(255,255,255,0.05);color:#F0F0F0;font-size:13px;width:210px;outline:none;font-family:inherit;transition:.15s}}
.search-box::placeholder{{color:rgba(255,255,255,0.27)}}
.search-box:focus{{border-color:rgba(255,255,255,0.2);background:rgba(255,255,255,0.08)}}
/* ── TABLE ── */
.twrap{{overflow-x:auto;border-radius:8px;border:1px solid rgba(255,255,255,0.08)}}
table{{width:100%;border-collapse:collapse;background:rgba(255,255,255,0.01)}}
th{{background:rgba(0,0,0,0.6);padding:10px 12px;text-align:left;font-size:11px;color:rgba(255,255,255,0.48);text-transform:uppercase;letter-spacing:.08em;font-weight:700;white-space:nowrap;cursor:pointer;user-select:none;border-bottom:1px solid rgba(255,255,255,0.1);transition:.1s}}
th:hover{{color:#F0F0F0}}
th.sa::after{{content:" ↑";color:#F59E0B}}
th.sd::after{{content:" ↓";color:#F59E0B}}
td{{padding:10px 12px;border-bottom:1px solid rgba(255,255,255,0.04);vertical-align:middle;white-space:nowrap}}
tr.erow{{cursor:pointer;transition:.08s}}
tr.erow:hover td{{background:rgba(255,255,255,0.04)}}
tr.expanded td{{background:rgba(255,255,255,0.05)}}
tr.drow td{{padding:0;background:rgba(0,0,0,0.35)}}
.din{{padding:20px 24px;display:grid;grid-template-columns:repeat(auto-fill,minmax(440px,1fr));gap:14px}}
/* ── SCORE BADGE ── */
.sbadge{{display:inline-flex;align-items:center;gap:6px;padding:3px 10px;border-radius:4px;font-weight:600;font-size:12px}}
.sbadge.verde{{background:rgba(34,197,94,0.14);color:#22C55E}}
.sbadge.amarelo{{background:rgba(245,158,11,0.14);color:#F59E0B}}
.sbadge.vermelho{{background:rgba(239,68,68,0.14);color:#EF4444}}
.sbadge.nd{{background:rgba(255,255,255,0.07);color:rgba(255,255,255,0.36)}}
.mtag-critico{{display:inline-block;padding:1px 6px;border-radius:3px;font-size:10px;font-weight:700;background:rgba(239,68,68,0.18);color:#EF4444;margin-left:6px;vertical-align:middle;letter-spacing:.04em}}
.mtag-alerta{{display:inline-block;padding:1px 6px;border-radius:3px;font-size:10px;font-weight:700;background:rgba(245,158,11,0.18);color:#F59E0B;margin-left:6px;vertical-align:middle;letter-spacing:.04em}}
/* ── DOT ── */
.dot{{display:inline-block;width:6px;height:6px;border-radius:9999px;flex-shrink:0}}
.dot.verde{{background:var(--verde)}}
.dot.amarelo{{background:var(--amarelo)}}
.dot.vermelho{{background:var(--vermelho)}}
.dot.nd{{background:rgba(255,255,255,0.32)}}
/* ── EFF ── */
.ecell{{display:flex;flex-direction:column;gap:2px;min-width:72px}}
.etrack{{width:68px;height:4px;background:rgba(255,255,255,0.08);border-radius:2px;overflow:hidden}}
.ebar{{height:100%;border-radius:2px}}
.ebar.verde{{background:var(--verde)}}
.ebar.amarelo{{background:var(--amarelo)}}
.ebar.vermelho{{background:var(--vermelho)}}
.ebar.nd{{background:rgba(255,255,255,0.18)}}
.eval{{font-size:11px;font-weight:600}}
.eval.verde{{color:var(--verde)}}
.eval.amarelo{{color:var(--amarelo)}}
.eval.vermelho{{color:var(--vermelho)}}
.eval.nd{{color:rgba(255,255,255,0.27)}}
/* ── DETAIL CARDS ── */
.obs-banner{{display:flex;align-items:flex-start;gap:10px;padding:11px 14px;border-radius:8px;border:1px solid;margin-bottom:10px;line-height:1.5}}
.obs-tag{{font-size:10px;font-weight:700;padding:2px 8px;border-radius:4px;white-space:nowrap;flex-shrink:0;letter-spacing:.06em;margin-top:1px}}
.obs-txt{{font-size:12px;color:rgba(255,255,255,0.78)}}
.dcard{{background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.08);border-radius:8px;overflow:hidden}}
.dcard-hdr{{padding:9px 14px;background:rgba(0,0,0,0.4);font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.07em;color:rgba(255,255,255,0.42);border-bottom:1px solid rgba(255,255,255,0.06);display:flex;justify-content:space-between;align-items:center}}
/* ── INDICATOR ROW ── */
.irow{{display:grid;grid-template-columns:1fr 72px 72px 86px;gap:6px;padding:8px 14px;border-bottom:1px solid rgba(255,255,255,0.04);align-items:center;font-size:12px}}
.irow:last-child{{border-bottom:none}}
.irow.isub{{padding-left:28px;opacity:.7}}
.iname{{display:flex;align-items:center;gap:6px;overflow:hidden}}
.iname-txt{{white-space:normal;overflow:hidden;display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical;line-height:1.35;word-break:break-word}}
.imeta{{color:rgba(255,255,255,0.3);text-align:right}}
.iexec{{text-align:right;font-weight:600}}
.ieff{{display:flex;flex-direction:column;align-items:flex-end;gap:2px}}
/* ── MATRIZ MENSAL ── */
.mmatriz-wrap{{margin-top:4px;padding:14px}}
.mmatriz-title{{color:rgba(255,255,255,0.4);font-size:10px;text-transform:uppercase;letter-spacing:.07em;font-weight:700;margin-bottom:10px}}
.mmatriz{{border-collapse:collapse;font-size:11px;width:100%}}
.mmatriz thead tr{{background:rgba(0,0,0,0.5)}}
.mth-lbl{{text-align:left;padding:5px 10px;color:rgba(255,255,255,0.36);font-weight:700;background:rgba(0,0,0,0.5);position:sticky;left:0;z-index:2;min-width:140px;white-space:nowrap;border-bottom:2px solid rgba(255,255,255,0.1)}}
.mth-mes{{text-align:center;padding:5px 4px;color:rgba(255,255,255,0.36);font-weight:700;min-width:64px;border-bottom:2px solid rgba(255,255,255,0.1);white-space:nowrap}}
.mtr-grp td{{background:rgba(255,255,255,0.06);padding:4px 10px;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.07em}}
.mtr-grp.g0 td{{color:#93c5fd;border-left:3px solid #3B82F6}}
.mtr-grp.g1 td{{color:#fbbf24;border-left:3px solid #F59E0B}}
.mtr-grp.g2 td{{color:#67e8f9;border-left:3px solid #06b6d4}}
.mtr-grp.g3 td{{color:#86efac;border-left:3px solid #22C55E}}
.mtr-grp.g4 td{{color:#c4b5fd;border-left:3px solid #a78bfa}}
.mtd-lbl{{color:rgba(255,255,255,0.4);padding:4px 10px;background:rgba(0,0,0,0.35);position:sticky;left:0;z-index:1;white-space:nowrap;border-right:1px solid rgba(255,255,255,0.05)}}
.mtd{{padding:3px 3px;text-align:center;border:1px solid rgba(0,0,0,0.3);min-width:64px}}
.mc-p{{color:rgba(255,255,255,0.3);font-size:10px;line-height:1.3}}
.mc-e{{color:#F0F0F0;font-weight:600;font-size:11px;line-height:1.3}}
.mtd.mc-nd{{color:rgba(255,255,255,0.14)}}
.mtd.mc-v{{background:rgba(34,197,94,0.1);border-color:rgba(34,197,94,0.15)}}
.mtd.mc-v .mc-e{{color:#86efac}}
.mtd.mc-a{{background:rgba(245,158,11,0.1);border-color:rgba(245,158,11,0.15)}}
.mtd.mc-a .mc-e{{color:#fde047}}
.mtd.mc-r{{background:rgba(239,68,68,0.1);border-color:rgba(239,68,68,0.15)}}
.mtd.mc-r .mc-e{{color:#fca5a5}}
/* ── TOOLTIP ── */
.tipwrap{{position:relative;display:inline-flex;align-items:center}}
.tip-icon{{width:14px;height:14px;border-radius:50%;background:rgba(255,255,255,0.1);color:rgba(255,255,255,0.4);font-size:9px;font-weight:700;display:inline-flex;align-items:center;justify-content:center;cursor:help;flex-shrink:0;transition:.1s}}
.tip-icon:hover{{background:rgba(255,255,255,0.18);color:#F0F0F0}}
.tipbox{{display:none;position:absolute;left:20px;top:-8px;z-index:999;width:260px;background:rgba(10,10,10,0.96);backdrop-filter:blur(8px);-webkit-backdrop-filter:blur(8px);border:1px solid rgba(255,255,255,0.15);border-radius:8px;padding:12px;box-shadow:0 8px 24px rgba(0,0,0,0.7);font-size:12px;line-height:1.5;color:#F0F0F0}}
.tipwrap:hover .tipbox{{display:block}}
.tipbox strong{{color:#F59E0B;display:block;margin-bottom:4px;font-size:13px}}
.tipbox .tip-why{{color:rgba(255,255,255,0.44);font-size:11px;margin-top:6px;border-top:1px solid rgba(255,255,255,0.08);padding-top:6px}}
.tipbox .tip-alert{{color:#EF4444;font-size:11px;margin-top:4px}}
/* ── DIAGNOSIS ── */
.diag-row{{display:flex;gap:8px;align-items:flex-start;padding:7px 14px;border-bottom:1px solid rgba(255,255,255,0.04);font-size:12px}}
.diag-row:last-child{{border-bottom:none}}
.diag-icon{{font-size:13px;flex-shrink:0;margin-top:1px}}
.diag-txt{{flex:1;line-height:1.4}}
/* ── SPARKLINE ── */
.spark{{display:flex;gap:2px;align-items:flex-end;height:28px;margin-top:4px}}
.spark-b{{width:8px;border-radius:2px 2px 0 0;min-height:2px;background:rgba(255,255,255,0.1);transition:.3s}}
.spark-b.has{{background:#F59E0B}}
/* ── CONSOLIDADO ── */
.cgrid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(300px,1fr));gap:16px}}
.ccard{{background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:8px;padding:18px;box-shadow:0 2px 12px rgba(0,0,0,0.25)}}
.ccard h3{{font-size:11px;text-transform:uppercase;letter-spacing:.06em;color:rgba(255,255,255,0.4);margin-bottom:14px;font-weight:700}}
/* ── BAR CHART ── */
.bchart{{display:flex;flex-direction:column;gap:8px}}
.bitem{{display:grid;grid-template-columns:130px 1fr 70px;gap:8px;align-items:center;font-size:11px}}
.btrack{{height:18px;background:rgba(255,255,255,0.06);border-radius:4px;overflow:hidden;cursor:help}}
.bfill{{height:100%;border-radius:4px;background:#F59E0B;transition:.5s}}
.bval{{color:rgba(255,255,255,0.4);text-align:right;white-space:nowrap}}
.blabel{{overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:#F0F0F0}}
/* ── MONTHLY CHART ── */
.mchart{{display:flex;gap:4px;align-items:flex-end;height:72px;padding-top:8px}}
.mbar-w{{flex:1;display:flex;flex-direction:column;align-items:center;gap:4px}}
.mbar{{width:100%;border-radius:3px 3px 0 0;min-height:2px;transition:.4s;cursor:help}}
.mlabel{{font-size:9px;color:rgba(255,255,255,0.3)}}
/* ── DIST ── */
.distrow{{display:flex;gap:16px;align-items:center;justify-content:center;flex-wrap:wrap;margin:10px 0}}
.distitem{{display:flex;flex-direction:column;align-items:center;gap:3px}}
.distnum{{font-size:30px;font-weight:800}}
.distlabel{{font-size:10px;color:rgba(255,255,255,0.4);text-transform:uppercase;letter-spacing:.05em}}
/* ── RANK LIST ── */
.rlist{{display:flex;flex-direction:column;gap:5px}}
.ritem{{display:grid;grid-template-columns:22px 1fr auto;gap:8px;align-items:center;padding:7px 10px;background:rgba(0,0,0,0.3);border-radius:6px;font-size:12px}}
.rpos{{color:rgba(255,255,255,0.35);font-weight:700;text-align:center;font-size:11px}}
.rname{{overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
.rval{{font-weight:700;color:#F59E0B;white-space:nowrap}}
/* ── SUCCESS CASE ── */
.case-card{{background:rgba(0,0,0,0.3);border:1px solid rgba(34,197,94,0.2);border-left:3px solid var(--verde);border-radius:8px;padding:12px 14px;margin-bottom:8px}}
.case-card h4{{color:var(--verde);font-size:12px;font-weight:700;margin-bottom:4px}}
.case-card p{{color:rgba(255,255,255,0.46);font-size:11px;line-height:1.5}}
.attn-card{{background:rgba(0,0,0,0.3);border:1px solid rgba(239,68,68,0.2);border-left:3px solid var(--vermelho);border-radius:8px;padding:12px 14px;margin-bottom:8px}}
.attn-card h4{{color:var(--vermelho);font-size:12px;font-weight:700;margin-bottom:4px}}
.attn-card p{{color:rgba(255,255,255,0.46);font-size:11px;line-height:1.5}}
/* ── LEGENDA SECTION ── */
.leg-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:10px}}
.leg-card{{background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:8px;padding:14px}}
.leg-card .leg-titulo{{font-weight:700;font-size:13px;color:#F0F0F0;margin-bottom:4px;display:flex;align-items:center;gap:6px}}
.leg-card .leg-tipo{{background:rgba(255,255,255,0.1);color:rgba(255,255,255,0.46);font-size:10px;padding:1px 6px;border-radius:4px;font-weight:600}}
.leg-card .leg-desc{{font-size:12px;color:rgba(255,255,255,0.5);line-height:1.5;margin-bottom:6px}}
.leg-card .leg-impacto{{font-size:11px;color:rgba(255,255,255,0.33);border-top:1px solid rgba(255,255,255,0.06);padding-top:6px}}
.leg-card .leg-alerta{{font-size:11px;color:#F59E0B;margin-top:4px}}
/* ── MISC ── */
.tag{{display:inline-block;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600;letter-spacing:.03em}}
.tag.positivo{{background:rgba(34,197,94,0.14);color:#22C55E}}
.tag.alerta{{background:rgba(245,158,11,0.14);color:#F59E0B}}
.tag.critico{{background:rgba(239,68,68,0.14);color:#EF4444}}
.nodata{{color:rgba(255,255,255,0.27);font-style:italic;padding:16px;text-align:center;font-size:13px}}
.hidden{{display:none!important}}
.section-title{{font-size:13px;font-weight:700;color:#F0F0F0;margin-bottom:14px;display:flex;align-items:center;gap:8px}}
.section-title .badge{{background:rgba(255,255,255,0.1);color:rgba(255,255,255,0.4);font-size:11px;padding:1px 8px;border-radius:10px}}
</style>
</head>
<body>

<header>
  <div class="logo">
    {logo_html}
    <span class="logo-sep">·</span>
    <span class="logo-title">Planos de Negócio 2026</span>
  </div>
  <div class="header-meta">Gerado em {now} &nbsp;·&nbsp; {n} empresas</div>
  <button class="sync-btn" id="syncBtn" onclick="triggerSync()">⟳ Sincronizar</button>
  <button class="sync-cfg" onclick="openSyncModal()" title="Configurar GitHub">⚙</button>
</header>

<div id="syncModal">
  <div class="smbox">
    <h3>Configurar Sincronização via GitHub</h3>
    <p>Dispara o workflow que baixa todas as planilhas e atualiza o dashboard.<br>
    O token precisa ter permissão <strong>workflow</strong>. Dados salvos apenas no seu navegador.</p>
    <label class="smlabel">Usuário GitHub (owner)</label>
    <input class="sminput" id="syncOwner" placeholder="ex: isael-solarz" type="text">
    <label class="smlabel">Nome do repositório</label>
    <input class="sminput" id="syncRepo" placeholder="ex: solarz-dashboard" type="text">
    <label class="smlabel">Personal Access Token (PAT)</label>
    <input class="sminput" id="syncPat" placeholder="ghp_xxxxxxxxxxxx" type="password">
    <div class="smactions">
      <button class="smcancel" onclick="closeSyncModal()">Cancelar</button>
      <button class="smsave" onclick="saveSyncConfig()">Salvar e Sincronizar</button>
    </div>
  </div>
</div>

<div class="tabs" id="tabBar">
  <div class="tab active"   onclick="switchTab('geral')">Visão Geral & Insights</div>
  <div class="tab"          onclick="switchTab('ranking')">Rankings Detalhados</div>
  <div class="tab"          onclick="switchTab('construindo')">Inteligência Operacional</div>
  <div class="tab"          onclick="switchTab('legendas')">Guia de Indicadores</div>
</div>

<!-- ═══════════════ PANE: GERAL ═══════════════ -->
<div id="pane-geral">
  <div class="kpi-tv" id="kpiGrid"></div>
  <div class="rank1st-divider"></div>
  <div class="rank1st-cols" id="rank1stGrid"></div>
  <div class="rank1st-divider"></div>
  <div class="portfolio-section" id="portfolioGrid"></div>
  <div class="content">
    <div class="insights-box">
      <div class="insights-title">
        <svg width="14" height="14" viewBox="0 0 14 14" fill="none"><circle cx="7" cy="7" r="6" stroke="#f59e0b" stroke-width="1.5"/><path d="M7 6v4M7 4.5v.5" stroke="#f59e0b" stroke-width="1.5" stroke-linecap="round"/></svg>
        Insights Automáticos
      </div>
      <div class="insight-list" id="insightsList"></div>
    </div>
  </div>
</div>

<!-- ═══════════════ PANE: RANKING ═══════════════ -->
<div id="pane-ranking" class="hidden">
  <div class="content">
    <div class="filters">
      <button class="fbtn active" onclick="setFilter('todos',this)">Todas</button>
      <button class="fbtn fr"     onclick="setFilter('critico',this)">CRÍTICO</button>
      <button class="fbtn fa"     onclick="setFilter('alerta',this)">ALERTA</button>
      <button class="fbtn"        onclick="setFilter('sem_tag',this)">Sem Tag</button>
      <input type="text" class="search-box" id="searchBox" placeholder="Buscar empresa…" oninput="renderTable()">
      <span id="tableCount" style="color:rgba(255,255,255,0.38);font-size:12px;margin-left:auto"></span>
    </div>
    <div class="filters" style="gap:6px;padding-top:6px;border-top:1px solid rgba(255,255,255,0.07)">
      <span style="color:rgba(255,255,255,0.4);font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.06em;align-self:center">Mês:</span>
      <button class="mfbtn active" onclick="setMes(0,this)">Todo o período</button>
      <button class="mfbtn" onclick="setMes(1,this)">Jan</button>
      <button class="mfbtn" onclick="setMes(2,this)">Fev</button>
      <button class="mfbtn" onclick="setMes(3,this)">Mar</button>
      <button class="mfbtn" onclick="setMes(4,this)">Abr</button>
      <button class="mfbtn" onclick="setMes(5,this)">Mai</button>
      <button class="mfbtn" onclick="setMes(6,this)">Jun</button>
      <button class="mfbtn" onclick="setMes(7,this)">Jul</button>
      <button class="mfbtn" onclick="setMes(8,this)">Ago</button>
      <button class="mfbtn" onclick="setMes(9,this)">Set</button>
      <button class="mfbtn" onclick="setMes(10,this)">Out</button>
      <button class="mfbtn" onclick="setMes(11,this)">Nov</button>
      <button class="mfbtn" onclick="setMes(12,this)">Dez</button>
      <span id="mesLabel" style="margin-left:6px;color:#F59E0B;font-size:11px;font-weight:700;align-self:center"></span>
    </div>
    <div id="fonteBanner" style="display:none;margin-bottom:10px;padding:8px 14px;background:rgba(59,130,246,0.08);border:1px solid rgba(59,130,246,0.18);border-radius:6px;font-size:11px;color:rgba(255,255,255,0.55);line-height:1.6">
      <span id="fonteTexto"></span>
    </div>
    <div class="twrap">
      <table id="mainTable">
        <thead>
          <tr>
            <th onclick="sortBy('rank')">#</th>
            <th onclick="sortBy('nome')">Empresa</th>
            <th onclick="sortBy('score')">Score ▾</th>
            <th onclick="sortBy('mrr_acc_e')">MRR Acumulado</th>
            <th onclick="sortBy('mrr_vend_e')">MRR Vendido</th>
            <th onclick="sortBy('carteira_e')">Carteira</th>
            <th onclick="sortBy('conv_acomp_e')">Conv. Acomp</th>
            <th onclick="sortBy('conv_srv_e')">Conv. Serviço</th>
            <th onclick="sortBy('conv_seg_e')">Conv. Seguro</th>
            <th onclick="sortBy('vol_abord_e')">Vol. Abordado</th>
            <th onclick="sortBy('seguro_e')">Seguro</th>
            <th onclick="sortBy('servicos_e')">Serviços</th>
            <th onclick="sortBy('indicacoes_e')">Indicações</th>
            <th onclick="sortBy('agentes_e')">Agentes PV</th>
          </tr>
        </thead>
        <tbody id="tableBody"></tbody>
      </table>
    </div>
  </div>
</div>

<!-- ═══════════════ PANE: CONSTRUINDO ═══════════════ -->
<div id="pane-construindo" class="hidden">
  <div class="content">
    <div class="io-topbar">
      <div class="section-title" style="margin:0">Análise da Carteira</div>
      <div class="io-filter-wrap">
        <span class="io-filter-label">Consultor:</span>
        <input class="io-filter-input" id="io-consultor-input" list="io-consultores-list"
               placeholder="Todos os consultores" oninput="onConsultorInput(this.value)">
        <datalist id="io-consultores-list"></datalist>
        <button class="io-clear-btn" onclick="clearConsultor()" title="Limpar filtro">✕</button>
      </div>
    </div>
    <!-- Linha 1: Status tag + Perf consultor + Receita fonte -->
    <div class="io-grid-3">
      <div class="io-card">
        <h3>Distribuição de Status</h3>
        <div class="io-donut-wrap">
          <div class="io-donut" id="io-donut-status">
            <div class="io-donut-hole">
              <div class="io-donut-num" id="io-donut-status-num">—</div>
              <div class="io-donut-lbl">EMPRESAS</div>
            </div>
          </div>
          <div id="io-donut-status-legend" class="io-legend"></div>
        </div>
      </div>
      <div class="io-card">
        <h3>Performance por Consultor (score médio)</h3>
        <div id="io-perf-consultor" class="io-hbar-list"></div>
      </div>
      <div class="io-card">
        <h3>Receita por Fonte (período completo)</h3>
        <div class="io-donut-wrap">
          <div class="io-donut" id="io-donut-receita">
            <div class="io-donut-hole">
              <div class="io-donut-num" id="io-donut-receita-num">—</div>
              <div class="io-donut-lbl">TOTAL</div>
            </div>
          </div>
          <div id="io-donut-receita-legend" class="io-legend"></div>
        </div>
      </div>
    </div>
    <!-- Linha 2: Dist score + Faixas MRR + Tempo vs Score -->
    <div class="io-grid-3">
      <div class="io-card">
        <h3>Distribuição de Score</h3>
        <div id="io-dist-score" class="io-vbar-list"></div>
      </div>
      <div class="io-card">
        <h3>Faixas de MRR Acumulado</h3>
        <div id="io-faixas-mrr" class="io-vbar-list"></div>
      </div>
      <div class="io-card">
        <h3>Tempo na Consultoria vs Score Médio</h3>
        <div id="io-tempo-score" class="io-vbar-list"></div>
      </div>
    </div>
    <!-- Linha 3: Agentes PV + abordagem -->
    <div class="io-card-wide">
      <h3>Agentes PV — Volume Abordado <span style="color:rgba(255,255,255,0.35);font-size:10px;font-weight:500;text-transform:none;letter-spacing:0">linha laranja = limite 50 abordagens</span></h3>
      <div id="io-agente-abord" class="io-hbar-list io-hbar-scroll"></div>
    </div>
    <!-- Painel de Observações CRÍTICO/ALERTA -->
    <div class="io-card-wide" id="io-obs-panel">
      <h3>Análise de Observações — CRÍTICO &amp; ALERTA <span id="io-obs-subtitle" style="color:rgba(255,255,255,0.35);font-size:10px;font-weight:500;text-transform:none;letter-spacing:0"></span></h3>
      <div id="io-obs-content"></div>
    </div>
  </div>
  <!-- Matriz detalhada -->
  <div class="content" style="padding-top:0">
    <div class="section-title" style="margin-bottom:12px">Matriz Detalhada</div>
    <div class="filters" style="gap:6px;padding-bottom:10px">
      <span style="color:rgba(255,255,255,0.4);font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.06em;align-self:center">Mês:</span>
      <button class="mfbtn active" id="io-mes-0" onclick="setMesIO(0,this)">Todo o período</button>
      <button class="mfbtn" onclick="setMesIO(1,this)">Jan</button>
      <button class="mfbtn" onclick="setMesIO(2,this)">Fev</button>
      <button class="mfbtn" onclick="setMesIO(3,this)">Mar</button>
      <button class="mfbtn" onclick="setMesIO(4,this)">Abr</button>
      <button class="mfbtn" onclick="setMesIO(5,this)">Mai</button>
      <button class="mfbtn" onclick="setMesIO(6,this)">Jun</button>
      <button class="mfbtn" onclick="setMesIO(7,this)">Jul</button>
      <button class="mfbtn" onclick="setMesIO(8,this)">Ago</button>
      <button class="mfbtn" onclick="setMesIO(9,this)">Set</button>
      <button class="mfbtn" onclick="setMesIO(10,this)">Out</button>
      <button class="mfbtn" onclick="setMesIO(11,this)">Nov</button>
      <button class="mfbtn" onclick="setMesIO(12,this)">Dez</button>
      <span id="io-tableCount" style="color:rgba(255,255,255,0.38);font-size:12px;margin-left:auto"></span>
    </div>
    <div class="twrap">
      <table>
        <thead>
          <tr>
            <th onclick="sortByIO('rank')">#</th>
            <th onclick="sortByIO('nome')">Empresa</th>
            <th onclick="sortByIO('score')">Score ▾</th>
            <th onclick="sortByIO('mrr_acc_e')">MRR Acumulado</th>
            <th onclick="sortByIO('mrr_vend_e')">MRR Vendido</th>
            <th onclick="sortByIO('carteira_e')">Carteira</th>
            <th onclick="sortByIO('conv_acomp_e')">Conv. Acomp</th>
            <th onclick="sortByIO('conv_srv_e')">Conv. Serviço</th>
            <th onclick="sortByIO('conv_seg_e')">Conv. Seguro</th>
            <th onclick="sortByIO('vol_abord_e')">Vol. Abordado</th>
            <th onclick="sortByIO('seguro_e')">Seguro</th>
            <th onclick="sortByIO('servicos_e')">Serviços</th>
            <th onclick="sortByIO('indicacoes_e')">Indicações</th>
            <th onclick="sortByIO('agentes_e')">Agentes PV</th>
          </tr>
        </thead>
        <tbody id="io-tableBody"></tbody>
      </table>
    </div>
  </div>
  <!-- Seção original mantida -->
  <div class="content" id="construindoContent"></div>
</div>

<!-- ═══════════════ PANE: LEGENDAS ═══════════════ -->
<div id="pane-legendas" class="hidden">
  <div class="content">
    <div class="section-title">Guia Completo de Indicadores</div>
    <div class="leg-grid" id="legendaGrid"></div>
  </div>
</div>

<script>
const DADOS = {data_json};
const CONSOL = {consol_json};
const INSIGHTS_DATA = {insights_json};
const MESES = {{1:'Jan',2:'Fev',3:'Mar',4:'Abr',5:'Mai',6:'Jun',7:'Jul',8:'Ago',9:'Set',10:'Out',11:'Nov',12:'Dez'}};

// ── LEGENDAS DOS INDICADORES ────────────────────────────────────────────────
const LEGENDAS = {{
  'mrr acumulado':{{
    titulo:'MRR Acumulado', tipo:'R$',
    desc:'Receita Recorrente Mensal total acumulada — soma de todos os contratos ativos pagantes desde o início da operação. O valor de Maio/2026 já inclui Janeiro a Maio.',
    impacto:'Principal métrica da operação de pós-vendas. Um MRR crescente indica construção de base sólida. É o indicador que define se o pós-vendas está gerando valor sustentável.',
    alerta:'⚠️ Não somar os meses — use apenas o valor do mês mais recente. Pode divergir da soma do MRR Vendido por efeito de Churn.'
  }},
  'mrr vendido':{{
    titulo:'MRR Vendido', tipo:'R$',
    desc:'Receita recorrente NOVA gerada no período, proveniente exclusivamente das novas vendas do plano de monitoramento. Não é acumulativo.',
    impacto:'Mede o ritmo de crescimento mês a mês. Zero indica que a equipe não está adicionando novos contratos à base — alerta de estagnação.',
    alerta:'⚠️ Valor zero por dois meses = urgente. Verificar se a equipe está realizando abordagem ativa.'
  }},
  'mrr churn':{{
    titulo:'MRR Churn', tipo:'R$',
    desc:'Receita recorrente perdida por cancelamentos de contratos no período. Representa o "vazamento" da base.',
    impacto:'Alto Churn corrói o MRR Acumulado mesmo quando há boas vendas novas. MRR Líquido = MRR Vendido − MRR Churn.',
    alerta:'ℹ️ Exibido sem alerta automático — o contexto determina se o valor é preocupante (ex: cancelamentos por cliente inadimplente vs insatisfeito).'
  }},
  'volume abordado':{{
    titulo:'Volume Abordado', tipo:'Qtd',
    desc:'Quantidade de clientes da carteira que foram efetivamente contatados pela equipe de pós-vendas no período.',
    impacto:'Base de toda a operação. Sem abordagem não há vendas, indicações nem relacionamento. É o motor que alimenta todos os demais indicadores.',
    alerta:'⚠️ Zero = paralisia operacional. Meta mínima da metodologia SolarZ: 80 abordagens/mês (somando todos os produtos).'
  }},
  'volume vendido':{{
    titulo:'Volume Vendido', tipo:'Qtd',
    desc:'Quantidade de contratos de plano de monitoramento efetivamente vendidos e assinados no período.',
    impacto:'Conversão da abordagem em contratos. Cruzar com Volume Abordado para calcular a Taxa de Conversão real.',
    alerta:'⚠️ Se há Volume Abordado mas Volume Vendido = 0, o problema é de conversão (pitch, treinamento, proposta).'
  }},
  'taxa de conversao':{{
    titulo:'Taxa de Conversão', tipo:'%',
    desc:'Percentual de clientes abordados que efetivamente compraram: Volume Vendido ÷ Volume Abordado × 100.',
    impacto:'Eficiência comercial da equipe. Uma baixa conversão com alto volume abordado indica problema no argumento de venda. Alta conversão com volume baixo pode ser normal em carteiras pequenas.',
    alerta:'ℹ️ Valores acima de 100% podem indicar vendas para clientes não contabilizados no Volume Abordado.'
  }},
  'ticket medio':{{
    titulo:'Ticket Médio', tipo:'R$',
    desc:'Valor médio por contrato vendido no período: MRR Vendido ÷ Volume Vendido.',
    impacto:'Reflete a qualidade das vendas. Ticket abaixo do esperado pode indicar venda de planos mais baratos ou descontos excessivos.',
    alerta:'ℹ️ Comparar com o ticket médio histórico da carteira atual e com a precificação padrão.'
  }},
  'quantidade de agente':{{
    titulo:'Agentes de Pós-Vendas', tipo:'Qtd',
    desc:'Número de colaboradores da integradora atuando exclusivamente no pós-vendas no período (gestores + operadores).',
    impacto:'Capacidade operacional. Define o limite máximo de abordagens possíveis.',
    alerta:'⚠️ Zero = ninguém operando o pós-vendas. Dois meses consecutivos = alerta crítico de reestruturação da equipe.'
  }},
  'safra de negocio':{{
    titulo:'Safra de Negócio', tipo:'Qtd',
    desc:'Volume de clientes em processo ativo de negociação no período — o pipeline de vendas do pós-vendas.',
    impacto:'Indica o futuro: se a Safra está vazia hoje, as vendas dos próximos meses estão em risco. Um pipeline saudável garante previsibilidade.',
    alerta:'⚠️ Zero Safra + Zero MRR Vendido = duplo alerta. Verificar se a equipe está prospectando ativamente.'
  }},
  'clientes em carteira':{{
    titulo:'Clientes em Carteira', tipo:'Qtd',
    desc:'Total de clientes ativos sendo gerenciados pelo pós-vendas da integradora. É um indicador de estoque — não acumulativo.',
    impacto:'Define a responsabilidade e capacidade de atendimento do consultor. Crescimento indica expansão; queda indica churn ou realocação.',
    alerta:'ℹ️ Carteira abaixo da meta pode indicar que novas integradoras ainda estão em fase de onboarding.'
  }},
  'volume indicacoes':{{
    titulo:'Volume de Indicações', tipo:'Qtd',
    desc:'Quantidade de leads indicados pelos clientes da carteira para novos contratos de monitoramento.',
    impacto:'Canal orgânico e de menor custo de aquisição. Clientes engajados que indicam sinalizam satisfação com o serviço.',
    alerta:'⚠️ Zero = clientes não estão sendo ativamente solicitados a indicar. Revisar processo de solicitação de indicações.'
  }},
  'vendido_seguro':{{
    titulo:'Seguro Vendido', tipo:'R$',
    desc:'Valor total de apólices de seguro solar vendidas para os clientes da carteira no período.',
    impacto:'Linha de receita complementar ao MRR. Produto de alta margem que diversifica a geração de valor e aumenta o ticket por cliente.',
    alerta:'⚠️ Zero = produto não está sendo ofertado. Verificar se a equipe recebeu treinamento de venda do seguro solar.'
  }},
  'vendido_servicos':{{
    titulo:'Serviços Vendidos', tipo:'R$',
    desc:'Valor total de serviços avulsos (limpeza, manutenção, expansão de usina) vendidos para a base no período.',
    impacto:'Receita não recorrente que complementa o MRR. Alta eficiência em Serviços indica carteira engajada e com demanda ativa.',
    alerta:'ℹ️ Sazonalidade esperada — serviços tendem a ser mais fortes em períodos específicos do ano.'
  }},
}};

// ── HELPERS ─────────────────────────────────────────────────────────────────
function fmtM(v){{
  if(v==null)return'—';
  const a=Math.abs(v);
  if(a>=1e6)return'R$ '+(v/1e6).toFixed(1).replace('.',',')+' Milhões';
  if(a>=1e5)return'R$ '+Math.round(v/1000)+' mil';
  if(a>=1000)return'R$ '+(v/1000).toFixed(1).replace('.',',')+' mil';
  return'R$ '+Math.round(v).toLocaleString('pt-BR');
}}
function fmtN(v){{if(v==null)return'—';return Math.round(v).toLocaleString('pt-BR')}}
function fmtP(v){{if(v==null)return'—';return v.toFixed(1)+'%'}}
function sColor(s){{return{{verde:'#22C55E',amarelo:'#F59E0B',vermelho:'#EF4444',nd:'rgba(255,255,255,0.32)'}}[s]||'rgba(255,255,255,0.32)'}}
function convColor(pct){{
  if(pct==null)return'rgba(255,255,255,0.32)';
  if(pct>=80)return'#22C55E';
  if(pct>=60)return'#F59E0B';
  return'#EF4444';
}}
function isMonetary(n){{
  const nm=n.toLowerCase();
  return nm.includes('mrr')||nm.includes('valor')||nm.includes('ticket');
}}
function isPercent(n){{
  return n.toLowerCase().includes('taxa de convers');
}}
function fmtVal(n,v){{
  if(v==null)return'—';
  return isMonetary(n)?fmtM(v):fmtN(v);
}}
function tipKey(indName){{
  const nm=indName.toLowerCase();
  if(nm.includes('mrr acumulado'))return'mrr acumulado';
  if(nm.includes('mrr vendido'))return'mrr vendido';
  if(nm.includes('mrr churn')||nm.includes('churn')&&nm.includes('mrr'))return'mrr churn';
  if(nm.includes('volume abordado'))return'volume abordado';
  if(nm.includes('volume vendido')||nm.includes('volume vendas'))return'volume vendido';
  if(nm.includes('taxa de convers'))return'taxa de conversao';
  if(nm.includes('ticket'))return'ticket medio';
  if(nm.includes('agente')||nm.includes('operador'))return'quantidade de agente';
  if(nm.includes('safra'))return'safra de negocio';
  if(nm.includes('carteira'))return'clientes em carteira';
  if(nm.includes('indica'))return'volume indicacoes';
  return null;
}}

function effCell(eff,s){{
  if(eff==null)return'<span style="color:rgba(255,255,255,0.27)">—</span>';
  const w=Math.min(Math.abs(eff),150)/1.5;
  return`<div class="ecell"><span class="eval ${{s}}">${{fmtP(eff)}}</span><div class="etrack"><div class="ebar ${{s}}" style="width:${{w}}%"></div></div></div>`;
}}
function scoreBadge(score,status){{
  if(score==null)return'<span class="sbadge nd">—</span>';
  return`<span class="sbadge ${{status}}">${{score.toFixed(0)}}%</span>`;
}}
function diagIcon(tipo){{
  return{{positivo:'✅',alerta:'⚠️',critico:'🔴',neutro:'ℹ️'}}[tipo]||'•';
}}

// ── SPARK ────────────────────────────────────────────────────────────────────
function sparkline(mensal){{
  const vals=[1,2,3,4,5,6,7,8,9,10,11,12].map(m=>{{
    const d=mensal[String(m)];
    return d&&d.e!=null?d.e:null;
  }});
  const nonNull=vals.filter(v=>v!==null);
  if(!nonNull.length)return'';
  const mx=Math.max(...nonNull,1);
  const bars=vals.map(v=>{{
    if(v==null)return`<div class="spark-b" style="height:2px"></div>`;
    const h=Math.max(Math.round((v/mx)*26),2);
    return`<div class="spark-b has" style="height:${{h}}px" title="${{fmtM(v)}} / ${{fmtN(v)}}"></div>`;
  }});
  return`<div class="spark">${{bars.join('')}}</div>`;
}}

// ── TOOLTIP HTML ─────────────────────────────────────────────────────────────
function tipHTML(indName){{
  const key=tipKey(indName);
  if(!key)return'';
  const l=LEGENDAS[key];
  if(!l)return'';
  return`<div class="tipwrap">
    <span class="tip-icon">?</span>
    <div class="tipbox">
      <strong>${{l.titulo}} <span style="background:rgba(255,255,255,0.1);color:rgba(255,255,255,0.46);font-size:10px;padding:1px 5px;border-radius:3px;font-weight:600">${{l.tipo}}</span></strong>
      ${{l.desc}}
      <div class="tip-why">💡 ${{l.impacto}}</div>
      ${{l.alerta?`<div class="tip-alert">${{l.alerta}}</div>`:''}}
    </div>
  </div>`;
}}

// ── TABS ─────────────────────────────────────────────────────────────────────
let activeTab='geral';
function switchTab(t){{
  activeTab=t;
  ['geral','ranking','construindo','legendas'].forEach((id,i)=>{{
    document.getElementById('pane-'+id).classList.toggle('hidden',id!==t);
    document.querySelectorAll('.tab')[i].classList.toggle('active',id===t);
  }});
  if(t==='construindo'){{populateConsultorList();renderConstruindo();renderAllIO();}}
  if(t==='legendas')renderLegendas();
}}

// ── SUMMARY ──────────────────────────────────────────────────────────────────
function renderSummary(){{
  const c=CONSOL;
  const defs=[
    {{l:'MRR Total Construído',s:'MRR Acumulado Executado',v:fmtM(c.total_mrr),ac:'#f59e0b',rgb:'245,158,11',
      ic:'<polyline points="23 6 13.5 15.5 8.5 10.5 1 18"/><polyline points="17 6 23 6 23 12"/>'}},
    {{l:'MRR Vendido',s:'Novas vendas no período',v:fmtM(c.total_mrr_vend),ac:'#10b981',rgb:'16,185,129',
      ic:'<line x1="12" y1="1" x2="12" y2="23"/><path d="M17 5H9.5a3.5 3.5 0 0 0 0 7h5a3.5 3.5 0 0 1 0 7H6"/>'}},
    {{l:'Clientes Carteirizados',s:'Total gerenciado',v:fmtN(c.total_carteira),ac:'#3b82f6',rgb:'59,130,246',
      ic:'<path d="M17 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"/><circle cx="9" cy="7" r="4"/><path d="M23 21v-2a4 4 0 0 0-3-3.87"/><path d="M16 3.13a4 4 0 0 1 0 7.75"/>'}},
    {{l:'Indicações Geradas',s:'Volume total de indicações',v:fmtN(c.total_indicacoes),ac:'#a855f7',rgb:'168,85,247',
      ic:'<path d="M22 16.92v3a2 2 0 0 1-2.18 2 19.79 19.79 0 0 1-8.63-3.07A19.5 19.5 0 0 1 4.69 13 19.79 19.79 0 0 1 1.61 4.4 2 2 0 0 1 3.6 2.18h3a2 2 0 0 1 2 1.72c.127.96.361 1.903.7 2.81a2 2 0 0 1-.45 2.11L7.91 9a16 16 0 0 0 6.18 6.18l.95-.95a2 2 0 0 1 2.11-.45c.907.339 1.85.573 2.81.7A2 2 0 0 1 22 16.92z"/>'}},
    {{l:'Seguro Vendido',s:'Acumulado no período',v:fmtM(c.total_seguro),ac:'#06b6d4',rgb:'6,182,212',
      ticket:c.ticket_seguro!=null?`Ticket médio: ${{fmtM(c.ticket_seguro)}} · ${{fmtN(c.qtd_seguro)}} vendas`:null,
      ic:'<path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"/>'}},
    {{l:'Serviços Vendidos',s:'Acumulado no período',v:fmtM(c.total_servicos),ac:'#f97316',rgb:'249,115,22',
      ticket:c.ticket_servicos!=null?`Ticket médio: ${{fmtM(c.ticket_servicos)}} · ${{fmtN(c.qtd_servicos)}} serviços`:null,
      ic:'<circle cx="12" cy="12" r="3"/><path d="M19.07 4.93a10 10 0 0 1 0 14.14M4.93 4.93a10 10 0 0 0 0 14.14"/>'}}
  ];
  document.getElementById('kpiGrid').innerHTML=defs.map(d=>`
    <div class="kpi-tv-card" style="--kac:${{d.ac}};--kac-rgb:${{d.rgb}};box-shadow:0 0 28px rgba(${{d.rgb}},.09)">
      <div class="kpi-tv-head">
        <div class="kpi-tv-icon"><svg viewBox="0 0 24 24">${{d.ic}}</svg></div>
        <div class="kpi-tv-lbl">${{d.l}}</div>
      </div>
      <div class="kpi-tv-val">${{d.v}}</div>
      <div class="kpi-tv-sub">${{d.s}}</div>
      ${{d.ticket?`<div class="kpi-tv-ticket">${{d.ticket}}</div>`:''}}
    </div>`).join('');
}}

// ── INSIGHTS ─────────────────────────────────────────────────────────────────
function renderInsights(){{
  const icons={{positivo:'✅',alerta:'⚠️',critico:'🔴'}};
  document.getElementById('insightsList').innerHTML=INSIGHTS_DATA.map(i=>`
    <div class="insight ${{i.tipo}}">
      <span class="insight-icon">${{icons[i.tipo]||'ℹ️'}}</span>
      <div class="insight-body">
        <div class="insight-txt">${{i.texto}}</div>
        ${{i.detalhe?`<div class="insight-det">${{i.detalhe}}</div>`:''}}
      </div>
    </div>`).join('') || '<div class="nodata">Nenhum insight disponível</div>';
}}

// ── RANKINGS (aba Visão Geral) ────────────────────────────────────────────────
function renderRankings1st(){{
  const RDEFS=[
    {{t:'MRR Atual',s:'Base construída · último mês preenchido',ac:'#f59e0b',rgb:'245,158,11',
      fn:d=>lastMonthExec(d,'acompanhamento','mrr acumulado'),fmt:fmtM,
      ic:'<rect x="2" y="3" width="20" height="14" rx="2"/><path d="M8 21h8M12 17v4"/><polyline points="7 10 10 13 14 9 17 11"/>'}},
    {{t:'Receita com Seguro',s:'Por empresa · período completo',ac:'#06b6d4',rgb:'6,182,212',
      fn:d=>sumMonthly(d,'seguro','vendido','valor vendido'),fmt:fmtM,
      ic:'<path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"/>'}},
    {{t:'Receita com Venda de Serviços',s:'Por empresa · período completo',ac:'#f97316',rgb:'249,115,22',
      fn:d=>sumMonthly(d,'servicos','valor vendido','vendido'),fmt:fmtM,
      ic:'<circle cx="12" cy="12" r="3"/><path d="M19.07 4.93a10 10 0 0 1 0 14.14M4.93 4.93a10 10 0 0 0 0 14.14"/>'}}
  ];
  const MEDALS=[
    {{bg:'rgba(245,158,11,.15)',bd:'rgba(245,158,11,.45)',cl:'#f59e0b'}},
    {{bg:'rgba(148,163,184,.15)',bd:'rgba(148,163,184,.45)',cl:'#94a3b8'}},
    {{bg:'rgba(180,83,9,.15)',bd:'rgba(180,83,9,.45)',cl:'#b45309'}}
  ];
  const grid=document.getElementById('rank1stGrid');
  if(!grid)return;
  grid.innerHTML='';
  RDEFS.forEach(col=>{{
    const top=[...DADOS]
      .map(d=>({{nome:d.nome,val:col.fn(d)}}))
      .filter(x=>x.val!=null&&x.val>0)
      .sort((a,b)=>b.val-a.val).slice(0,5);
    const maxVal=top.length>0?top[0].val:1;
    const colEl=document.createElement('div');
    colEl.className='rank1st-col';
    let h=`
      <div class="rank1st-hdr">
        <div class="rank1st-hdr-icon" style="background:rgba(${{col.rgb}},.12)">
          <svg viewBox="0 0 24 24" width="18" height="18" stroke="${{col.ac}}" stroke-width="2" fill="none" stroke-linecap="round" stroke-linejoin="round">${{col.ic}}</svg>
        </div>
        <div>
          <div class="rank1st-hdr-title">${{col.t}}</div>
          <div class="rank1st-hdr-sub">${{col.s}}</div>
        </div>
      </div>
      <div class="rank1st-items">`;
    if(top.length===0){{
      h+=`<div style="padding:16px 18px;color:rgba(255,255,255,.3);font-size:12px">Sem dados no período</div>`;
    }}else{{
      top.forEach((item,i)=>{{
        const m=i<3?MEDALS[i]:null;
        const pct=Math.round(item.val/maxVal*100);
        const bStyle=m
          ?`--rbg:${{m.bg}};--rbd:${{m.bd}};--rcl:${{m.cl}};background:${{m.bg}};border-color:${{m.bd}};color:${{m.cl}}`
          :`background:rgba(255,255,255,.06);border-color:rgba(255,255,255,.15);color:rgba(255,255,255,.45)`;
        h+=`
          <div class="rank1st-item">
            <div class="rank1st-top">
              <div class="rank1st-badge" style="${{bStyle}}">${{i+1}}</div>
              <div class="rank1st-nome">${{item.nome}}</div>
            </div>
            <div class="rank1st-val" style="color:${{col.ac}}">${{col.fmt(item.val)}}</div>
            <div class="rank1st-bar">
              <div class="rank1st-bar-bg"><div class="rank1st-bar-fill" style="width:${{pct}}%;background:${{col.ac}}"></div></div>
              <div class="rank1st-bar-pct">${{pct}}%</div>
            </div>
          </div>`;
      }});
    }}
    h+=`</div>`;
    colEl.innerHTML=h;
    grid.appendChild(colEl);
  }});
}}

// ── LINHA DE R$10K + COBERTURA DE PRODUTOS ────────────────────────────────────
function getMVAt(d,bloco,month,...kws){{
  const bk=d.blocos?.[bloco]; if(!bk)return null;
  const ind=bk.inds.find(i=>!i.sub&&kws.some(kw=>normStr(i.n).includes(kw)));
  if(!ind?.mensal)return null;
  return ind.mensal[String(month)]?.e??null;
}}
function combinedRevAt(d,m){{
  const mrrA=getMVAt(d,'acompanhamento',m,'mrr acumulado');
  const seg=getMVAt(d,'seguro',m,'vendido','valor vendido');
  const srv=getMVAt(d,'servicos',m,'valor vendido','vendido');
  if(mrrA==null&&seg==null&&srv==null)return null;
  return(mrrA||0)+(seg||0)+(srv||0);
}}
function combinedMonthly(d){{
  const r=[];
  for(let m=1;m<=12;m++){{const v=combinedRevAt(d,m);if(v!=null)r.push({{m,v}});}}
  return r;
}}
const LINHA_10K=10000;
function analyze10K(d){{
  const pts=combinedMonthly(d);
  if(pts.length===0)return null;
  const crossIdx=pts.findIndex(p=>p.v>=LINHA_10K);
  if(crossIdx>=0){{
    const cur=pts[pts.length-1];
    return{{status:'crossed',nome:d.nome,crossMonth:pts[crossIdx].m,msToCross:crossIdx+1,currentVal:cur.v,currentMonth:cur.m}};
  }}
  if(pts.length<2)return null;
  const last=pts[pts.length-1],prev=pts[pts.length-2];
  if(last.v<=prev.v||last.v<=0)return null;
  const growthPM=last.v-prev.v;
  const msToLine=Math.ceil((LINHA_10K-last.v)/growthPM);
  if(msToLine<=0||msToLine>12)return null;
  return{{status:'approaching',nome:d.nome,currentVal:last.v,currentMonth:last.m,growthPM,msToLine,pct:Math.min(99,Math.round(last.v/LINHA_10K*100))}};
}}

function renderGrowthCoverage(){{
  const grid=document.getElementById('portfolioGrid');
  if(!grid)return;
  const MN=['','Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'];

  const analyses=DADOS.map(d=>analyze10K(d)).filter(a=>a!=null);
  const approaching=analyses.filter(a=>a.status==='approaching').sort((a,b)=>a.msToLine-b.msToLine).slice(0,5);

  // Ranking de crescimento absoluto: receita combinada último mês vs mês anterior
  const growthRanking=DADOS.map(d=>{{
    const pts=combinedMonthly(d);
    if(pts.length<2)return null;
    const last=pts[pts.length-1],prev=pts[pts.length-2];
    return{{nome:d.nome,delta:last.v-prev.v,lastVal:last.v,prevVal:prev.v,lastMonth:last.m,prevMonth:prev.m}};
  }}).filter(g=>g!=null).sort((a,b)=>b.delta-a.delta).slice(0,5);

  // Cobertura — sempre período completo, independe de curMes
  const total=DADOS.length;
  const withMrr=DADOS.filter(d=>(sumMonthly(d,'acompanhamento','mrr vendido')||0)>0).length;
  const withSeg=DADOS.filter(d=>(sumMonthly(d,'seguro','vendido','valor vendido')||0)>0).length;
  const withSrv=DADOS.filter(d=>(sumMonthly(d,'servicos','valor vendido','vendido')||0)>0).length;

  const MEDALS=[
    {{bg:'rgba(245,158,11,.14)',bd:'rgba(245,158,11,.4)',cl:'#f59e0b'}},
    {{bg:'rgba(148,163,184,.12)',bd:'rgba(148,163,184,.4)',cl:'#94a3b8'}},
    {{bg:'rgba(180,83,9,.12)',bd:'rgba(180,83,9,.4)',cl:'#b45309'}},
    {{bg:'rgba(255,255,255,.05)',bd:'rgba(255,255,255,.14)',cl:'rgba(255,255,255,.38)'}},
    {{bg:'rgba(255,255,255,.05)',bd:'rgba(255,255,255,.14)',cl:'rgba(255,255,255,.38)'}}
  ];
  const growthHtml=growthRanking.length===0
    ?`<div style="padding:14px 16px;color:rgba(255,255,255,.28);font-size:11px">Dados insuficientes — mínimo 2 meses por empresa</div>`
    :growthRanking.map((g,i)=>{{
      const m=MEDALS[i]||MEDALS[3];
      const isPos=g.delta>=0;
      const deltaColor=isPos?'#10b981':'#ef4444';
      const deltaSign=isPos?'+':'';
      return`<div class="growth-item">
        <div class="growth-badge" style="background:${{m.bg}};border-color:${{m.bd}};color:${{m.cl}}">${{i+1}}</div>
        <div class="growth-info">
          <div class="growth-name">${{g.nome}}</div>
          <div class="growth-detail">${{MN[g.prevMonth]}} ${{fmtM(g.prevVal)}} → ${{MN[g.lastMonth]}} ${{fmtM(g.lastVal)}}</div>
        </div>
        <div style="text-align:right;flex-shrink:0">
          <div style="font-size:13px;font-weight:800;color:${{deltaColor}};line-height:1;white-space:nowrap">${{deltaSign}}${{fmtM(g.delta)}}</div>
        </div>
      </div>`;
    }}).join('');

  const approachHtml=approaching.length===0
    ?`<div style="padding:14px 16px;color:rgba(255,255,255,.28);font-size:11px">Sem empresas avançando para cruzar em breve</div>`
    :approaching.map(g=>`
      <div class="growth-item" style="flex-direction:column;gap:6px;align-items:stretch">
        <div style="display:flex;justify-content:space-between;align-items:center">
          <div class="growth-name" style="font-size:11px">${{g.nome}}</div>
          <div style="font-size:13px;font-weight:800;color:#3b82f6;flex-shrink:0;margin-left:8px">≈${{g.msToLine}} ${{g.msToLine===1?'mês':'meses'}}</div>
        </div>
        <div style="display:flex;align-items:center;gap:8px">
          <div style="flex:1;height:5px;background:rgba(255,255,255,.07);border-radius:3px;overflow:hidden">
            <div style="height:100%;width:${{g.pct}}%;background:linear-gradient(90deg,#1d4ed8,#3b82f6);border-radius:3px"></div>
          </div>
          <div style="font-size:10px;color:#3b82f6;font-weight:700;min-width:32px;text-align:right">${{g.pct}}%</div>
        </div>
        <div style="font-size:10px;color:rgba(255,255,255,.32)">${{fmtM(g.currentVal)}} atual · +${{fmtM(g.growthPM)}}/mês</div>
      </div>`).join('');

  const covHtml=[
    {{label:'MRR Novo Gerado',count:withMrr,ac:'#10b981'}},
    {{label:'Seguro Vendido',count:withSeg,ac:'#06b6d4'}},
    {{label:'Serviços Vendidos',count:withSrv,ac:'#f97316'}},
  ].map(c=>{{
    const pct=Math.round(c.count/total*100);
    return`<div class="coverage-item">
      <div class="coverage-head">
        <span class="coverage-label">${{c.label}}</span>
        <span class="coverage-count">${{c.count}} de ${{total}} empresas</span>
      </div>
      <div class="coverage-bar-bg"><div class="coverage-bar-fill" style="width:${{pct}}%;background:${{c.ac}}"></div></div>
      <div class="coverage-pct" style="color:${{c.ac}}">${{pct}}%</div>
    </div>`;
  }}).join('');

  grid.innerHTML=`
    <div class="portfolio-card">
      <div class="portfolio-card-hdr">
        <div class="portfolio-card-hdr-icon" style="background:rgba(245,158,11,.12)">
          <svg viewBox="0 0 24 24" width="18" height="18" stroke="#f59e0b" stroke-width="2" fill="none" stroke-linecap="round" stroke-linejoin="round">
            <line x1="12" y1="20" x2="12" y2="10"/><line x1="18" y1="20" x2="18" y2="4"/><line x1="6" y1="20" x2="6" y2="16"/>
          </svg>
        </div>
        <div>
          <div class="portfolio-card-title">Evolução da Receita Combinada</div>
          <div class="portfolio-card-sub">MRR Acumulado + Seguro + Serviços · comparativo mensal</div>
        </div>
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr;border-top:1px solid rgba(255,255,255,.05)">
        <div style="border-right:1px solid rgba(255,255,255,.05)">
          <div style="padding:8px 16px 4px;font-size:9px;font-weight:700;letter-spacing:.08em;text-transform:uppercase;color:rgba(255,255,255,.28)">📈 Maior crescimento no último mês</div>
          <div class="growth-list">${{growthHtml}}</div>
        </div>
        <div>
          <div style="padding:8px 16px 4px;font-size:9px;font-weight:700;letter-spacing:.08em;text-transform:uppercase;color:rgba(255,255,255,.28)">🚀 Avançando para R$10K</div>
          <div class="growth-list">${{approachHtml}}</div>
        </div>
      </div>
    </div>
    <div class="portfolio-card">
      <div class="portfolio-card-hdr">
        <div class="portfolio-card-hdr-icon" style="background:rgba(168,85,247,.12)">
          <svg viewBox="0 0 24 24" width="18" height="18" stroke="#a855f7" stroke-width="2" fill="none" stroke-linecap="round" stroke-linejoin="round">
            <path d="M21.21 15.89A10 10 0 1 1 8 2.83"/><path d="M22 12A10 10 0 0 0 12 2v10z"/>
          </svg>
        </div>
        <div>
          <div class="portfolio-card-title">Cobertura de Produtos</div>
          <div class="portfolio-card-sub">% de empresas com venda no período completo</div>
        </div>
      </div>
      <div class="coverage-list">${{covHtml}}</div>
    </div>`;
}}

// ── TABLE ────────────────────────────────────────────────────────────────────
let curFilter='todos', curSort={{col:'score',dir:'desc'}}, expandedId=null, curMes=0;

// ── FILTRO DE MÊS ─────────────────────────────────────────────────────────────
function updateFonteBanner(m){{
  const banner=document.getElementById('fonteBanner');
  const texto=document.getElementById('fonteTexto');
  if(!banner||!texto)return;
  if(m===0){{
    banner.style.display='block';
    texto.innerHTML=
      '<strong style="color:#93c5fd">Todo o período</strong> &nbsp;·&nbsp; '
      +'<span style="color:rgba(255,255,255,0.72)">MRR Vendido · Seguro · Serviços · Indicações</span> = soma de todos os meses executados'
      +'&nbsp;&nbsp;|&nbsp;&nbsp;<span style="color:rgba(255,255,255,0.72)">Carteira · MRR Acumulado · Agentes PV</span> = último mês preenchido'
      +'&nbsp;&nbsp;|&nbsp;&nbsp;<span style="color:rgba(255,255,255,0.72)">Conversão</span> = indisponível (selecione um mês)';
  }}else{{
    banner.style.display='block';
    texto.innerHTML=
      `<strong style="color:#F59E0B">${{MESES[m]}}/26</strong> &nbsp;·&nbsp; `
      +`todos os indicadores exibem o executado de ${{MESES[m]}}. `
      +`Conversão = Vol.Vendido ÷ Vol.Abordado do mês.`;
  }}
}}
function setMes(m, btn){{
  curMes=m; expandedId=null;
  document.querySelectorAll('.mfbtn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  updateFonteBanner(m);
  renderTable();
}}

// ── HELPERS MENSAIS ────────────────────────────────────────────────────────────
function normStr(s){{return(s||'').toLowerCase().normalize('NFD').replace(/[^a-z0-9 ]/g,'').trim();}}
function getMV(d,bloco,...kws){{
  const bk=d.blocos?.[bloco]; if(!bk)return null;
  const ind=bk.inds.find(i=>!i.sub&&kws.some(kw=>normStr(i.n).includes(kw)));
  if(!ind?.mensal)return null;
  return ind.mensal[String(curMes)]?.e??null;
}}
function getMP(d,bloco,...kws){{
  const bk=d.blocos?.[bloco]; if(!bk)return null;
  const ind=bk.inds.find(i=>!i.sub&&kws.some(kw=>normStr(i.n).includes(kw)));
  if(!ind?.mensal)return null;
  return ind.mensal[String(curMes)]?.p??null;
}}
function effS(eff){{
  if(eff==null)return'nd';
  if(eff>=80)return'verde';
  if(eff>=60)return'amarelo';
  return'vermelho';
}}
function calcEff(e,p){{return(e!=null&&p!=null&&p>0)?e/p*100:null;}}
function sumMonthly(d,bloco,...kws){{
  const bk=d.blocos?.[bloco]; if(!bk)return null;
  const ind=bk.inds.find(i=>!i.sub&&kws.some(kw=>normStr(i.n).includes(kw)));
  if(!ind?.mensal)return null;
  const vals=Object.values(ind.mensal).map(v=>v?.e).filter(v=>v!=null);
  return vals.length>0?vals.reduce((a,b)=>a+b,0):null;
}}
function sumMonthlyP(d,bloco,...kws){{
  const bk=d.blocos?.[bloco]; if(!bk)return null;
  const ind=bk.inds.find(i=>!i.sub&&kws.some(kw=>normStr(i.n).includes(kw)));
  if(!ind?.mensal)return null;
  const vals=Object.values(ind.mensal).map(v=>v?.p).filter(v=>v!=null);
  return vals.length>0?vals.reduce((a,b)=>a+b,0):null;
}}
function lastMonthExec(d,bloco,...kws){{
  const bk=d.blocos?.[bloco]; if(!bk)return null;
  const ind=bk.inds.find(i=>!i.sub&&kws.some(kw=>normStr(i.n).includes(kw)));
  if(!ind?.mensal)return ind?.e??null;
  const ms=Object.keys(ind.mensal).map(Number).filter(m=>ind.mensal[String(m)]?.e!=null).sort((a,b)=>b-a);
  return ms.length>0?ind.mensal[String(ms[0])].e:(ind.e??null);
}}

// Retorna valores efetivos (período ou mês selecionado) para uma empresa
function getAgentesPV(d){{
  const bk=d.blocos?.acompanhamento; if(!bk)return null;
  const ind=bk.inds.find(i=>!i.sub&&(normStr(i.n).includes('quantidade de agente')||normStr(i.n).includes('quantidade de operador')||normStr(i.n).includes('quantida de operador')));
  if(!ind)return null;
  if(curMes>0) return ind.mensal?.[String(curMes)]?.e??null;
  // Todo período → último mês preenchido
  const mFilled=Object.keys(ind.mensal||{{}}).map(Number).filter(m=>ind.mensal[String(m)]?.e!=null).sort((a,b)=>b-a);
  if(mFilled.length>0)return ind.mensal[String(mFilled[0])].e;
  return ind.e??null;
}}

function eVals(d){{
  const agE=getAgentesPV(d);
  if(!curMes){{
    // MRR Vendido, Seguro, Serviços, Indicações → soma dos executados mensais
    const mrrVE=sumMonthly(d,'acompanhamento','mrr vendido');
    const segE=sumMonthly(d,'seguro','vendido','valor vendido');
    const srvE=sumMonthly(d,'servicos','valor vendido','vendido');
    const indE=sumMonthly(d,'acompanhamento','volume indicacoes','indicacoes');
    // Carteira → último mês executado preenchido
    const carteiraE=lastMonthExec(d,'acompanhamento','clientes em carteira','total em carteira');
    // Vol. Abordado → soma mensal dos 4 componentes
    const vaA=sumMonthly(d,'acompanhamento','volume abordado');
    const vaS=sumMonthly(d,'servicos','volume abordado');
    const vaG=sumMonthly(d,'seguro','volume abordado');
    const vaSf=sumMonthly(d,'acompanhamento','safra');
    const hasVA=vaA!=null||vaS!=null||vaG!=null||vaSf!=null;
    const volAE=hasVA?(vaA??0)+(vaS??0)+(vaG??0)+(vaSf??0):null;
    const vaPsum=(sumMonthlyP(d,'acompanhamento','volume abordado')??0)
               +(sumMonthlyP(d,'servicos','volume abordado')??0)
               +(sumMonthlyP(d,'seguro','volume abordado')??0)
               +(sumMonthlyP(d,'acompanhamento','safra')??0);
    const volAEff=calcEff(volAE,vaPsum>0?vaPsum:null);
    // Conversão → traço no período completo
    return{{
      mrr_acc_e:d.mrr_acc_e, mrr_acc_eff:d.mrr_acc_eff, mrr_acc_s:d.mrr_acc_s,
      mrr_vend_e:mrrVE, mrr_vend_eff:null, mrr_vend_s:'nd',
      carteira_e:carteiraE, carteira_s:d.carteira_s,
      conversao_e:null, conversao_s:'nd',
      conv_acomp_e:null, conv_acomp_s:'nd',
      conv_srv_e:null, conv_srv_s:'nd',
      conv_seg_e:null, conv_seg_s:'nd',
      vol_abord_e:volAE, vol_abord_eff:volAEff, vol_abord_s:effS(volAEff),
      seguro_e:segE, seguro_eff:null, seguro_s:'nd',
      servicos_e:srvE, servicos_eff:null, servicos_s:'nd',
      indicacoes_e:indE, agentes_e:agE
    }};
  }}
  // Valores mensais
  const mrrAE=getMV(d,'acompanhamento','mrr acumulado');
  const mrrAEff=calcEff(mrrAE,getMP(d,'acompanhamento','mrr acumulado'));
  const mrrVE=getMV(d,'acompanhamento','mrr vendido');
  const mrrVEff=calcEff(mrrVE,getMP(d,'acompanhamento','mrr vendido'));
  // Vol. Abordado mensal = soma das 4 abordagens
  const vaAcc=getMV(d,'acompanhamento','volume abordado')??0;
  const vaSrv=getMV(d,'servicos','volume abordado')??0;
  const vaSeg=getMV(d,'seguro','volume abordado')??0;
  const vaSafra=getMV(d,'acompanhamento','safra')??0;
  const hasVA=(getMV(d,'acompanhamento','volume abordado')!=null||getMV(d,'servicos','volume abordado')!=null||getMV(d,'seguro','volume abordado')!=null||getMV(d,'acompanhamento','safra')!=null);
  const volA=hasVA?vaAcc+vaSrv+vaSeg+vaSafra:null;
  const volAP=(getMP(d,'acompanhamento','volume abordado')??0)+(getMP(d,'servicos','volume abordado')??0)+(getMP(d,'seguro','volume abordado')??0)+(getMP(d,'acompanhamento','safra')??0);
  const volAEff=calcEff(volA,volAP>0?volAP:null);
  const volV=getMV(d,'acompanhamento','volume vendido');
  // Conversão monitoramento: vol_vendido / vol_abordado (consistente com planilha)
  const convDenom=getMV(d,'acompanhamento','volume abordado');
  const conv=(convDenom!=null&&convDenom>0&&volV!=null)?volV/convDenom*100:null;
  // Taxa de conversão mensal por bloco (vol_vendido / vol_abordado × 100)
  function mConvCalc(bk){{
    const vv=getMV(d,bk,'volume vendido','volume vendas'), va=getMV(d,bk,'volume abordado');
    return (va!=null&&va>0&&vv!=null)?vv/va*100:null;
  }}
  function mConvStatus(e,bk){{
    if(e==null)return'nd';
    const p=getMP(d,bk,'taxa de conversao'); if(p==null||p<=0)return'nd';
    return effS(calcEff(e, p<=1.5?p*100:p));
  }}
  const cAE=mConvCalc('acompanhamento'), cSrE=mConvCalc('servicos'), cSeE=mConvCalc('seguro');
  const segE=getMV(d,'seguro','vendido','valor vendido');
  const segEff=calcEff(segE,getMP(d,'seguro','vendido','valor vendido'));
  const srvE=getMV(d,'servicos','valor vendido','vendido');
  const srvEff=calcEff(srvE,getMP(d,'servicos','valor vendido','vendido'));
  const indE=getMV(d,'acompanhamento','volume indicacoes','indicacoes');
  return{{
    mrr_acc_e:mrrAE, mrr_acc_eff:mrrAEff, mrr_acc_s:effS(mrrAEff),
    mrr_vend_e:mrrVE, mrr_vend_eff:mrrVEff, mrr_vend_s:effS(mrrVEff),
    carteira_e:d.carteira_e, carteira_s:d.carteira_s,
    conversao_e:conv, conversao_s:effS(conv),
    conv_acomp_e:cAE, conv_acomp_s:mConvStatus(cAE,'acompanhamento'),
    conv_srv_e:cSrE, conv_srv_s:mConvStatus(cSrE,'servicos'),
    conv_seg_e:cSeE, conv_seg_s:mConvStatus(cSeE,'seguro'),
    vol_abord_e:volA, vol_abord_eff:volAEff, vol_abord_s:effS(volAEff),
    seguro_e:segE, seguro_eff:segEff, seguro_s:effS(segEff),
    servicos_e:srvE, servicos_eff:srvEff, servicos_s:effS(srvEff),
    indicacoes_e:indE, agentes_e:agE
  }};
}}

function setFilter(f,btn){{
  curFilter=f; expandedId=null;
  document.querySelectorAll('.fbtn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  renderTable();
}}

function sortBy(col){{
  curSort=curSort.col===col?{{col,dir:curSort.dir==='asc'?'desc':'asc'}}:{{col,dir:'desc'}};
  expandedId=null; renderTable();
}}

function filteredData(){{
  const q=(document.getElementById('searchBox')?.value||'').toLowerCase();
  return DADOS.filter(d=>{{
    const st=(d.status_tag||'').toUpperCase();
    let fm;
    if(curFilter==='todos')fm=true;
    else if(curFilter==='critico')fm=st==='CRÍTICO'||st==='CRITICO';
    else if(curFilter==='alerta')fm=st==='ALERTA';
    else if(curFilter==='sem_tag')fm=!st;
    else fm=true;
    const sm=!q||d.nome.toLowerCase().includes(q)||d.id.includes(q);
    return fm&&sm;
  }});
}}

function sortedData(data){{
  const col=curSort.col,dir=curSort.dir;
  // Para colunas com dados mensais, ordenar pelo valor efetivo
  const mensalCols=new Set(['mrr_acc_e','mrr_vend_e',
    'conv_acomp_e','conv_srv_e','conv_seg_e',
    'vol_abord_e','seguro_e','servicos_e','indicacoes_e','agentes_e']);
  return[...data].sort((a,b)=>{{
    let va,vb;
    if(col==='rank'){{va=DADOS.indexOf(a);vb=DADOS.indexOf(b);}}
    else if(col==='nome'){{va=a.nome.toLowerCase();vb=b.nome.toLowerCase();}}
    else if(mensalCols.has(col)){{va=eVals(a)[col];vb=eVals(b)[col];}}
    else{{va=a[col];vb=b[col];}}
    if(va==null)va=dir==='asc'?Infinity:-Infinity;
    if(vb==null)vb=dir==='asc'?Infinity:-Infinity;
    return va<vb?(dir==='asc'?-1:1):va>vb?(dir==='asc'?1:-1):0;
  }});
}}

function renderTable(){{
  const data=sortedData(filteredData());
  const cols=['rank','nome','score','mrr_acc_e','mrr_vend_e','carteira_e','conv_acomp_e','conv_srv_e','conv_seg_e','vol_abord_e','seguro_e','servicos_e','indicacoes_e','agentes_e'];
  document.querySelectorAll('th').forEach((th,i)=>{{
    th.classList.remove('sa','sd');
    if(cols[i]===curSort.col)th.classList.add(curSort.dir==='asc'?'sa':'sd');
  }});
  const mesTag=curMes?` <span style="font-size:10px;background:#1c3a5e;color:#93c5fd;border-radius:3px;padding:1px 5px;margin-left:4px">${{MESES[curMes]}}/26</span>`:'';
  document.getElementById('tableCount').textContent='';
  document.getElementById('tableCount').innerHTML=data.length+' empresa(s)'+mesTag;

  let html='';
  data.forEach(d=>{{
    const ev=eVals(d);
    const rank=DADOS.indexOf(d)+1;
    const isExp=expandedId===d.id;
    const per=d.periodo?`${{MESES[d.periodo.inicio]||d.periodo.inicio}}${{d.periodo.inicio!==d.periodo.fim?'−'+MESES[d.periodo.fim]:''}}/26`:'—';
    const naStyle=curMes?'color:rgba(255,255,255,0.18)':'';
    const stag=(d.status_tag||'').toUpperCase();
    const tagHtml=stag==='CRÍTICO'||stag==='CRITICO'?'<span class="mtag-critico">CRÍTICO</span>':stag==='ALERTA'?'<span class="mtag-alerta">ALERTA</span>':'';
    const tempoHtml=(()=>{{
      if(!d.data_inicio)return'';
      const pts=d.data_inicio.split('/');
      if(pts.length!==3)return'';
      const dt=new Date(+pts[2],+pts[1]-1,+pts[0]);
      if(isNaN(dt)||dt.getFullYear()<2000||dt.getFullYear()>2100)return'';
      const now=new Date();
      const meses=Math.max(0,Math.round((now-dt)/(1000*60*60*24*30.44)));
      if(meses<1)return'';
      if(meses<12)return meses+(meses===1?' mês':' meses');
      const anos=Math.floor(meses/12),rm=meses%12;
      return anos+'a'+(rm?' '+rm+'m':'');
    }})();
    const subInfo=[d.id?'ID '+d.id:'',d.consultor||'',tempoHtml].filter(x=>x).join(' · ');
    html+=`<tr class="erow${{isExp?' expanded':''}}" onclick="toggleDet('${{d.id}}')">
      <td><span style="color:rgba(255,255,255,0.35);font-weight:700">${{rank}}</span></td>
      <td>
        <div style="display:flex;flex-direction:column;gap:1px">
          <span style="font-weight:600">${{d.nome}}${{tagHtml}}</span>
          <span style="color:rgba(255,255,255,0.35);font-size:11px">${{subInfo}}</span>
        </div>
      </td>
      <td>${{scoreBadge(d.score,d.status)}}</td>
      <td style="font-weight:500">${{ev.mrr_acc_e!=null?fmtM(ev.mrr_acc_e):'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
      <td style="font-weight:500">${{ev.mrr_vend_e!=null?fmtM(ev.mrr_vend_e):'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
      <td style="${{naStyle}}">${{fmtN(ev.carteira_e)}}</td>
      <td style="color:${{convColor(ev.conv_acomp_e)}};font-weight:600">${{ev.conv_acomp_e!=null?ev.conv_acomp_e.toFixed(1)+'%':'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
      <td style="color:${{convColor(ev.conv_srv_e)}};font-weight:600">${{ev.conv_srv_e!=null?ev.conv_srv_e.toFixed(1)+'%':'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
      <td style="color:${{convColor(ev.conv_seg_e)}};font-weight:600">${{ev.conv_seg_e!=null?ev.conv_seg_e.toFixed(1)+'%':'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
      <td style="color:${{sColor(ev.vol_abord_s)}}">${{ev.vol_abord_e!=null?fmtN(ev.vol_abord_e):'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
      <td style="font-weight:500">${{ev.seguro_e!=null?fmtM(ev.seguro_e):'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
      <td style="font-weight:500">${{ev.servicos_e!=null?fmtM(ev.servicos_e):'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
      <td style="color:rgba(255,255,255,0.5)">${{ev.indicacoes_e!=null?fmtN(ev.indicacoes_e):'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
      <td style="color:#7dd3fc;font-weight:500">${{ev.agentes_e!=null?fmtN(ev.agentes_e):'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
    </tr>`;
    if(isExp)html+=`<tr class="drow"><td colspan="14"><div class="din">${{renderDetail(d)}}</div></td></tr>`;
  }});
  if(!html)html='<tr><td colspan="14" class="nodata">Nenhuma empresa encontrada</td></tr>';
  document.getElementById('tableBody').innerHTML=html;
}}

function toggleDet(id){{expandedId=expandedId===id?null:id;renderTable();}}

// ── MATRIZ MENSAL ─────────────────────────────────────────────────────────────
function getIndBlocos(d,bk,...kws){{
  const bloco=d.blocos?.[bk]; if(!bloco)return null;
  return bloco.inds.find(i=>!i.sub&&kws.some(kw=>normStr(i.n).includes(kw)))||null;
}}

function renderMatriz(d){{
  const GRUPOS=[
    {{g:'Pipeline',cls:'g0',inds:[
      {{label:'Safra de Negócio',bk:'acompanhamento',kws:['safra'],tipo:'vol'}}
    ]}},
    {{g:'Abordagens',cls:'g1',inds:[
      {{label:'Abord. Monitoramento',bk:'acompanhamento',kws:['volume abordado'],tipo:'vol'}},
      {{label:'Abord. Serviço',bk:'servicos',kws:['volume abordado'],tipo:'vol'}},
      {{label:'Abord. Seguro',bk:'seguro',kws:['volume abordado'],tipo:'vol'}}
    ]}},
    {{g:'Volume Vendido (qtd)',cls:'g2',inds:[
      {{label:'Vol. Monitor.',bk:'acompanhamento',kws:['volume vendido'],tipo:'vol'}},
      {{label:'Vol. Serviço',bk:'servicos',kws:['volume vendido','volume vendas'],tipo:'vol'}},
      {{label:'Vol. Seguro',bk:'seguro',kws:['volume vendido','volume vendas'],tipo:'vol'}}
    ]}},
    {{g:'Valor Vendido (R$)',cls:'g3',inds:[
      {{label:'MRR Vendido',bk:'acompanhamento',kws:['mrr vendido'],tipo:'mon'}},
      {{label:'Valor Serviço',bk:'servicos',kws:['valor vendido','vendido'],tipo:'mon'}},
      {{label:'Valor Seguro',bk:'seguro',kws:['vendido','valor vendido'],tipo:'mon'}}
    ]}},
    {{g:'Equipe',cls:'g4',inds:[
      {{label:'Qtd. Agentes PV',bk:'acompanhamento',kws:['quantidade de agente','operador'],tipo:'vol'}}
    ]}}
  ];

  // Descobrir quais meses têm algum dado
  const mSet=new Set();
  GRUPOS.forEach(grp=>grp.inds.forEach(cfg=>{{
    const ind=getIndBlocos(d,cfg.bk,...cfg.kws);
    if(ind?.mensal) Object.keys(ind.mensal).forEach(m=>{{
      const mv=ind.mensal[m];
      if(mv&&(mv.p!=null||mv.e!=null)) mSet.add(parseInt(m));
    }});
  }}));
  if(!mSet.size) return '<div style="padding:14px;color:#475569;font-size:12px">Sem dados mensais disponíveis</div>';
  const meses=[...mSet].sort((a,b)=>a-b);

  function fmtC(v,tipo){{
    if(v==null)return'—';
    if(tipo==='mon'){{
      if(Math.abs(v)>=1000)return'R$'+(v/1000).toFixed(1)+'K';
      return'R$'+Math.round(v);
    }}
    return v>=1000?(v/1000).toFixed(1)+'K':String(Math.round(v));
  }}
  function cellCls(p,e){{
    if(e==null)return'mc-nd';
    if(p==null||p===0)return e>0?'mc-nd':'mc-nd';
    const r=e/p;
    if(r>=0.8)return'mc-v';
    if(r>=0.6)return'mc-a';
    return'mc-r';
  }}

  let t=`<div class="mmatriz-wrap">
    <div class="mmatriz-title">Matriz de Acompanhamento Mensal — Plano / Executado</div>
    <div style="overflow-x:auto"><table class="mmatriz">
      <thead><tr>
        <th class="mth-lbl">Indicador</th>
        ${{meses.map(m=>`<th class="mth-mes">${{MESES[m]}}</th>`).join('')}}
      </tr></thead><tbody>`;

  GRUPOS.forEach(grp=>{{
    t+=`<tr class="mtr-grp ${{grp.cls}}"><td colspan="${{meses.length+1}}">${{grp.g}}</td></tr>`;
    grp.inds.forEach(cfg=>{{
      const ind=getIndBlocos(d,cfg.bk,...cfg.kws);
      t+=`<tr class="mtr-ind"><td class="mtd-lbl">${{cfg.label}}</td>`;
      meses.forEach(m=>{{
        const mv=ind?.mensal?.[String(m)];
        if(!mv||(mv.p==null&&mv.e==null)){{
          t+=`<td class="mtd mc-nd" style="color:rgba(255,255,255,0.15)">—</td>`;
        }}else{{
          const cls=cellCls(mv.p,mv.e);
          t+=`<td class="mtd ${{cls}}"><div class="mc-p">${{fmtC(mv.p,cfg.tipo)}}</div><div class="mc-e">${{fmtC(mv.e,cfg.tipo)}}</div></td>`;
        }}
      }});
      t+=`</tr>`;
    }});
  }});

  t+=`</tbody></table></div></div>`;
  return t;
}}

// ── DETAIL ───────────────────────────────────────────────────────────────────
function renderDetail(d){{
  const BL={{acompanhamento:'Acompanhamento',seguro:'Seguro Solar',servicos:'Serviços Avulsos'}};
  let html='';

  // Banner de observação (apenas CRÍTICO/ALERTA com texto preenchido)
  const stag=(d.status_tag||'').toUpperCase();
  if((stag==='CRÍTICO'||stag==='CRITICO'||stag==='ALERTA')&&d.observacao){{
    const isCrit=stag!=='ALERTA';
    const bg=isCrit?'rgba(239,68,68,0.10)':'rgba(245,158,11,0.10)';
    const bc=isCrit?'rgba(239,68,68,0.30)':'rgba(245,158,11,0.30)';
    const tc=isCrit?'#EF4444':'#F59E0B';
    const lbl=isCrit?'CRÍTICO':'ALERTA';
    html+=`<div class="obs-banner" style="background:${{bg}};border-color:${{bc}}">
      <span class="obs-tag" style="background:${{isCrit?'rgba(239,68,68,0.18)':'rgba(245,158,11,0.18)'}};color:${{tc}}">${{lbl}}</span>
      <span class="obs-txt">${{d.observacao}}</span>
    </div>`;
  }}

  // Diagnóstico
  const hasDiag=d.diagnosis&&d.diagnosis.length;
  html+=`<div class="dcard">
    <div class="dcard-hdr"><span>Diagnóstico Automático</span><span style="color:rgba(255,255,255,0.35)">${{d.total_inds}} ind. avaliados</span></div>
    ${{hasDiag?d.diagnosis.map(obs=>`
      <div class="diag-row">
        <span class="diag-icon">${{diagIcon(obs.tipo)}}</span>
        <span class="diag-txt" style="color:${{obs.tipo==='positivo'?'#22C55E':obs.tipo==='critico'?'#EF4444':obs.tipo==='alerta'?'#F59E0B':'rgba(255,255,255,0.55)'}}">${{obs.texto}}</span>
      </div>`).join(''):'<div class="nodata">Sem diagnóstico disponível</div>'}}
    <div style="padding:10px 14px;display:flex;gap:16px;flex-wrap:wrap;border-top:1px solid rgba(255,255,255,0.05)">
      ${{[['verde','#22C55E',d.verde,'Verde (≥80%)'],['amarelo','#F59E0B',d.amarelo,'Atenção (60-80%)'],['vermelho','#EF4444',d.vermelho,'Crítico (<60%)']].map(([s,c,n,l])=>`
        <div style="display:flex;align-items:center;gap:6px;font-size:12px">
          <span class="dot ${{s}}"></span><span style="color:${{c}};font-weight:700">${{n}}</span><span style="color:rgba(255,255,255,0.35)">${{l}}</span>
        </div>`).join('')}}
    </div>
  </div>`;

  // Blocos
  ['acompanhamento','seguro','servicos'].forEach(bk=>{{
    const bloco=d.blocos[bk];
    if(!bloco){{
      html+=`<div class="dcard"><div class="dcard-hdr">${{BL[bk]}}</div><div class="nodata">Bloco não encontrado na planilha</div></div>`;
      return;
    }}
    const mainInds=bloco.inds.filter(i=>!i.sub);
    const rows=mainInds.map(ind=>{{
      const e=curMes>0?(ind.mensal?.[String(curMes)]?.e??null):ind.e;
      let m;
      if(curMes>0){{
        m=ind.mensal?.[String(curMes)]?.p??null;
      }}else{{
        m=ind.m;
        if(m==null){{
          const mv=Object.values(ind.mensal||{{}});
          const ps=mv.reduce((a,v)=>a+(v?.p??0),0);
          m=ps>0?ps:null;
        }}
      }}
      const eff=calcEff(e,m);
      const s=effS(eff);
      const effFmt=eff!=null?fmtP(eff):'—';
      const bar=eff!=null?`<div class="etrack" style="width:52px"><div class="ebar ${{s}}" style="width:${{Math.min(Math.abs(eff),150)/1.5}}%"></div></div>`:'';
      const tk=tipHTML(ind.n);
      const sp=sparkline(ind.mensal||{{}});
      const isMonet=isMonetary(ind.n);
      const isPct=isPercent(ind.n);
      function fmtV(v){{if(v==null)return'—';if(isPct)return fmtP(v*100);if(isMonet)return fmtM(v);return fmtN(v);}}
      return`<div class="irow">
        <div class="iname">
          <div style="display:flex;flex-direction:column;gap:1px">
            <div style="display:flex;align-items:center;gap:5px">
              <span class="iname-txt">${{ind.n}}</span>
              ${{tk}}
            </div>
            ${{sp}}
          </div>
        </div>
        <div class="imeta">${{fmtV(m)}}</div>
        <div class="iexec" style="color:${{sColor(s)}}">${{fmtV(e)}}</div>
        <div class="ieff"><span class="eval ${{s}}">${{effFmt}}</span>${{bar}}</div>
      </div>`;
    }}).join('');
    html+=`<div class="dcard">
      <div class="dcard-hdr"><span>${{BL[bk]}}</span><span style="color:rgba(255,255,255,0.35)">${{mainInds.length}} indicadores</span></div>
      <div style="display:grid;grid-template-columns:1fr 72px 72px 86px;gap:6px;padding:6px 14px;border-bottom:1px solid rgba(255,255,255,0.08)">
        <span style="color:rgba(255,255,255,0.38);font-size:10px;text-transform:uppercase;letter-spacing:.05em">Indicador</span>
        <span style="color:rgba(255,255,255,0.38);font-size:10px;text-align:right">Meta</span>
        <span style="color:rgba(255,255,255,0.38);font-size:10px;text-align:right">Exec</span>
        <span style="color:rgba(255,255,255,0.38);font-size:10px;text-align:right">Eficiência</span>
      </div>
      ${{rows||'<div class="nodata">Sem dados</div>'}}
    </div>`;
  }});

  // Matriz mensal
  html+=`<div class="dcard" style="grid-column:1/-1"><div class="dcard-hdr"><span>Matriz de Acompanhamento Mensal</span><span style="color:rgba(255,255,255,0.35);font-size:11px">Plano / Executado por mês</span></div>${{renderMatriz(d)}}</div>`;

  if(d.erros&&d.erros.length){{
    html+=`<div class="dcard"><div class="dcard-hdr">⚠️ Avisos de Parsing</div>
      <div style="padding:10px 14px;color:#ef4444;font-size:12px">${{d.erros.join('<br>')}}</div></div>`;
  }}
  return html;
}}

// ── CONSTRUINDO ──────────────────────────────────────────────────────────────
function renderConstruindo(){{
  const c=CONSOL;
  const topMRR=[...DADOS].sort((a,b)=>(b.mrr_acc_e||0)-(a.mrr_acc_e||0)).slice(0,8);
  const maxMRR=topMRR[0]?.mrr_acc_e||1;
  const topSeg=[...DADOS].filter(d=>d.seguro_e).sort((a,b)=>(b.seguro_e||0)-(a.seguro_e||0)).slice(0,6);
  const topSrv=[...DADOS].filter(d=>d.servicos_e).sort((a,b)=>(b.servicos_e||0)-(a.servicos_e||0)).slice(0,6);
  const topInd=[...DADOS].filter(d=>d.indicacoes_e).sort((a,b)=>(b.indicacoes_e||0)-(a.indicacoes_e||0)).slice(0,6);
  const meses=[1,2,3,4,5,6,7,8,9,10,11,12];
  const mrrV=meses.map(m=>c.mrr_vend_mensal[String(m)]||0);
  const maxMV=Math.max(...mrrV,1);

  document.getElementById('construindoContent').innerHTML=`
  <div class="cgrid">

    <div class="ccard">
      <h3>Performance Geral</h3>
      <div class="distrow">
        <div class="distitem"><div class="distnum" style="color:var(--verde)">${{c.counts.verde}}</div><div class="distlabel">Verde</div></div>
        <div class="distitem"><div class="distnum" style="color:var(--amarelo)">${{c.counts.amarelo}}</div><div class="distlabel">Atenção</div></div>
        <div class="distitem"><div class="distnum" style="color:var(--vermelho)">${{c.counts.vermelho}}</div><div class="distlabel">Crítico</div></div>
      </div>
      <div style="display:flex;flex-direction:column;gap:8px;margin-top:12px">
        ${{[
          ['MRR Total Construído',fmtM(c.total_mrr),'#F59E0B'],
          ['MRR Vendido no Período',fmtM(c.total_mrr_vend),'#F0F0F0'],
          ['Total Clientes em Carteira',fmtN(c.total_carteira),'#F0F0F0'],
          ['Total Indicações Geradas',fmtN(c.total_indicacoes),'#F0F0F0'],
          ['Seguro Total Vendido',fmtM(c.total_seguro),'#F0F0F0'],
          ['Serviços Total Vendido',fmtM(c.total_servicos),'#F0F0F0'],
          ['Conversão Geral',c.conversao_geral!=null?fmtP(c.conversao_geral*100):'—','#F0F0F0'],
        ].map(([l,v,clr])=>`<div style="display:flex;justify-content:space-between;font-size:12px;border-bottom:1px solid rgba(255,255,255,0.07);padding-bottom:6px">
          <span style="color:rgba(255,255,255,0.48)">${{l}}</span><span style="color:${{clr}};font-weight:700">${{v}}</span>
        </div>`).join('')}}
      </div>
    </div>

    <div class="ccard">
      <h3>MRR Vendido por Mês (soma das empresas)</h3>
      <div class="mchart">
        ${{meses.map(m=>{{
          const val=c.mrr_vend_mensal[String(m)]||0;
          const h=maxMV>0?Math.max(Math.round((val/maxMV)*64),val>0?4:0):0;
          return`<div class="mbar-w">
            <div class="mbar" style="height:${{h}}px;background:${{val>0?'#F59E0B':'rgba(255,255,255,0.1)'}}" title="${{MESES[m]}}: ${{fmtM(val)}}"></div>
            <div class="mlabel">${{MESES[m]}}</div>
          </div>`;
        }}).join('')}}
      </div>
      <div style="margin-top:10px;color:rgba(255,255,255,0.35);font-size:11px;text-align:center">Passe o cursor sobre as barras para ver o valor</div>
    </div>

    <div class="ccard">
      <h3>Top MRR Acumulado Construído</h3>
      <div class="bchart">
        ${{topMRR.map(d=>`<div class="bitem">
          <div class="blabel" title="${{d.nome}}">${{d.nome}}</div>
          <div class="btrack" title="${{fmtM(d.mrr_acc_e)}}">
            <div class="bfill" style="width:${{((d.mrr_acc_e||0)/maxMRR*100).toFixed(1)}}%;background:${{sColor(d.mrr_acc_s||'nd')}}"></div>
          </div>
          <div class="bval">${{fmtM(d.mrr_acc_e)}}</div>
        </div>`).join('')}}
      </div>
    </div>

    <div class="ccard">
      <h3>Top Indicações Geradas</h3>
      ${{topInd.length?`<div class="rlist">${{topInd.map((d,i)=>`
        <div class="ritem">
          <span class="rpos">${{i+1}}</span>
          <span class="rname" title="${{d.nome}}">${{d.nome}}</span>
          <span class="rval">${{fmtN(d.indicacoes_e)}}</span>
        </div>`).join('')}}</div>`:'<div class="nodata">Sem dados</div>'}}
    </div>

    <div class="ccard">
      <h3>Top Seguro Vendido</h3>
      ${{topSeg.length?`<div class="rlist">${{topSeg.map((d,i)=>`
        <div class="ritem">
          <span class="rpos">${{i+1}}</span>
          <span class="rname" title="${{d.nome}}">${{d.nome}}</span>
          <span class="rval">${{fmtM(d.seguro_e)}}</span>
        </div>`).join('')}}</div>`:'<div class="nodata">Sem dados</div>'}}
    </div>

    <div class="ccard">
      <h3>Top Serviços Vendidos</h3>
      ${{topSrv.length?`<div class="rlist">${{topSrv.map((d,i)=>`
        <div class="ritem">
          <span class="rpos">${{i+1}}</span>
          <span class="rname" title="${{d.nome}}">${{d.nome}}</span>
          <span class="rval">${{fmtM(d.servicos_e)}}</span>
        </div>`).join('')}}</div>`:'<div class="nodata">Sem dados</div>'}}
    </div>

    <div class="ccard">
      <h3>Empresas com Maior Carteira</h3>
      <div class="rlist">
        ${{[...DADOS].filter(d=>d.carteira_e).sort((a,b)=>(b.carteira_e||0)-(a.carteira_e||0)).slice(0,6).map((d,i)=>`
          <div class="ritem">
            <span class="rpos">${{i+1}}</span>
            <span class="rname" title="${{d.nome}}">${{d.nome}}</span>
            <span class="rval">${{fmtN(d.carteira_e)}} cli</span>
          </div>`).join('')||'<div class="nodata">Sem dados</div>'}}
      </div>
    </div>

    <div class="ccard">
      <h3>Prioridade de Intervenção (por score)</h3>
      <div class="rlist">
        ${{[...DADOS].filter(d=>d.status==='vermelho'||d.status==='nd').sort((a,b)=>(a.score||0)-(b.score||0)).slice(0,8).map(d=>`
          <div class="ritem">
            <span class="dot ${{d.status}}"></span>
            <span class="rname" title="${{d.nome}}">${{d.nome}}</span>
            <span style="color:${{sColor(d.status)}};font-weight:600;font-size:12px;white-space:nowrap">${{d.score!=null?d.score.toFixed(0)+'%':'—'}}</span>
          </div>`).join('')||'<div class="nodata">Todas saudáveis ✅</div>'}}
      </div>
    </div>

  </div>`;
}}

// ── LEGENDAS ─────────────────────────────────────────────────────────────────
function renderLegendas(){{
  const all=[
    ['mrr acumulado','mrr acumulado'],
    ['mrr vendido','mrr vendido'],
    ['mrr churn','mrr churn'],
    ['volume abordado','volume abordado'],
    ['volume vendido','volume vendido'],
    ['taxa de conversao','taxa de conversao'],
    ['ticket medio','ticket medio'],
    ['quantidade de agente','quantidade de agente'],
    ['safra de negocio','safra de negocio'],
    ['clientes em carteira','clientes em carteira'],
    ['volume indicacoes','volume indicacoes'],
    ['vendido_seguro','vendido_seguro'],
    ['vendido_servicos','vendido_servicos'],
  ];
  document.getElementById('legendaGrid').innerHTML=all.map(([k])=>{{
    const l=LEGENDAS[k];
    if(!l)return'';
    return`<div class="leg-card">
      <div class="leg-titulo">
        ${{l.titulo}}
        <span class="leg-tipo">${{l.tipo}}</span>
      </div>
      <div class="leg-desc">${{l.desc}}</div>
      <div class="leg-impacto">💡 ${{l.impacto}}</div>
      ${{l.alerta?`<div class="leg-alerta">${{l.alerta}}</div>`:''}}
    </div>`;
  }}).join('');
}}

// ── INTELIGÊNCIA OPERACIONAL ─────────────────────────────────────────────────
let curConsultor='', curSortIO={{col:'score',dir:'desc'}}, expandedIdIO=null, curMesIO=0;
let ioInitialized=false;

function getIOData(){{
  return DADOS.filter(d=>!curConsultor||d.consultor===curConsultor);
}}
function onConsultorInput(val){{
  const all=[...new Set(DADOS.map(d=>d.consultor).filter(c=>c))];
  curConsultor=(val===''||val==='Todos')?'':all.includes(val)?val:'';
  if(val===''||all.includes(val))renderAllIO();
}}
function clearConsultor(){{
  document.getElementById('io-consultor-input').value='';
  curConsultor=''; renderAllIO();
}}
function populateConsultorList(){{
  const all=[...new Set(DADOS.map(d=>d.consultor).filter(c=>c))].sort();
  document.getElementById('io-consultores-list').innerHTML=all.map(c=>`<option value="${{c}}">`).join('');
}}
function renderObsPanel(data){{
  const STOP=new Set(['para','que','uma','com','por','não','mais','mas','dos','das','nos','nas','está','são','tem','foi','ela','ele','isso','esse','esta','este','essa','como','muito','também','ainda','quando','após','desde','mesmo','sobre','pelos','pelas','pelo','pela','nao','esta','esse','esse','num','numa','onde','qual','quem','todo','toda','todos','todas','cada','entre','sem','sim','ser','ter','fazer','estar','pois','logo','caso','tipo','isso','aqui','lá','já','ate','até','vai','vem','nos','nas','seu','sua','seus','suas','meu','minha','seus','nós','eles','elas','vocês','deles','delas','apenas','porém','então','assim','tanto','outros','outras','outro','outra','algum','alguma','nenhum','nenhuma','agora','antes','depois','sempre','nunca','talvez','bem','mal','grande','pequeno','novo','nova']);
  const isCrit=d=>{{const s=(d.status_tag||'').toUpperCase();return s==='CRÍTICO'||s==='CRITICO';}};
  const isAlert=d=>(d.status_tag||'').toUpperCase()==='ALERTA';
  const withObs=d=>(d.observacao||'').trim().length>0;

  const criticos=data.filter(d=>isCrit(d));
  const alertas=data.filter(d=>isAlert(d));
  const critObs=criticos.filter(withObs);
  const alertObs=alertas.filter(withObs);
  const total=criticos.length+alertas.length;
  const totalObs=critObs.length+alertObs.length;

  // Subtitle
  const sub=document.getElementById('io-obs-subtitle');
  if(sub) sub.textContent=`${{total}} empresas · ${{totalObs}} com observação preenchida`;

  // Frequência de palavras (por empresa, não por ocorrência, para evitar viés)
  const wordMap={{}};
  [...critObs,...alertObs].forEach(d=>{{
    const words=[...new Set(normStr(d.observacao).split(/\s+/))];
    words.forEach(w=>{{
      if(w.length>4&&!STOP.has(w)&&!/^\d+$/.test(w)){{
        wordMap[w]=(wordMap[w]||0)+1;
      }}
    }});
  }});
  const topKws=Object.entries(wordMap)
    .filter(([,n])=>n>=2)
    .sort(([,a],[,b])=>b-a)
    .slice(0,15);

  let html='';

  // Contadores
  html+=`<div class="obs-counts">
    <div class="obs-count-pill" style="border-color:rgba(239,68,68,0.3);background:rgba(239,68,68,0.07)">
      <span class="obs-count-num" style="color:#EF4444">${{criticos.length}}</span>
      <span style="color:rgba(255,255,255,0.5)">CRÍTICO<br><span style="font-size:10px;font-weight:400">${{critObs.length}} com obs.</span></span>
    </div>
    <div class="obs-count-pill" style="border-color:rgba(245,158,11,0.3);background:rgba(245,158,11,0.07)">
      <span class="obs-count-num" style="color:#F59E0B">${{alertas.length}}</span>
      <span style="color:rgba(255,255,255,0.5)">ALERTA<br><span style="font-size:10px;font-weight:400">${{alertObs.length}} com obs.</span></span>
    </div>
    <div class="obs-count-pill" style="border-color:rgba(255,255,255,0.1);background:rgba(255,255,255,0.03)">
      <span class="obs-count-num" style="color:rgba(255,255,255,0.4)">${{total-totalObs}}</span>
      <span style="color:rgba(255,255,255,0.35)">Sem observação<br><span style="font-size:10px;font-weight:400">preenchida</span></span>
    </div>
  </div>`;

  // Termos recorrentes
  if(topKws.length>0){{
    html+=`<div class="obs-kw-row">
      <span class="obs-kw-label">Termos recorrentes:</span>
      ${{topKws.map(([w,n])=>`<span class="obs-kw">${{w}} <span class="obs-kw-n">×${{n}}</span></span>`).join('')}}
    </div>`;
  }}

  // Função auxiliar de renderizar lista
  const renderList=(list,color,tag)=>{{
    if(!list.length) return`<div class="obs-empty">Nenhuma empresa ${{tag}} com observação preenchida.</div>`;
    return list.map(d=>`
      <div class="obs-item">
        <div class="obs-item-left">
          <div class="obs-item-nome">${{d.nome}}</div>
          <div class="obs-item-id">ID ${{d.id}}${{d.consultor?' · '+d.consultor:''}}</div>
        </div>
        <div class="obs-item-text">${{d.observacao}}</div>
      </div>`).join('');
  }};

  // Grupo CRÍTICO
  if(critObs.length>0){{
    html+=`<div class="obs-group-title" style="color:#EF4444">CRÍTICO — ${{critObs.length}} observaç${{critObs.length===1?'ão':'ões'}}</div>`;
    html+=renderList(critObs,'#EF4444','CRÍTICO');
  }}

  // Separador
  if(critObs.length>0&&alertObs.length>0){{
    html+=`<div style="height:14px"></div>`;
  }}

  // Grupo ALERTA
  if(alertObs.length>0){{
    html+=`<div class="obs-group-title" style="color:#F59E0B">ALERTA — ${{alertObs.length}} observaç${{alertObs.length===1?'ão':'ões'}}</div>`;
    html+=renderList(alertObs,'#F59E0B','ALERTA');
  }}

  if(totalObs===0){{
    html+=`<div class="obs-empty">Nenhuma observação preenchida para as empresas em CRÍTICO ou ALERTA${{curConsultor?' deste consultor':''}}.</div>`;
  }}

  document.getElementById('io-obs-content').innerHTML=html;
}}

function renderAllIO(){{
  const data=getIOData();
  renderDonutStatus(data);
  renderPerfConsultor(data);
  renderReceitaFonte(data);
  renderDistScore(data);
  renderFaixasMRR(data);
  renderTempoScore(data);
  renderAgenteAbord(data);
  renderObsPanel(data);
  renderIOTable();
}}
function setDonut(elId,segments){{
  const el=document.getElementById(elId); if(!el)return;
  const total=segments.reduce((s,g)=>s+g.pct,0);
  if(total===0){{el.style.background='rgba(255,255,255,0.08)';return;}}
  let pos=0;
  const stops=segments.map(g=>{{
    const from=pos,to=pos+g.pct; pos=to;
    return`${{g.clr}} ${{from.toFixed(1)}}% ${{to.toFixed(1)}}%`;
  }});
  el.style.background=`conic-gradient(${{stops.join(',')}})`;
}}
function renderDonutStatus(data){{
  const critico=data.filter(d=>{{const s=(d.status_tag||'').toUpperCase();return s==='CRÍTICO'||s==='CRITICO';}}).length;
  const alerta=data.filter(d=>(d.status_tag||'').toUpperCase()==='ALERTA').length;
  const semTag=data.filter(d=>!(d.status_tag||'').trim()).length;
  const total=data.length;
  setDonut('io-donut-status',[
    {{pct:total?critico/total*100:0,clr:'#EF4444'}},
    {{pct:total?alerta/total*100:0,clr:'#F59E0B'}},
    {{pct:total?semTag/total*100:0,clr:'rgba(255,255,255,0.15)'}},
  ]);
  const numEl=document.getElementById('io-donut-status-num');
  if(numEl)numEl.textContent=total;
  const leg=document.getElementById('io-donut-status-legend');
  if(leg)leg.innerHTML=`
    <div class="io-leg-item"><span class="io-leg-dot" style="background:#EF4444"></span><span>CRÍTICO <b>${{critico}}</b></span></div>
    <div class="io-leg-item"><span class="io-leg-dot" style="background:#F59E0B"></span><span>ALERTA <b>${{alerta}}</b></span></div>
    <div class="io-leg-item"><span class="io-leg-dot" style="background:rgba(255,255,255,0.25)"></span><span>Sem Tag <b>${{semTag}}</b></span></div>`;
}}
function renderPerfConsultor(data){{
  const groups={{}};
  data.forEach(d=>{{
    const c=d.consultor||'Sem consultor';
    if(!groups[c])groups[c]={{scores:[],count:0}};
    if(d.score!=null)groups[c].scores.push(d.score);
    groups[c].count++;
  }});
  const sorted=Object.entries(groups)
    .map(([name,g])=>{{
      const avg=g.scores.length?g.scores.reduce((a,b)=>a+b,0)/g.scores.length:0;
      return{{name,avgScore:avg,count:g.count}};
    }})
    .sort((a,b)=>b.avgScore-a.avgScore);
  const maxS=Math.max(...sorted.map(g=>g.avgScore),1);
  const el=document.getElementById('io-perf-consultor'); if(!el)return;
  el.innerHTML=sorted.map(g=>{{
    const clr=g.avgScore>=60?'#22C55E':g.avgScore>=35?'#F59E0B':'#EF4444';
    const pct=(g.avgScore/maxS*100).toFixed(1);
    const firstName=g.name.split(' ')[0];
    return`<div class="io-hbar">
      <div class="io-hbar-label" title="${{g.name}}">${{firstName}}</div>
      <div class="io-hbar-track"><div class="io-hbar-fill" style="width:${{pct}}%;background:${{clr}}"></div></div>
      <div class="io-hbar-val">${{g.avgScore.toFixed(0)}}% (${{g.count}})</div>
    </div>`;
  }}).join('');
}}
function renderReceitaFonte(data){{
  const mrrT=data.reduce((s,d)=>s+(sumMonthly(d,'acompanhamento','mrr vendido')||0),0);
  const segT=data.reduce((s,d)=>s+(sumMonthly(d,'seguro','vendido','valor vendido')||0),0);
  const srvT=data.reduce((s,d)=>s+(sumMonthly(d,'servicos','valor vendido','vendido')||0),0);
  const total=mrrT+segT+srvT;
  setDonut('io-donut-receita',[
    {{pct:total?mrrT/total*100:33,clr:'#3b82f6'}},
    {{pct:total?segT/total*100:33,clr:'#06b6d4'}},
    {{pct:total?srvT/total*100:34,clr:'#f97316'}},
  ]);
  const numEl=document.getElementById('io-donut-receita-num');
  if(numEl)numEl.textContent=total?fmtM(total):'—';
  const leg=document.getElementById('io-donut-receita-legend');
  if(leg)leg.innerHTML=`
    <div class="io-leg-item"><span class="io-leg-dot" style="background:#3b82f6"></span><span>MRR <b>${{fmtM(mrrT)}}</b></span></div>
    <div class="io-leg-item"><span class="io-leg-dot" style="background:#06b6d4"></span><span>Seguro <b>${{fmtM(segT)}}</b></span></div>
    <div class="io-leg-item"><span class="io-leg-dot" style="background:#f97316"></span><span>Serviços <b>${{fmtM(srvT)}}</b></span></div>`;
}}
function renderVBarChart(elId,bins){{
  const el=document.getElementById(elId); if(!el)return;
  const maxCount=Math.max(...bins.map(b=>b.val),1);
  el.innerHTML=bins.map(b=>{{
    const h=b.val>0?Math.max(Math.round(b.val/maxCount*100),4):0;
    return`<div class="io-vbar">
      <div class="io-vbar-val">${{b.val}}</div>
      <div class="io-vbar-track"><div class="io-vbar-fill" style="height:${{h}}%;background:${{b.clr}}"></div></div>
      <div class="io-vbar-label">${{b.label}}</div>
    </div>`;
  }}).join('');
}}
function renderDistScore(data){{
  renderVBarChart('io-dist-score',[
    {{label:'0–20%',val:data.filter(d=>d.score!=null&&d.score<20).length,clr:'#EF4444'}},
    {{label:'20–40%',val:data.filter(d=>d.score!=null&&d.score>=20&&d.score<40).length,clr:'#F97316'}},
    {{label:'40–60%',val:data.filter(d=>d.score!=null&&d.score>=40&&d.score<60).length,clr:'#F59E0B'}},
    {{label:'60–80%',val:data.filter(d=>d.score!=null&&d.score>=60&&d.score<80).length,clr:'#84cc16'}},
    {{label:'80–100%',val:data.filter(d=>d.score!=null&&d.score>=80).length,clr:'#22C55E'}},
    {{label:'S/Dado',val:data.filter(d=>d.score==null).length,clr:'rgba(255,255,255,0.2)'}},
  ]);
}}
function renderFaixasMRR(data){{
  renderVBarChart('io-faixas-mrr',[
    {{label:'Sem dado',val:data.filter(d=>!d.mrr_acc_e).length,clr:'rgba(255,255,255,0.2)'}},
    {{label:'0–5K',val:data.filter(d=>d.mrr_acc_e>0&&d.mrr_acc_e<5000).length,clr:'#EF4444'}},
    {{label:'5–10K',val:data.filter(d=>d.mrr_acc_e>=5000&&d.mrr_acc_e<10000).length,clr:'#F59E0B'}},
    {{label:'+10K',val:data.filter(d=>d.mrr_acc_e>=10000).length,clr:'#22C55E'}},
  ]);
}}
function calcMesesConsultoria(data_inicio){{
  if(!data_inicio)return null;
  const pts=data_inicio.split('/');
  if(pts.length!==3)return null;
  const dt=new Date(+pts[2],+pts[1]-1,+pts[0]);
  if(isNaN(dt)||dt.getFullYear()<2000||dt.getFullYear()>2100)return null;
  return Math.max(0,Math.round((new Date()-dt)/(1000*60*60*24*30.44)));
}}
function renderTempoScore(data){{
  const bins=[
    {{label:'0–6m',min:0,max:6,clr:'#93c5fd'}},
    {{label:'6–12m',min:6,max:12,clr:'#60a5fa'}},
    {{label:'+12m',min:12,max:9999,clr:'#3b82f6'}},
    {{label:'Sem data',min:null,max:null,clr:'rgba(255,255,255,0.2)'}},
  ];
  const groups=bins.map(b=>{{
    const companies=b.min===null
      ?data.filter(d=>calcMesesConsultoria(d.data_inicio)===null)
      :data.filter(d=>{{const m=calcMesesConsultoria(d.data_inicio);return m!==null&&m>=b.min&&m<b.max;}});
    const scores=companies.filter(d=>d.score!=null).map(d=>d.score);
    const avg=scores.length?scores.reduce((a,b)=>a+b,0)/scores.length:0;
    return{{label:b.label+'<br><span style="font-size:9px;opacity:.55">'+companies.length+' emp</span>',val:Math.round(avg),clr:b.clr}};
  }});
  renderVBarChart('io-tempo-score',groups);
}}
function setMesIO(m,btn){{
  curMesIO=m; expandedIdIO=null;
  document.querySelectorAll('#pane-construindo .mfbtn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  renderAllIO();
}}
function sortByIO(col){{
  curSortIO=curSortIO.col===col?{{col,dir:curSortIO.dir==='asc'?'desc':'asc'}}:{{col,dir:'desc'}};
  expandedIdIO=null; renderIOTable();
}}
function toggleDetIO(id){{expandedIdIO=expandedIdIO===id?null:id;renderIOTable();}}
function eValsIO(d){{
  const m=curMesIO;
  function mvIO(bk,...kws){{
    const bloco=d.blocos?.[bk]; if(!bloco)return{{e:null,s:'nd',eff:null}};
    const ind=bloco.inds.find(i=>!i.sub&&kws.some(kw=>normStr(i.n).includes(kw)));
    if(!ind)return{{e:null,s:'nd',eff:null}};
    if(m===0)return{{e:ind.e,s:ind.s||'nd',eff:ind.eff}};
    const mv=ind.mensal?.[String(m)]; if(!mv)return{{e:null,s:'nd',eff:null}};
    return{{e:mv.e,s:mv.s||'nd',eff:mv.eff}};
  }}
  const mrrA=mvIO('acompanhamento','mrr acumulado');
  const mrrV=mvIO('acompanhamento','mrr vendido');
  const cart=mvIO('acompanhamento','clientes em carteira','total em carteira');
  const cAc=mvIO('acompanhamento','taxa de conversao');
  const cSrv=mvIO('servicos','taxa de conversao');
  const cSeg=mvIO('seguro','taxa de conversao');
  const _vaKs=[['acompanhamento','volume abordado'],['servicos','volume abordado'],['seguro','volume abordado'],['acompanhamento','safra de negocio']];
  const _vaE=_vaKs.map(([bk,kw])=>mvIO(bk,kw).e).filter(v=>v!=null);
  const va_e=_vaE.length?_vaE.reduce((a,b)=>a+b,0):null;
  const seg=mvIO('seguro','vendido','valor vendido');
  const srv=mvIO('servicos','valor vendido','vendido');
  const ind=mvIO('acompanhamento','volume indicacoes','indicacoes');
  return{{
    mrr_acc_e:mrrA.e, mrr_vend_e:mrrV.e, carteira_e:cart.e,
    conv_acomp_e:cAc.e!=null?+(cAc.e*100).toFixed(1):null,
    conv_srv_e:cSrv.e!=null?+(cSrv.e*100).toFixed(1):null,
    conv_seg_e:cSeg.e!=null?+(cSeg.e*100).toFixed(1):null,
    vol_abord_e:va_e, vol_abord_s:'nd',
    seguro_e:seg.e, servicos_e:srv.e, indicacoes_e:ind.e,
    agentes_e:d.agentes_e,
  }};
}}
function renderIOTable(){{
  const data=getIOData();
  const col=curSortIO.col,dir=curSortIO.dir;
  const mensalCols=new Set(['mrr_acc_e','mrr_vend_e','conv_acomp_e','conv_srv_e','conv_seg_e','vol_abord_e','seguro_e','servicos_e','indicacoes_e','agentes_e']);
  const sorted=[...data].sort((a,b)=>{{
    let va,vb;
    if(col==='rank'){{va=DADOS.indexOf(a);vb=DADOS.indexOf(b);}}
    else if(col==='nome'){{va=a.nome.toLowerCase();vb=b.nome.toLowerCase();}}
    else if(mensalCols.has(col)){{va=eValsIO(a)[col];vb=eValsIO(b)[col];}}
    else{{va=a[col];vb=b[col];}}
    if(va==null)va=dir==='asc'?Infinity:-Infinity;
    if(vb==null)vb=dir==='asc'?Infinity:-Infinity;
    return va<vb?(dir==='asc'?-1:1):va>vb?(dir==='asc'?1:-1):0;
  }});
  const mesTag=curMesIO?` <span style="font-size:10px;background:#1c3a5e;color:#93c5fd;border-radius:3px;padding:1px 5px;margin-left:4px">${{MESES[curMesIO]}}/26</span>`:'';
  const cnt=document.getElementById('io-tableCount');
  if(cnt)cnt.innerHTML=sorted.length+' empresa(s)'+mesTag;
  const naStyle=curMesIO?'color:rgba(255,255,255,0.18)':'';
  let html='';
  sorted.forEach(d=>{{
    const ev=eValsIO(d);
    const rank=DADOS.indexOf(d)+1;
    const isExp=expandedIdIO===d.id;
    const per=d.periodo?`${{MESES[d.periodo.inicio]||d.periodo.inicio}}${{d.periodo.inicio!==d.periodo.fim?'−'+MESES[d.periodo.fim]:''}}/26`:'—';
    const stag=(d.status_tag||'').toUpperCase();
    const tagHtml=stag==='CRÍTICO'||stag==='CRITICO'?'<span class="mtag-critico">CRÍTICO</span>':stag==='ALERTA'?'<span class="mtag-alerta">ALERTA</span>':'';
    const tempoHtml=(()=>{{
      if(!d.data_inicio)return'';
      const pts=d.data_inicio.split('/');
      if(pts.length!==3)return'';
      const dt=new Date(+pts[2],+pts[1]-1,+pts[0]);
      if(isNaN(dt)||dt.getFullYear()<2000||dt.getFullYear()>2100)return'';
      const meses=Math.max(0,Math.round((new Date()-dt)/(1000*60*60*24*30.44)));
      if(meses<1)return'';
      if(meses<12)return meses+(meses===1?' mês':' meses');
      const anos=Math.floor(meses/12),rm=meses%12;
      return anos+'a'+(rm?' '+rm+'m':'');
    }})();
    const subInfo=[d.id?'ID '+d.id:'',d.consultor||'',tempoHtml].filter(x=>x).join(' · ');
    html+=`<tr class="erow${{isExp?' expanded':''}}" onclick="toggleDetIO('${{d.id}}')">
      <td><span style="color:rgba(255,255,255,0.35);font-weight:700">${{rank}}</span></td>
      <td>
        <div style="display:flex;flex-direction:column;gap:1px">
          <span style="font-weight:600">${{d.nome}}${{tagHtml}}</span>
          <span style="color:rgba(255,255,255,0.35);font-size:11px">${{subInfo}}</span>
        </div>
      </td>
      <td>${{scoreBadge(d.score,d.status)}}</td>
      <td style="font-weight:500">${{ev.mrr_acc_e!=null?fmtM(ev.mrr_acc_e):'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
      <td style="font-weight:500">${{ev.mrr_vend_e!=null?fmtM(ev.mrr_vend_e):'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
      <td style="${{naStyle}}">${{fmtN(ev.carteira_e)}}</td>
      <td style="color:${{convColor(ev.conv_acomp_e)}};font-weight:600">${{ev.conv_acomp_e!=null?ev.conv_acomp_e.toFixed(1)+'%':'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
      <td style="color:${{convColor(ev.conv_srv_e)}};font-weight:600">${{ev.conv_srv_e!=null?ev.conv_srv_e.toFixed(1)+'%':'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
      <td style="color:${{convColor(ev.conv_seg_e)}};font-weight:600">${{ev.conv_seg_e!=null?ev.conv_seg_e.toFixed(1)+'%':'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
      <td style="color:${{sColor(ev.vol_abord_s)}}">${{ev.vol_abord_e!=null?fmtN(ev.vol_abord_e):'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
      <td style="font-weight:500">${{ev.seguro_e!=null?fmtM(ev.seguro_e):'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
      <td style="font-weight:500">${{ev.servicos_e!=null?fmtM(ev.servicos_e):'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
      <td style="color:rgba(255,255,255,0.5)">${{ev.indicacoes_e!=null?fmtN(ev.indicacoes_e):'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
      <td style="color:#7dd3fc;font-weight:500">${{ev.agentes_e!=null?fmtN(ev.agentes_e):'<span style="color:rgba(255,255,255,0.18)">—</span>'}}</td>
    </tr>`;
    if(isExp)html+=`<tr class="drow"><td colspan="14"><div class="din">${{renderDetail(d)}}</div></td></tr>`;
  }});
  if(!html)html='<tr><td colspan="14" class="nodata">Nenhuma empresa encontrada</td></tr>';
  const tb=document.getElementById('io-tableBody');
  if(tb)tb.innerHTML=html;
}}
function renderAgenteAbord(data){{
  const withAgents=data.filter(d=>(d.agentes_e||0)>0);
  const LIMITE=50;
  const vals=withAgents.map(d=>{{
    const ev=eVals(d); return{{d,v:ev.vol_abord_e}};
  }}).sort((a,b)=>(a.v||0)-(b.v||0));
  const maxVal=Math.max(...vals.map(x=>x.v||0),LIMITE*1.5,1);
  const el=document.getElementById('io-agente-abord'); if(!el)return;
  if(!vals.length){{el.innerHTML='<div class="nodata">Nenhuma empresa com Agente PV</div>';return;}}
  el.innerHTML=vals.map(x=>{{
    const abaixo=x.v!=null&&x.v<LIMITE;
    const clr=abaixo?'#EF4444':x.v!=null?'#22C55E':'rgba(255,255,255,0.2)';
    const pct=x.v!=null?Math.min((x.v/maxVal*100),100).toFixed(1):0;
    const thPct=(LIMITE/maxVal*100).toFixed(1);
    return`<div class="io-hbar">
      <div class="io-hbar-label" style="color:${{abaixo?'#EF4444':''}}" title="${{x.d.nome}}">${{x.d.nome.split(' ')[0]}}</div>
      <div class="io-hbar-track">
        <div class="io-hbar-fill" style="width:${{pct}}%;background:${{clr}}"></div>
        <div class="io-threshold" style="left:${{thPct}}%"></div>
      </div>
      <div class="io-hbar-val" style="color:${{abaixo?'#EF4444':''}}">${{x.v!=null?x.v:'—'}}</div>
    </div>`;
  }}).join('');
}}

// ── SYNC GITHUB ──────────────────────────────────────────────────────────────
function getSyncCfg(){{try{{return JSON.parse(localStorage.getItem('sz_sync')||'null');}}catch{{return null;}}}}
function setSyncCfg(c){{localStorage.setItem('sz_sync',JSON.stringify(c));}}
function openSyncModal(){{
  const c=getSyncCfg()||{{}};
  document.getElementById('syncOwner').value=c.owner||'';
  document.getElementById('syncRepo').value=c.repo||'';
  document.getElementById('syncPat').value=c.pat||'';
  document.getElementById('syncModal').style.display='flex';
}}
function closeSyncModal(){{document.getElementById('syncModal').style.display='none';}}
function saveSyncConfig(){{
  const owner=document.getElementById('syncOwner').value.trim();
  const repo=document.getElementById('syncRepo').value.trim();
  const pat=document.getElementById('syncPat').value.trim();
  if(!owner||!repo||!pat){{alert('Preencha todos os campos.');return;}}
  setSyncCfg({{owner,repo,pat}});
  closeSyncModal();
  triggerSync();
}}
async function triggerSync(){{
  const cfg=getSyncCfg();
  if(!cfg||!cfg.pat||!cfg.owner||!cfg.repo){{openSyncModal();return;}}
  const btn=document.getElementById('syncBtn');
  btn.disabled=true;btn.textContent='⟳ Sincronizando...';btn.style.color='';
  try{{
    const r=await fetch(
      `https://api.github.com/repos/${{cfg.owner}}/${{cfg.repo}}/actions/workflows/update-dashboard.yml/dispatches`,
      {{method:'POST',
        headers:{{'Authorization':'token '+cfg.pat,'Accept':'application/vnd.github.v3+json','Content-Type':'application/json'}},
        body:JSON.stringify({{ref:'main'}})}}
    );
    if(r.status===204){{
      btn.textContent='✓ Disparado! (~2 min)';btn.style.color='#22C55E';
      setTimeout(()=>{{btn.textContent='⟳ Sincronizar';btn.disabled=false;btn.style.color='';}}
      ,7000);
    }}else{{
      const e=await r.json().catch(()=>({{}}));
      throw new Error(e.message||'HTTP '+r.status);
    }}
  }}catch(e){{
    btn.textContent='✗ '+e.message;btn.style.color='#EF4444';
    setTimeout(()=>{{btn.textContent='⟳ Sincronizar';btn.disabled=false;btn.style.color='';}}
    ,6000);
  }}
}}

// ── INIT ──────────────────────────────────────────────────────────────────────
renderSummary();
renderInsights();
renderRankings1st();
renderGrowthCoverage();
updateFonteBanner(0);
renderTable();
</script>
</body>
</html>"""

# ── TV DASHBOARD ─────────────────────────────────────────────────────────────

def generate_tv_html(all_data, consolidado, insights):
    js_data=[]
    for d in all_data:
        blocos_js={}
        for bk,bloco in d['blocos'].items():
            inds=[{'n':i['n'],'sub':i['sub'],'e':i['e'],
                   'mensal':{str(k):{'e':v.get('e')} for k,v in i.get('mensal',{}).items()}}
                  for i in bloco['indicadores']]
            blocos_js[bk]={'inds':inds}
        js_data.append({'id':d['id'],'nome':d['nome'],'blocos':blocos_js})

    data_json=json.dumps(js_data,ensure_ascii=False,default=str)
    consol_json=json.dumps(consolidado,ensure_ascii=False,default=str)
    now=datetime.now().strftime('%d/%m/%Y às %H:%M')
    n=len(all_data)

    _logo_path=r'C:\Users\usuario\Desktop\sz-document-design\assets\solarz-logo-branca.png'
    try:
        with open(_logo_path,'rb') as _lf:
            _logo_b64=base64.b64encode(_lf.read()).decode('ascii')
        logo_html='<img src="data:image/png;base64,'+_logo_b64+'" height="32" alt="SolarZ">'
    except Exception:
        logo_html='<span style="color:#FF5500;font-weight:800;font-size:22px;letter-spacing:-0.02em">SolarZ</span>'

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<title>SolarZ TV · Planos de Negócio 2026</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
html,body{{width:1920px;height:1080px;overflow:hidden;font-family:'Inter',system-ui,sans-serif;color:#f0f0f0}}
body{{
  background:linear-gradient(135deg,#0a0f1e 0%,#0d1835 40%,#0a1628 70%,#0a0f1e 100%);
  display:grid;grid-template-rows:64px 196px 3px 1fr;
}}
@keyframes fadeInUp{{from{{opacity:0;transform:translateY(16px)}}to{{opacity:1;transform:translateY(0)}}}}
.anim{{animation:fadeInUp .5s ease both}}

/* HEADER */
header{{
  display:flex;align-items:center;padding:0 40px;gap:16px;
  background:rgba(255,255,255,0.03);
  border-bottom:1px solid rgba(255,255,255,0.07);
  backdrop-filter:blur(20px);-webkit-backdrop-filter:blur(20px);
}}
.hlogo{{display:flex;align-items:center;gap:12px}}
.hlogo img{{height:32px;filter:drop-shadow(0 0 8px rgba(255,85,0,0.35))}}
.hlogo-sep{{color:rgba(255,255,255,0.18);font-size:20px;font-weight:200;margin:0 4px}}
.hlogo-title{{font-size:15px;font-weight:600;color:rgba(255,255,255,0.65);letter-spacing:-0.01em}}
.hmeta{{margin-left:auto;display:flex;align-items:center;gap:20px}}
.hcount{{padding:4px 14px;border-radius:20px;background:rgba(245,158,11,0.1);border:1px solid rgba(245,158,11,0.25);color:#f59e0b;font-size:12px;font-weight:700;letter-spacing:0.06em;text-transform:uppercase}}
.htimestamp{{font-size:12px;color:rgba(255,255,255,0.3);font-weight:500}}

/* KPI SECTION */
.kpi-section{{
  padding:16px 40px;display:grid;grid-template-columns:repeat(6,1fr);gap:16px;align-items:stretch;
}}
.kpi-card{{
  background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:16px;
  padding:20px 22px;backdrop-filter:blur(20px);-webkit-backdrop-filter:blur(20px);
  display:flex;flex-direction:column;gap:10px;position:relative;overflow:hidden;
}}
.kpi-card::before{{
  content:'';position:absolute;top:0;left:0;right:0;height:2px;
  background:linear-gradient(90deg,transparent,var(--accent),transparent);opacity:.55;
}}
.kpi-head{{display:flex;align-items:center;gap:10px}}
.kpi-icon{{
  width:36px;height:36px;border-radius:10px;display:grid;place-items:center;flex-shrink:0;
  background:rgba(var(--accent-rgb),0.12);
}}
.kpi-icon svg{{width:18px;height:18px;stroke:var(--accent);stroke-width:2;fill:none;stroke-linecap:round;stroke-linejoin:round}}
.kpi-label{{font-size:10px;font-weight:700;letter-spacing:0.08em;text-transform:uppercase;color:rgba(255,255,255,0.45);line-height:1.3}}
.kpi-value{{font-size:30px;font-weight:800;letter-spacing:-0.03em;color:#fff;font-variant-numeric:tabular-nums;line-height:1}}
.kpi-sub{{font-size:10px;color:rgba(255,255,255,0.3);font-weight:500;margin-top:auto}}

/* DIVIDER */
.divider{{background:linear-gradient(90deg,transparent,rgba(255,255,255,0.07) 20%,rgba(255,255,255,0.07) 80%,transparent);flex-shrink:0}}

/* RANKINGS */
.rankings-section{{
  padding:20px 40px;display:grid;grid-template-columns:repeat(3,1fr);gap:24px;overflow:hidden;
}}
.rank-col{{
  background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:16px;
  backdrop-filter:blur(20px);-webkit-backdrop-filter:blur(20px);
  display:flex;flex-direction:column;overflow:hidden;
}}
.rank-col-header{{
  padding:20px 26px 16px;border-bottom:1px solid rgba(255,255,255,0.06);
  display:flex;align-items:center;gap:14px;flex-shrink:0;
}}
.rank-col-icon{{
  width:44px;height:44px;border-radius:12px;display:grid;place-items:center;flex-shrink:0;
  background:rgba(var(--col-rgb),0.12);
}}
.rank-col-title{{font-size:13px;font-weight:700;letter-spacing:0.04em;text-transform:uppercase;color:rgba(255,255,255,0.8)}}
.rank-col-sub{{font-size:11px;color:rgba(255,255,255,0.32);margin-top:3px}}
.rank-items{{flex:1;display:flex;flex-direction:column;overflow:hidden;min-height:0}}
.rank-item{{
  flex:1;min-height:0;padding:8px 22px;border-bottom:1px solid rgba(255,255,255,0.05);
  display:flex;flex-direction:column;gap:5px;
  transition:background .2s;
}}
.rank-item:last-child{{border-bottom:none}}
.rank-item:hover{{background:rgba(255,255,255,0.02)}}
.rank-top{{display:flex;align-items:center;gap:10px}}
.rank-badge{{
  width:26px;height:26px;border-radius:7px;display:grid;place-items:center;
  font-size:12px;font-weight:800;flex-shrink:0;
  border:1px solid var(--bd,rgba(255,255,255,.2));background:var(--bg,rgba(255,255,255,.06));color:var(--cl,rgba(255,255,255,.5));
}}
.rank-nome{{font-size:13px;font-weight:700;color:#fff;line-height:1.2;flex:1}}
.rank-value{{
  font-size:22px;font-weight:800;letter-spacing:-0.03em;
  color:var(--col-accent);font-variant-numeric:tabular-nums;line-height:1;padding-left:36px;
}}
.rank-bar-wrap{{padding-left:36px;display:flex;align-items:center;gap:8px}}
.rank-bar-bg{{flex:1;height:4px;background:rgba(255,255,255,0.07);border-radius:2px;overflow:hidden}}
.rank-bar-fill{{height:100%;border-radius:2px;transition:width .8s ease}}
.rank-bar-pct{{font-size:10px;color:rgba(255,255,255,0.35);font-weight:600;min-width:30px}}
.rank-empty{{flex:1;display:grid;place-items:center;color:rgba(255,255,255,0.22);font-size:13px}}
</style>
</head>
<body>

<header>
  <div class="hlogo">
    {logo_html}
    <span class="hlogo-sep">|</span>
    <span class="hlogo-title">Planos de Negócio · Pós-Vendas 2026</span>
  </div>
  <div class="hmeta">
    <span class="hcount">{n} empresas analisadas</span>
    <span class="htimestamp">Atualizado em {now}</span>
  </div>
</header>

<section class="kpi-section" id="kpiSection"></section>
<div class="divider"></div>
<section class="rankings-section" id="rankingsSection"></section>

<script>
const DADOS={data_json};
const CONSOL={consol_json};

function fmtM(v){{
  if(v==null)return'—';
  const a=Math.abs(v);
  if(a>=1e6)return'R$ '+(v/1e6).toFixed(1)+'M';
  if(a>=1e3)return'R$ '+(v/1e3).toFixed(1)+'K';
  return'R$ '+v.toFixed(0);
}}
function fmtN(v){{
  if(v==null)return'—';
  return Math.round(v).toLocaleString('pt-BR');
}}
function normStr(s){{
  return s.toLowerCase().normalize('NFD').replace(/[\\u0300-\\u036f]/g,'');
}}
function companyVal(d,bloco,...kws){{
  const bk=d.blocos?.[bloco];if(!bk)return null;
  const ind=bk.inds.find(i=>!i.sub&&kws.some(kw=>normStr(i.n).includes(kw)));
  if(!ind)return null;
  if(ind.mensal&&Object.keys(ind.mensal).length>0){{
    const vals=Object.values(ind.mensal).map(v=>v?.e).filter(v=>v!=null);
    if(vals.length>0)return vals.reduce((a,b)=>a+b,0);
  }}
  return ind.e??null;
}}
function topN(fn,n=3){{
  return DADOS.map(d=>({{nome:d.nome,val:fn(d)}}))
    .filter(x=>x.val!=null&&x.val>0)
    .sort((a,b)=>b.val-a.val).slice(0,n);
}}

const KPI_DEFS=[
  {{l:'MRR Total Construído',s:'MRR Acumulado Executado',v:()=>fmtM(CONSOL.total_mrr),
    ac:'#f59e0b',rgb:'245,158,11',
    ic:'<polyline points="23 6 13.5 15.5 8.5 10.5 1 18"/><polyline points="17 6 23 6 23 12"/>'}},
  {{l:'MRR Vendido',s:'Novas vendas no período',v:()=>fmtM(CONSOL.total_mrr_vend),
    ac:'#10b981',rgb:'16,185,129',
    ic:'<line x1="12" y1="1" x2="12" y2="23"/><path d="M17 5H9.5a3.5 3.5 0 0 0 0 7h5a3.5 3.5 0 0 1 0 7H6"/>'}},
  {{l:'Clientes Carteirizados',s:'Total gerenciado',v:()=>fmtN(CONSOL.total_carteira),
    ac:'#3b82f6',rgb:'59,130,246',
    ic:'<path d="M17 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"/><circle cx="9" cy="7" r="4"/><path d="M23 21v-2a4 4 0 0 0-3-3.87"/><path d="M16 3.13a4 4 0 0 1 0 7.75"/>'}},
  {{l:'Indicações Geradas',s:'Volume total de indicações',v:()=>fmtN(CONSOL.total_indicacoes),
    ac:'#a855f7',rgb:'168,85,247',
    ic:'<path d="M22 16.92v3a2 2 0 0 1-2.18 2 19.79 19.79 0 0 1-8.63-3.07A19.5 19.5 0 0 1 4.69 13 19.79 19.79 0 0 1 1.61 4.4 2 2 0 0 1 3.6 2.18h3a2 2 0 0 1 2 1.72c.127.96.361 1.903.7 2.81a2 2 0 0 1-.45 2.11L7.91 9a16 16 0 0 0 6.18 6.18l.95-.95a2 2 0 0 1 2.11-.45c.907.339 1.85.573 2.81.7A2 2 0 0 1 22 16.92z"/>'}},
  {{l:'Seguro Vendido',s:'Acumulado no período',v:()=>fmtM(CONSOL.total_seguro),
    ac:'#06b6d4',rgb:'6,182,212',
    ic:'<path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"/>'}},
  {{l:'Serviços Vendidos',s:'Acumulado no período',v:()=>fmtM(CONSOL.total_servicos),
    ac:'#f97316',rgb:'249,115,22',
    ic:'<circle cx="12" cy="12" r="3"/><path d="M19.07 4.93a10 10 0 0 1 0 14.14M4.93 4.93a10 10 0 0 0 0 14.14"/>'}}
];

const RANK_DEFS=[
  {{t:'Maior MRR Vendido',s:'Soma executada · todo o período',ac:'#f59e0b',rgb:'245,158,11',
    fn:d=>companyVal(d,'acompanhamento','mrr vendido'),fmt:fmtM,
    ic:'<polyline points="23 6 13.5 15.5 8.5 10.5 1 18"/><polyline points="17 6 23 6 23 12"/>'}},
  {{t:'Maior Seguro Vendido',s:'Soma executada · todo o período',ac:'#06b6d4',rgb:'6,182,212',
    fn:d=>companyVal(d,'seguro','vendido','valor vendido'),fmt:fmtM,
    ic:'<path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"/>'}},
  {{t:'Maior Serviços Vendidos',s:'Soma executada · todo o período',ac:'#f97316',rgb:'249,115,22',
    fn:d=>companyVal(d,'servicos','valor vendido','vendido'),fmt:fmtM,
    ic:'<circle cx="12" cy="12" r="3"/><path d="M19.07 4.93a10 10 0 0 1 0 14.14M4.93 4.93a10 10 0 0 0 0 14.14"/>'}}
];

const MEDALS=[
  {{bg:'rgba(245,158,11,.15)',bd:'rgba(245,158,11,.45)',cl:'#f59e0b'}},
  {{bg:'rgba(148,163,184,.15)',bd:'rgba(148,163,184,.45)',cl:'#94a3b8'}},
  {{bg:'rgba(180,83,9,.15)',bd:'rgba(180,83,9,.45)',cl:'#b45309'}}
];

function renderKPIs(){{
  const sec=document.getElementById('kpiSection');
  KPI_DEFS.forEach((k,i)=>{{
    const el=document.createElement('div');
    el.className='kpi-card anim';
    el.style.cssText=`--accent:${{k.ac}};--accent-rgb:${{k.rgb}};box-shadow:0 0 40px rgba(${{k.rgb}},.1);animation-delay:${{i*.07}}s`;
    el.innerHTML=`
      <div class="kpi-head">
        <div class="kpi-icon"><svg viewBox="0 0 24 24">${{k.ic}}</svg></div>
        <div class="kpi-label">${{k.l}}</div>
      </div>
      <div class="kpi-value">${{k.v()}}</div>
      <div class="kpi-sub">${{k.s}}</div>`;
    sec.appendChild(el);
  }});
}}

function renderRankings(){{
  const sec=document.getElementById('rankingsSection');
  RANK_DEFS.forEach((col,ci)=>{{
    const top=topN(col.fn,10);
    const maxVal=top.length>0?top[0].val:1;
    const el=document.createElement('div');
    el.className='rank-col anim';
    el.style.cssText=`--col-accent:${{col.ac}};--col-rgb:${{col.rgb}};box-shadow:0 0 50px rgba(${{col.rgb}},.06);animation-delay:${{.3+ci*.1}}s`;
    let h=`
      <div class="rank-col-header">
        <div class="rank-col-icon">
          <svg viewBox="0 0 24 24" width="20" height="20" stroke="${{col.ac}}" stroke-width="2" fill="none" stroke-linecap="round" stroke-linejoin="round">${{col.ic}}</svg>
        </div>
        <div>
          <div class="rank-col-title">${{col.t}}</div>
          <div class="rank-col-sub">${{col.s}}</div>
        </div>
      </div>
      <div class="rank-items">`;
    if(top.length===0){{
      h+=`<div class="rank-empty">Sem dados no período</div>`;
    }}else{{
      top.forEach((item,i)=>{{
        const pct=Math.round(item.val/maxVal*100);
        const m=i<3?MEDALS[i]:null;
        const badgeStyle=m
          ?`background:${{m.bg}};border-color:${{m.bd}};color:${{m.cl}}`
          :`background:rgba(255,255,255,.06);border-color:rgba(255,255,255,.15);color:rgba(255,255,255,.45)`;
        h+=`
          <div class="rank-item">
            <div class="rank-top">
              <div class="rank-badge" style="${{badgeStyle}}">${{i+1}}</div>
              <div class="rank-nome">${{item.nome}}</div>
            </div>
            <div class="rank-value">${{col.fmt(item.val)}}</div>
            <div class="rank-bar-wrap">
              <div class="rank-bar-bg"><div class="rank-bar-fill" style="width:${{pct}}%;background:${{col.ac}}"></div></div>
              <div class="rank-bar-pct">${{pct}}%</div>
            </div>
          </div>`;
      }});
    }}
    h+=`</div>`;
    el.innerHTML=h;
    sec.appendChild(el);
  }});
}}

renderKPIs();
renderRankings();
</script>
</body>
</html>"""

if __name__=='__main__':
    main()
